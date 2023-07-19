import os
from pathlib import Path
from openpyxl import load_workbook
import datetime
import time
from win32com import client
import pandas as pd
import shutil
import openpyxl as xl
from openpyxl import load_workbook
from win32com import client
import win32com.client


def copia ():
  excel_caminho = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Calculo_Mensal_VariosEstudos_Intersemanal.xlsx"
  caminho_arquivo_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Criacao_Estudos_intersemanal_auto.xlsm'
  #caminho do arquivo base onde estão registrados os casos
  aplicacao_excel = client.DispatchEx("Excel.Application")
  aplicacao_excel.Visible = True
  arquivo_excel = aplicacao_excel.Workbooks.Open(Filename =caminho_arquivo_base)
  formulario = arquivo_excel.Worksheets('formulario')


  caracter = excel_caminho.rfind('\\') + 1
  nome_arquivo = excel_caminho[caracter:]
  caminho_rodada = formulario.Cells(7, 14).Value #caminho da rede onde a rodada será armazenada

  shutil.copy2(Path(excel_caminho),
              Path(caminho_rodada))

def apagar_linhas(): 
 excel_caminho = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Calculo_Mensal_VariosEstudos_Intersemanal.xlsx"
 caminho =  r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario_v1.xlsx" 

 wb = xl.load_workbook(excel_caminho)
 wb_bp = xl.load_workbook(caminho)
 ws = wb['compila_cmo_medio']
 ws_ena = wb['compila_ena']
 ws_ea = wb['compila_ea']

 ws_bp = wb_bp['compila_cmo_medio']
 ws_bp_ena = wb_bp['compila_ena_ons'] 
 ws_bp_ea = wb_bp['compila_ea']

#ws_ea = wb.active
#ws.delete_rows(firstrow, numberofrows) #for multiple row deletion
 ws.delete_rows(1, ws.max_row+1) # for entire sheet
 ws_ena.delete_rows(1, ws_ena.max_row+1) # for entire sheet
 ws_ea.delete_rows(1, ws_ea.max_row+1) # for entire sheet
#ws.delete_rows(rownum) #for single row
 wb.save(str(excel_caminho)) 

### apaga as linhas do Base Prospec auxiliar 

 ws_bp.delete_rows(2, ws_bp.max_row+1) # for entire sheet
 ws_bp_ena.delete_rows(2, ws_bp_ena.max_row+1) # for entire sheet
 ws_bp_ea.delete_rows(2, ws_bp_ea.max_row+1) # for entire sheet
#ws.delete_rows(rownum) #for single row
 wb_bp.save(str(caminho)) 
 

def funcao_copia_cola(StudyId): 
     filename1 = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario.xlsx"
     wb1 = xl.load_workbook(filename1) #arquivo_base_prospec
     ws1 = wb1['compila_cmo_medio']
     ws2 = wb1['compila_ena'] 
     ws3 = wb1['compila_ea']  
     ws4 = wb1['compila_ea_inicial']
     ws5 = wb1['compila_ena_ree']
     ws6 = wb1['compila_gh']
     ws7 = wb1['compila_gt']

     mr = ws1.max_row 
     mc = ws1.max_column 
     linha = ws1.max_row + 1
     aux = 2
     
     for i in range (1, mr + 1): 
         if ws1.cell(row = i, column = 1).value == StudyId:  
           for j in range (2, 5):          
                c = ws1.cell(row = i, column = j)
                ws2.cell(row = aux, column = j).value = c.value
                ws3.cell(row = aux, column = j).value = c.value
                ws4.cell(row = aux, column = j).value = c.value
                ws5.cell(row = aux, column = j).value = c.value
                ws6.cell(row = aux, column = j).value = c.value 
                ws7.cell(row = aux, column = j).value = c.value  
           #aux +=1
        
     wb1.save(str(filename1)) 

def arquiva_pld ():

#Chama a funcao que apaga as linhas


 filename1 = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario_v1.xlsx"
 caminho_estudo = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Calculo_Mensal_VariosEstudos_Intersemanal.xlsx"
 caminho_decomp =  r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Base_PLD_decomp.xlsx'

 

## CMO_MEDIO - Preenchendo o CMO da Base_Prospec na Planilha do Calculo Mensal

 wb1 = xl.load_workbook(filename1) #arquivo_base_prospec
 ws1 = wb1['compila_cmo_medio'] 
 

 #arquivo_excel_rodada = xl.load_workbook(caminho_estudo, read_only=False, keep_vba=True) #calculo_mensal
 arquivo_excel_rodada = xl.load_workbook(caminho_estudo, read_only = False)
 ws2 = arquivo_excel_rodada['compila_cmo_medio'] 

 
 mr = ws1.max_row 
 mc = ws1.max_column 
 linha = ws1.max_row + 1
 aux = 1

 for i in range (1, mr + 1): 
  # if ws1.cell(row = i, column = 1).value == StudyId:  
     for j in range (1, mc + 1):          
          c = ws1.cell(row = i + 1, column = j)
          ws2.cell(row = i, column = j).value = c.value
     #aux +=1
      #j = j+1 
      #i = i+1  
 arquivo_excel_rodada.save(str(caminho_estudo))

## ENA - Preenchendo o ENA da Base_Prospec na Planilha do Calculo Mensal

 ws1 = wb1['compila_ena_ons'] 

 #arquivo_excel_rodada = xl.load_workbook(caminho_estudo, read_only=False, keep_vba=True) #calculo_mensal
 #ws2 = arquivo_excel_rodada.active
 ws2 = arquivo_excel_rodada['compila_ena'] 



 mr = ws1.max_row 
 mc = ws1.max_column 
 linha = ws1.max_row + 1

 aux_ena = 1

 for i in range (1, mr + 1): 
   #if ws1.cell(row = i, column = 1).value == StudyId:  
     for j in range (1, mc + 1):          
          c = ws1.cell(row = i + 1, column = j)
          ws2.cell(row = i, column = j).value = c.value
     #aux_ena +=1  
 arquivo_excel_rodada.save(str(caminho_estudo))  

### Energia Armazenada  - Preenchendo o EA da Base_Prospec na Planilha do Calculo Mensal

 ws1 = wb1['compila_ea'] 

 #arquivo_excel_rodada = xl.load_workbook(caminho_estudo, read_only=False, keep_vba=True) #calculo_mensal
 #ws2 = arquivo_excel_rodada.active
 ws2 = arquivo_excel_rodada['compila_ea'] 


 mr = ws1.max_row 
 mc = ws1.max_column 
 linha = ws1.max_row + 1

 aux_ea = 1

 for i in range (1, mr + 1): 
   #if ws1.cell(row = i, column = 1).value == StudyId:  
     for j in range (1, mc + 1):          
          c = ws1.cell(row = i + 1, column = j)
          ws2.cell(row = i, column = j).value = c.value
     #aux_ea +=1
 arquivo_excel_rodada.save(str(caminho_estudo)) 


  
 

##################################################################################################################################################
##################################################################################################################################################

def pld_realizado ():

 filename1 = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario.xlsx"
 caminho_estudo = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Calculo_Mensal_VariosEstudos_Intersemanal.xlsx"
 caminho_decomp =  r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Base_PLD_decomp.xlsx'  

 wb1 = xl.load_workbook(filename1) #arquivo_base_prospec
 ws1 = wb1['compila_cmo_medio'] 
 

#arquivo_excel_rodada = xl.load_workbook(caminho_estudo, read_only=False, keep_vba=True) #calculo_mensal
 arquivo_excel_rodada = xl.load_workbook(caminho_estudo, read_only = False)
 
#ws2 = arquivo_excel_rodada.active

 ws2 = arquivo_excel_rodada['compila_cmo_medio'] 
 ws2_cadastro = arquivo_excel_rodada['cadastro'] 
 ws2_pld_sudeste = arquivo_excel_rodada['Calculo_cmo_medio_SE'] 


## Preenchimento do PLD medio
###################################################################################
####### Para o Sudeste 


 base_decomp = xl.load_workbook(caminho_decomp, data_only = True) 
 ws3 = base_decomp.active
 ws3 = base_decomp['PLD_DECOMP_PATAMAR'] 
 aux = 1
#ax_ws3 = len(ws3(column = 12))
 aux_ws3 = max((c.row for c in ws3['L'] if c.value is not None))
 print(aux_ws3)
#aux_ws3 = 66

 mr = ws2_pld_sudeste.max_row 
 mc = ws2_pld_sudeste.max_column  
 print(ws3.cell(row = 66, column = 8).value)
#print(ws2_pld_sudeste.cell(row = 11, column = 1).value)
 print(mr)
 for i in range (11, mr+1): 
      if ws2_cadastro.cell(row = 3, column = 2).value == 1:  
                  c = ws3.cell(row = aux_ws3, column = 8).value
                  ws2_pld_sudeste.cell(row = aux, column = 6).value = c
      if ws2_cadastro.cell(row = 3, column = 2).value == 2:            
                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
                  c = ws3.cell(row = aux_ws3, column = 8).value
                  ws2_pld_sudeste.cell(row = i, column = 6).value = c_1
                  ws2_pld_sudeste.cell(row = i, column = 7).value = c

      if ws2_cadastro.cell(row = 3, column = 2).value == 3:             
                  c = ws3.cell(row = aux_ws3, column = 8).value
                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
                  c_2 = ws3.cell(row = aux_ws3 - 2, column = 8).value
                  ws2_pld_sudeste.cell(row = i, column = 6).value = c_2
                  ws2_pld_sudeste.cell(row = i, column = 7).value = c_1
                  ws2_pld_sudeste.cell(row = i, column = 8).value = c
            
      if ws2_cadastro.cell(row = 3, column = 2).value == 4:   
                  c = ws3.cell(row = aux_ws3, column = 8).value
                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
                  c_2 = ws3.cell(row = aux_ws3 - 2, column = 8).value
                  c_3 = ws3.cell(row = aux_ws3 - 3, column = 8).value
                  ws2_pld_sudeste.cell(row = i, column = 6).value = c_3
                  ws2_pld_sudeste.cell(row = i, column = 7).value = c_2
                  ws2_pld_sudeste.cell(row = i, column = 8).value = c_1
                  ws2_pld_sudeste.cell(row = i, column = 9).value = c
                  
                               
 arquivo_excel_rodada.save(str(caminho_estudo))


##############################################################################
########## Para o SUL


 ws2_pld_sul = arquivo_excel_rodada['Calculo_cmo_medio_S'] 
 base_decomp = xl.load_workbook(caminho_decomp, data_only = True) 
 ws3 = base_decomp.active
 ws3 = base_decomp['PLD_DECOMP_PATAMAR'] 
 aux = 1
#ax_ws3 = len(ws3(column = 12))
 aux_ws3 = max((c.row for c in ws3['L'] if c.value is not None))
 print(aux_ws3)
 
 mr = ws2_pld_sul.max_row 
 mc = ws2_pld_sul.max_column  
 #print(ws3.cell(row = aux_ws3, column = 8).value) 
 
 for i in range (11, mr+1): 
      if ws2_cadastro.cell(row = 3, column = 2).value == 1:  
                  c = ws3.cell(row = aux_ws3, column = 9).value
                  ws2_pld_sul.cell(row = i, column = 6).value = c
            #aux +=1
      if ws2_cadastro.cell(row = 3, column = 2).value == 2: 
         #for j in range (6, 8):          
                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 9).value
                  c = ws3.cell(row = aux_ws3, column = 9).value
                  ws2_pld_sul.cell(row = i, column = 6).value = c_1
                  ws2_pld_sul.cell(row = i, column = 7).value = c
                      
      if ws2_cadastro.cell(row = 3, column = 2).value == 3:            
                  c = ws3.cell(row = aux_ws3, column = 9).value
                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 9).value
                  c_2 = ws3.cell(row = aux_ws3 - 2, column = 9).value
                  ws2_pld_sul.cell(row = i, column = 6).value = c_2
                  ws2_pld_sul.cell(row = i, column = 7).value = c_1
                  ws2_pld_sul.cell(row = i, column = 8).value = c
                  
      if ws2_cadastro.cell(row = 3, column = 2).value == 4:           
                  c = ws3.cell(row = aux_ws3, column = 9).value
                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 9).value
                  c_2 = ws3.cell(row = aux_ws3 - 2, column = 9).value
                  c_3 = ws3.cell(row = aux_ws3 - 3, column = 9).value
                  ws2_pld_sul.cell(row = i, column = 6).value = c_3
                  ws2_pld_sul.cell(row = i, column = 7).value = c_2
                  ws2_pld_sul.cell(row = i, column = 8).value = c_1
                  ws2_pld_sul.cell(row = i, column = 9).value = c

 arquivo_excel_rodada.save(str(caminho_estudo))
############################################################################################################
############# Para o Norte


 ws2_pld_norte = arquivo_excel_rodada['Calculo_cmo_medio_N'] 
 base_decomp = xl.load_workbook(caminho_decomp, data_only = True) 
 
 mr = ws2_pld_norte.max_row 
 mc = ws2_pld_norte.max_column  

 for i in range (11, mr+1): 
      if ws2_cadastro.cell(row = 3, column = 2).value == 1:  
                  c = ws3.cell(row = aux_ws3, column = 11)
                  ws2_pld_norte.cell(row = i, column = 6).value = c.value
                  
                  
      if ws2_cadastro.cell(row = 3, column = 2).value == 2:            
                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 11).value
                  c = ws3.cell(row = aux_ws3, column = 11).value
                  ws2_pld_norte.cell(row = i, column = 6).value = c_1
                  ws2_pld_norte.cell(row = i, column = 7).value = c
             
      if ws2_cadastro.cell(row = 3, column = 2).value == 3:            
                  c = ws3.cell(row = aux_ws3, column = 11).value
                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 11).value
                  c_2 = ws3.cell(row = aux_ws3 - 2, column = 11).value
                  ws2_pld_norte.cell(row = i, column = 6).value = c_2
                  ws2_pld_norte.cell(row = i, column = 7).value = c_1
                  ws2_pld_norte.cell(row = i, column = 8).value = c
                  
      if ws2_cadastro.cell(row = 3, column = 2).value == 4:            
                  c = ws3.cell(row = aux_ws3, column = 11).value
                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 11).value
                  c_2 = ws3.cell(row = aux_ws3 - 2, column = 11).value
                  c_3 = ws3.cell(row = aux_ws3 - 3, column = 11).value
                  ws2_pld_norte.cell(row = i, column = 6).value = c_3
                  ws2_pld_norte.cell(row = i, column = 7).value = c_2
                  ws2_pld_norte.cell(row = i, column = 8).value = c_1
                  ws2_pld_norte.cell(row = i, column = 9).value = c

                  
 arquivo_excel_rodada.save(str(caminho_estudo))


#######################################################################################
############ Para o Nordeste




 ws2_pld_ne = arquivo_excel_rodada['Calculo_cmo_medio_NE']  
 base_decomp = xl.load_workbook(caminho_decomp, data_only = True) 

 mr = ws2_pld_ne.max_row 
 mc = ws2_pld_ne.max_column  

 for i in range (11, mr+1): 
      if ws2_cadastro.cell(row = 3, column = 2).value == 1:              
                  c = ws3.cell(row = aux_ws3, column = 10)
                  ws2_pld_ne.cell(row = i, column = 6).value = c.value
                  
      if ws2_cadastro.cell(row = 3, column = 2).value == 2:                 
                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 10).value
                  c = ws3.cell(row = aux_ws3, column = 10).value
                  ws2_pld_ne.cell(row = i, column = 6).value = c_1
                  ws2_pld_ne.cell(row = i, column = 7).value = c

      if ws2_cadastro.cell(row = 3, column = 2).value == 3:                        
                  c = ws3.cell(row = aux_ws3, column = 10).value
                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 10).value
                  c_2 = ws3.cell(row = aux_ws3 - 2, column = 10).value
                  ws2_pld_ne.cell(row = i, column = 6).value = c_2
                  ws2_pld_ne.cell(row = i, column = 7).value = c_1
                  ws2_pld_ne.cell(row = i, column = 8).value = c
                  
      if ws2_cadastro.cell(row = 3, column = 2).value == 4:                  
                  c = ws3.cell(row = aux_ws3, column = 10).value
                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 10).value
                  c_2 = ws3.cell(row = aux_ws3 - 2, column = 10).value
                  c_3 = ws3.cell(row = aux_ws3 - 3, column = 10).value
                  ws2_pld_ne.cell(row = i, column = 6).value = c_3
                  ws2_pld_ne.cell(row = i, column = 7).value = c_2
                  ws2_pld_ne.cell(row = i, column = 8).value = c_1
                  ws2_pld_ne.cell(row = i, column = 9).value = c           
                  
 arquivo_excel_rodada.save(str(caminho_estudo))


#################################################################################################
#################################################################################################

####################################################################################################################################################################
#extrair valor da aba cadastro do excel base prospec, preencher o excel calculo-mensal e preencher o base prospec com o valor de saída do base prospec
###################################################################################################################################################################

def salva_arquivo ():

 #filename1 = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario_v2.xlsx"
 #caminho_estudo = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Calculo_Mensal_VariosEstudos_Intersemanal.xlsx"
 #caminho_decomp =  r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Base_PLD_decomp.xlsx'  


# Opening and saving XLSX file, so results for each stored formula can be evaluated and cashed so OpenPyXL can read them.
 excel_file = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Calculo_Mensal_VariosEstudos_Intersemanal.xlsx"
 excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
 excel.DisplayAlerts = False # disabling prompts to overwrite existing file
 excel.Workbooks.Open(excel_file )
 excel.ActiveWorkbook.SaveAs(excel_file, FileFormat=51, ConflictResolution=2)
 excel.DisplayAlerts = True # enabling prompts
 excel.ActiveWorkbook.Close()

 wb = load_workbook(excel_file)


######################################################################################################
######################################################################################################

 
def arquiva_valor ():
####################################################################
#A partir desse ponto o codigo vai coletar e preencher o cmo - medio
####################################################################

 filename1 = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario_v2.xlsx"
 caminho_estudo = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Calculo_Mensal_VariosEstudos_Intersemanal.xlsx"
 caminho_decomp =  r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Base_PLD_decomp.xlsx'  


 # Opening and saving XLSX file, so results for each stored formula can be evaluated and cashed so OpenPyXL can read them.
 excel_file = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Calculo_Mensal_VariosEstudos_Intersemanal.xlsx"
 excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
 excel.DisplayAlerts = False # disabling prompts to overwrite existing file
 excel.Workbooks.Open(excel_file )
 excel.ActiveWorkbook.SaveAs(excel_file, FileFormat=51, ConflictResolution=2)
 excel.DisplayAlerts = True # enabling prompts
 excel.ActiveWorkbook.Close()



 # Aqui começa o código



 wb1 = xl.load_workbook(filename1) #arquivo_base_prospec
 ws1 = wb1['compila_cmo_medio'] 
 arquivo_excel_rodada = xl.load_workbook(caminho_estudo, read_only = False, data_only = True)

 
 ws1_1 = wb1.active
 ws1_1 = wb1['compila_cmo_medio'] #aba_cadastro


 ws2_2 = arquivo_excel_rodada.active
 ws2_2 = arquivo_excel_rodada['Add_cmo_medio'] 
 ws_cd = arquivo_excel_rodada['cadastro']

 mr_cmo = 41 #ws2_2.max_row
 mc_cmo = 8
 #linha = ws1_1.max_row + 1
 #print (linha) 
 aux_1 =  ws1_1.max_row 

 print(ws2_2.cell(row = 11, column = 1).value )
 #print(arquivo_excel_rodada['cadastro'].cell(row = 2, column = 2).value)

 aux_ws1 = max((i for i in range(7,30) if ws_cd.cell(row = i, column = 2).value != None))
 print(ws_cd.cell(row = 8, column = 2).value)
 linha = aux_ws1 - 6
 print(linha) 

 if arquivo_excel_rodada['cadastro'].cell(row = 2, column = 2).value == 1:
   for i in range (1, mr_cmo + 1): 
       #if ws2_2.cell(row = i, column = 1).value == StudyId:
           for j in range (1, mc_cmo + 1):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + aux_1, column = j).value = c 
           #aux_1 += 1
   wb1.save(str(filename1))
   
 if arquivo_excel_rodada['cadastro'].cell(row = 2, column = 2).value == 2:
   for i in range (1, linha + 1):
       #if ws2_2.cell(row = i, column = 1).value == StudyId:
           for j in range (1, mc_cmo + 1):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + aux_1, column = j).value = c 
   for i in range (1, linha + 1):            
           for j in range (11, 19):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + linha + aux_1, column = j-10).value = c
           #aux_1 += 1
   wb1.save(str(filename1))
   
 if arquivo_excel_rodada['cadastro'].cell(row = 2, column = 2).value == 3:
   for i in range (1, linha + 1):
       #if ws2_2.cell(row = i, column = 1).value == StudyId:
           for j in range (1, mc_cmo + 1):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + aux_1 , column = j).value = c 
   for i in range (1, linha + 1):            
           for j in range (11, 19):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + linha + aux_1, column = j-10).value = c
   for i in range (1, linha + 1):            
           for j in range (21, 19):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + linha + linha + aux_1, column = j-10).value = c            
           #aux_1 += 1
   wb1.save(str(filename1))
####################################################################
#A partir desse ponto o codigo vai coletar a ENA - media
####################################################################
  
 #arquivo_excel_rodada = xl.load_workbook(caminho_estudo, data_only = True) 
 wb1 = xl.load_workbook(filename1) #arquivo_base_prospec

 ws1_1 = wb1.active
 ws1_1 = wb1['compila_ena_ons'] #aba_cadastro


 ws2_2 = arquivo_excel_rodada.active
 ws2_2 = arquivo_excel_rodada['Add_ena'] 

 mr_cmo = 41 #ws2_2.max_row
 #print(mr_cmo)
 mc_cmo = 8 
 aux_1 = ws1_1.max_row 
 print(ws2_2.cell(row = 11, column = 1).value )
 #print(arquivo_excel_rodada['cadastro'].cell(row = 2, column = 2).value)

 aux_ws1 = max((i for i in range(7,30) if ws_cd.cell(row = i, column = 2).value != None))
 print(ws_cd.cell(row = 8, column = 2).value)
 linha = aux_ws1 - 6
 print(linha)

 if arquivo_excel_rodada['cadastro'].cell(row = 2, column = 2).value == 1:
   for i in range (1, mr_cmo + 1): 
       #if ws2_2.cell(row = i, column = 1).value == StudyId:
           for j in range (1, mc_cmo + 1):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + aux_1, column = j).value = c 
           #aux_1 += 1
   wb1.save(str(filename1))


 if arquivo_excel_rodada['cadastro'].cell(row = 2, column = 2).value == 2:
   for i in range (1, linha + 1):
           for j in range (1, mc_cmo + 1):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + aux_1, column = j).value = c 
   for i in range (1, linha + 1):            
           for j in range (11, 19):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + linha + aux_1, column = j-10).value = c
           #aux_1 += 1
   wb1.save(str(filename1))
   
 if arquivo_excel_rodada['cadastro'].cell(row = 2, column = 2).value == 3:
   for i in range (1, linha + 1):
       #if ws2_2.cell(row = i, column = 1).value == StudyId:
           for j in range (1, mc_cmo + 1):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + aux_1, column = j).value = c 
   for i in range (1, linha + 1):            
           for j in range (11, 19):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + linha + aux_1, column = j-10).value = c
   for i in range (1, linha + 1):            
           for j in range (21, 19):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + linha + linha + aux_1, column = j-10).value = c            
           #aux_1 += 1
   wb1.save(str(filename1))

####################################################################
#A partir desse ponto o codigo vai coletar a EA - media
####################################################################

 wb1 = xl.load_workbook(filename1) #arquivo_base_prospec
 ws1 = wb1['compila_ea'] 


 #arquivo_excel_rodada = xl.load_workbook(caminho_estudo, read_only=False, keep_vba=True) #calculo_mensal


 ws1_1 = wb1.active
 ws1_1 = wb1['compila_ea'] 

 ws2_2 = arquivo_excel_rodada.active
 ws2_2 = arquivo_excel_rodada['Add_ea'] 

 mr_cmo = 41 #ws2_2.max_row
 #print(mr_cmo)
 mc_cmo = 8
 aux_1 = ws1.max_row 
 print(ws2_2.cell(row = 11, column = 1).value )
 #print(arquivo_excel_rodada['cadastro'].cell(row = 2, column = 2).value) 

 aux_ws1 = max((i for i in range(7,30) if ws_cd.cell(row = i, column = 2).value != None))
 print(ws_cd.cell(row = 8, column = 2).value)
 linha = aux_ws1 - 6
 print(linha)

 
 #  print(arquivo_excel_rodada['cadastro'].cell(row = 2, column = 2).value)
 if arquivo_excel_rodada['cadastro'].cell(row = 2, column = 2).value == 1:
   for i in range (1, mr_cmo + 1): 
       #if ws2_2.cell(row = i, column = 1).value == StudyId:
           for j in range (1, mc_cmo + 1):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + aux_1, column = j).value = c 
           #aux_1 += 1
   wb1.save(str(filename1))
 if arquivo_excel_rodada['cadastro'].cell(row = 2, column = 2).value == 2:
   for i in range (1, linha + 1):
       #if ws2_2.cell(row = i, column = 1).value == StudyId:
           for j in range (1, mc_cmo + 1):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + aux_1 , column = j).value = c 
   for i in range (1, linha + 1):            
           for j in range (11, 19):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + linha + aux_1, column = j-10).value = c
           #aux_1 += 1
   wb1.save(str(filename1))
   
 if arquivo_excel_rodada['cadastro'].cell(row = 2, column = 2).value == 3:
   for i in range (1, linha + 1):
       #if ws2_2.cell(row = i, column = 1).value == StudyId:
           for j in range (1, mc_cmo + 1):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + aux_1, column = j).value = c 
   for i in range (1, linha + 1):            
           for j in range (11, 19):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + linha + aux_1, column = j-10).value = c
   for i in range (1, linha + 1):            
           for j in range (21, 19):      
               c = ws2_2.cell(row = i +10, column = j).value   
               ws1_1.cell(row = i + linha + linha + aux_1, column = j-10).value = c            
           #aux_1 += 1
   wb1.save(str(filename1))


####################################################################################################################################################################
############



def criar_copia ():
 filename1 = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario.xlsx"
 caminho_estudo = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Calculo_Mensal_VariosEstudos_Intersemanal.xlsx"
 caminho_decomp =  r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Base_PLD_decomp.xlsx'  

#####################################################################################################################################
#########  CRIANDO ARQUIVO TEMPORÁRIO

    
 caracter = caminho_estudo.rfind('\\') + 1
 nome_arquivo = caminho_estudo[caracter:]
 caminho_local = Path(r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\06.Semanal\temporarios')
 nome_arquivo_temp = nome_arquivo[:-5] + '_temp' + '.xlsx'
 #copia arquivo de histórico de vazões para uma pasta local
 print('Inciando cópia de arquivo original da rede no caminho: ', caminho_estudo)
 shutil.copy2(Path(caminho_estudo),
              Path.joinpath(caminho_local, nome_arquivo_temp))
 caminho_estudo_temp = Path.joinpath(Path(caminho_estudo), Path.joinpath(caminho_local, nome_arquivo_temp))
 print('Arquivo temporário copiado para pasta local: ', nome_arquivo_temp)
 time.sleep(3)
    

 caracter_base = caminho_decomp.rfind('\\') + 1
 nome_arquivo_base = caminho_decomp[caracter:]
 caminho_local_base = Path(r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\06.Semanal\temporarios')
 nome_arquivo_temp_base = nome_arquivo_base[:-5] + '_temp' + '.xlsx'
 #copia arquivo de histórico de vazões para uma pasta local
 print('Inciando cópia de arquivo original da rede no caminho: ', caminho_decomp)
 shutil.copy2(Path(caminho_decomp),
              Path.joinpath(caminho_local_base, nome_arquivo_temp_base))
 #caminho_local_base_temp = Path.joinpath(Path(filename1), Path.joinpath(caminho_local_base, nome_arquivo_temp_base))
 caminho_decomp_temp = Path.joinpath(caminho_local_base, nome_arquivo_temp_base)
 print('Arquivo temporário copiado para pasta local: ', nome_arquivo_temp_base)
 time.sleep(3)

#########################################################################################################################################
########################################################################################################################################

def retorna_arquivo ():
# Retornando os aqruivos temporarios para a rede
 caracter = caminho_estudo.rfind('\\') + 1
 nome_arquivo = caminho_estudo[caracter:]
 caminho_local = Path(r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\06.Semanal\temporarios')
 nome_arquivo_temp = nome_arquivo[:-5] + '_temp' + '.xlsx'
 #copia arquivo de histórico de vazões para uma pasta local
 print('Inciando cópia de arquivo original da rede no caminho: ', caminho_estudo)
 shutil.copy2(Path(caminho_estudo),
              Path.joinpath(caminho_local, nome_arquivo_temp))
 caminho_estudo_temp = Path.joinpath(Path(caminho_estudo), Path.joinpath(caminho_local, nome_arquivo_temp))
 print('Arquivo temporário copiado para pasta local: ', nome_arquivo_temp)
 time.sleep(3)
 caracter_base = caminho_decomp.rfind('\\') + 1
 nome_arquivo_base = caminho_decomp[caracter:]
 caminho_local_base = Path(r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\06.Semanal\temporarios')
 nome_arquivo_temp_base = nome_arquivo_base[:-5] + '_temp' + '.xlsx'
 shutil.copy2(Path.joinpath(caminho_local_base, nome_arquivo_temp), Path(caminho_estudo))
 shutil.copy2(Path.joinpath(caminho_local_base, nome_arquivo_temp_base), Path(caminho_decomp)) 
 print('Arquivo temporário copiado de volta para a rede') 

#########################################################################################################################################
########################################################################################################################################
# Preencher o Id de cada estudo na planilha calculo mensal

def StudyId (StudyId):
 filename = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Criacao_Estudos_intersemanal_auto.xlsm'
 caminho_estudo = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Calculo_Mensal_VariosEstudos_Intersemanal.xlsx"
 arquivo_excel_rodada = xl.load_workbook(caminho_estudo, read_only = False)
 wb1 = xl.load_workbook(filename, read_only = False, keep_vba= True, data_only = True)
 ws2 = arquivo_excel_rodada.active
 ws2 = arquivo_excel_rodada['cadastro']
 ws1 =wb1.active
 ws1 = wb1['formulario']

 mr = ws1.max_row
 mc = ws1.max_column


 for i in range (7, mr + 1):
    #if ws1.cell(row = i, column = 3).value == StudyId:
        for j in range (1, mc + 1):      
               c = ws1.cell(row = i, column = 3).value  
               ws2.cell(row = i , column = 2).value = c 
        
 arquivo_excel_rodada.save(str(caminho_estudo))
 


###########################################################################
###########################################################################
 
def ena_realizada (): 

 from openpyxl import load_workbook
 import openpyxl as xl
 from pathlib import Path
 import datetime
 import time
 import os

 #Im
 def atualizar_dados(data):

   m_ti = time.ctime(data) 
   t_obj = time.strptime(m_ti) 
   T_stamp = time.strftime("%d/%m/%Y %H:%M:%S", t_obj)
   #print(T_stamp) 
   return T_stamp
###############################DESCOBRINDO o arquivo da pasta RDH
 caminho_rdh =Path( r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\06.RDH\2022\2022-03")   #J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES
 lista_arquivos_rdh = os.listdir(caminho_rdh) 

 lista_datas_rdh = []
 for arquivo_rdh in lista_arquivos_rdh:
      # descobrir a data desse arquivo
      if ".xlsx" or ".xls" or ".xlsm" in arquivo_rdh:
            data_rdh = os.path.getmtime(f"{caminho_rdh}/{arquivo_rdh}")
            lista_datas_rdh.append((data_rdh, arquivo_rdh))
    
        # data inicial = 01/01/2021
        # data1 = 02/01/2021 -> 10.000
        # data2 = 15/02/2021 -> 150.000
    
 lista_datas_rdh.sort(reverse=True) 
 ultimo_arquivo_rdh = lista_datas_rdh[0] 
 data_rdh = lista_datas_rdh[0] 
#print(data_rdh[0]) 
#ultima_data_rdh = time.ctime(data_rdh[0]) 
 ultima_data_rdh = atualizar_dados(data_rdh[0])#.strftime('%d/%m/%Y') 
 ultima_data_rdh = datetime.datetime.strptime(ultima_data_rdh, '%d/%m/%Y %H:%M:%S') 
 ultima = atualizar_dados(data_rdh[0]) 
 ultima_data_rdh = ultima_data_rdh.strftime("%d/%m/%Y") 
 arquivo =  ultimo_arquivo_rdh[1][ultimo_arquivo_rdh[1].find('R'):18] 
 #print(arquivo)
 #print(ultimo_arquivo_rdh[1], ultima_data_rdh)  
 caminho_base_1 = Path.joinpath(caminho_rdh,arquivo) 
 caminho_base = Path.joinpath(caminho_base_1, '.xlsm') 
 caminho_base_rdh = str(caminho_base)
 caminho_base_rdh = caminho_base_rdh.replace('\.xlsm', '.xlsx') 
 #print(caminho_base_rdh) 
 base = xl.load_workbook(caminho_base_1, keep_vba= True, data_only = True) 
 ws3 = base.active 
 ws3 = base['Hidroenergética-Subsistemas'] 
 print(ws3.cell(row = 7, column = 3).value) 


 caminho = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Calculo_Mensal_VariosEstudos_Intersemanal.xlsx"



 arquivo_excel_rodada = xl.load_workbook(caminho, read_only = False) 
 ws2_ena_sudeste = arquivo_excel_rodada['Calculo_ena_SE']
 mr = ws2_ena_sudeste.max_row 
 mc = ws2_ena_sudeste.max_column  
 mc = ws2_ena_sudeste.max_column  
 ws2_cadastro = arquivo_excel_rodada['cadastro']
 
### Sudeste 

 s_4 = ws3.cell(row = 10, column = 4).value
 s_3 = ws3.cell(row = 10, column = 5).value
 s_2 = ws3.cell(row = 10, column = 6).value
 s_1 = ws3.cell(row = 10, column = 7).value 
 #print(s_1)

 for i in range (11, mr+1): 
      if ws2_cadastro.cell(row = 3, column = 2).value == 1:  
#            #for j in range (6, 7):
#                  c = ws3.cell(row = aux_ws3, column = 8).value
                  ws2_ena_sudeste.cell(row = i, column = 6).value = s_1
#            #aux +=1
      if ws2_cadastro.cell(row = 3, column = 2).value == 2:  
           # for j in range (6, 8):          
                  #c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
                  #c = ws3.cell(row = aux_ws3, column = 8).value
                  ws2_ena_sudeste.cell(row = i, column = 6).value = s_1
                  ws2_ena_sudeste.cell(row = i, column = 7).value = s_2
            #aux +=1

      if ws2_cadastro.cell(row = 3, column = 2).value == 3:   
            #for j in range (6, 9):          
                  #c = ws3.cell(row = aux_ws3, column = 8).value
                  #c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
                  #c_2 = ws3.cell(row = aux_ws3 - 2, column = 8).value
                  ws2_ena_sudeste.cell(row = i, column = 6).value = s_1
                  ws2_ena_sudeste.cell(row = i, column = 7).value = s_2
                  ws2_ena_sudeste.cell(row = i, column = 8).value = s_3
            #aux += 1 
            
      if ws2_cadastro.cell(row = 3, column = 2).value == 4:   
#           # for j in range (6, 10):
#                 c = ws3.cell(row = aux_ws3, column = 8).value
#                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
#                  c_2 = ws3.cell(row = aux_ws3 - 2, column = 8).value
#                  c_3 = ws3.cell(row = aux_ws3 - 3, column = 8).value
                  ws2_ena_sudeste.cell(row = i, column = 6).value = s_1
                  ws2_ena_sudeste.cell(row = i, column = 7).value = s_2
                  ws2_ena_sudeste.cell(row = i, column = 8).value = s_3
                  ws2_ena_sudeste.cell(row = i, column = 9).value = s_4
                  
                               
 arquivo_excel_rodada.save(str(caminho))

## Para o SUL

 ws2_ena_sul = arquivo_excel_rodada['Calculo_ena_S']
 mr = ws2_ena_sul.max_row 
 mc = ws2_ena_sul.max_column  
 mc = ws2_ena_sul.max_column 

 s_4 = ws3.cell(row = 18, column = 4).value
 s_3 = ws3.cell(row = 18, column = 5).value
 s_2 = ws3.cell(row = 18, column = 6).value
 s_1 = ws3.cell(row = 18, column = 7).value 
 print(s_1)

 for i in range (11, mr+1): 
      if ws2_cadastro.cell(row = 3, column = 2).value == 1:  
#            #for j in range (6, 7):
#                  c = ws3.cell(row = aux_ws3, column = 8).value
                  ws2_ena_sul.cell(row = i, column = 6).value = s_1
#            #aux +=1
      if ws2_cadastro.cell(row = 3, column = 2).value == 2:  
           # for j in range (6, 8):          
                  #c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
                  #c = ws3.cell(row = aux_ws3, column = 8).value
                  ws2_ena_sul.cell(row = i, column = 6).value = s_1
                  ws2_ena_sul.cell(row = i, column = 7).value = s_2
            #aux +=1

      if ws2_cadastro.cell(row = 3, column = 2).value == 3:   
            #for j in range (6, 9):          
                  #c = ws3.cell(row = aux_ws3, column = 8).value
                  #c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
                  #c_2 = ws3.cell(row = aux_ws3 - 2, column = 8).value
                  ws2_ena_sul.cell(row = i, column = 6).value = s_1
                  ws2_ena_sul.cell(row = i, column = 7).value = s_2
                  ws2_ena_sul.cell(row = i, column = 8).value = s_3
            #aux += 1 
            
      if ws2_cadastro.cell(row = 3, column = 2).value == 4:   
#           # for j in range (6, 10):
#                 c = ws3.cell(row = aux_ws3, column = 8).value
#                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
#                  c_2 = ws3.cell(row = aux_ws3 - 2, column = 8).value
#                  c_3 = ws3.cell(row = aux_ws3 - 3, column = 8).value
                  ws2_ena_sul.cell(row = i, column = 6).value = s_1
                  ws2_ena_sul.cell(row = i, column = 7).value = s_2
                  ws2_ena_sul.cell(row = i, column = 8).value = s_3
                  ws2_ena_sul.cell(row = i, column = 9).value = s_4
                  
                               
 arquivo_excel_rodada.save(str(caminho))

#### Para o Norte 

 ws2_ena_ne = arquivo_excel_rodada['Calculo_ena_NE']
 mr = ws2_ena_ne.max_row 
 mc = ws2_ena_ne.max_column  
 mc = ws2_ena_ne.max_column 

 s_4 = ws3.cell(row = 26, column = 4).value
 s_3 = ws3.cell(row = 26, column = 5).value
 s_2 = ws3.cell(row = 26, column = 6).value
 s_1 = ws3.cell(row = 26, column = 7).value 

 for i in range (11, mr+1): 
      if ws2_cadastro.cell(row = 3, column = 2).value == 1:  
#            #for j in range (6, 7):
#                  c = ws3.cell(row = aux_ws3, column = 8).value
                  ws2_ena_ne.cell(row = i, column = 6).value = s_1
#            #aux +=1
      if ws2_cadastro.cell(row = 3, column = 2).value == 2:  
           # for j in range (6, 8):          
                  #c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
                  #c = ws3.cell(row = aux_ws3, column = 8).value
                  ws2_ena_ne.cell(row = i, column = 6).value = s_1
                  ws2_ena_ne.cell(row = i, column = 7).value = s_2
            #aux +=1

      if ws2_cadastro.cell(row = 3, column = 2).value == 3:   
            #for j in range (6, 9):          
                  #c = ws3.cell(row = aux_ws3, column = 8).value
                  #c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
                  #c_2 = ws3.cell(row = aux_ws3 - 2, column = 8).value
                  ws2_ena_ne.cell(row = i, column = 6).value = s_1
                  ws2_ena_ne.cell(row = i, column = 7).value = s_2
                  ws2_ena_ne.cell(row = i, column = 8).value = s_3
            #aux += 1 
            
      if ws2_cadastro.cell(row = 3, column = 2).value == 4:   
#           # for j in range (6, 10):
#                 c = ws3.cell(row = aux_ws3, column = 8).value
#                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
#                  c_2 = ws3.cell(row = aux_ws3 - 2, column = 8).value
#                  c_3 = ws3.cell(row = aux_ws3 - 3, column = 8).value
                  ws2_ena_ne.cell(row = i, column = 6).value = s_1
                  ws2_ena_ne.cell(row = i, column = 7).value = s_2
                  ws2_ena_ne.cell(row = i, column = 8).value = s_3
                  ws2_ena_ne.cell(row = i, column = 9).value = s_4
                  
                               
 arquivo_excel_rodada.save(str(caminho))

#########  Para o Norte

## Preenchimento da ENA Realizada
###################################################################################
####### Para o Norte 

 ws2_ena_n = arquivo_excel_rodada['Calculo_ena_N']
 mr = ws2_ena_n.max_row 
 mc = ws2_ena_n.max_column  
 mc = ws2_ena_n.max_column  

 s_4 = ws3.cell(row = 34, column = 4).value
 s_3 = ws3.cell(row = 34, column = 5).value
 s_2 = ws3.cell(row = 34, column = 6).value
 s_1 = ws3.cell(row = 34, column = 7).value

 for i in range (11, mr+1): 
      if ws2_cadastro.cell(row = 3, column = 2).value == 1:  
#            #for j in range (6, 7):
#                  c = ws3.cell(row = aux_ws3, column = 8).value
                  ws2_ena_n.cell(row = i, column = 6).value = s_1
#            #aux +=1
      if ws2_cadastro.cell(row = 3, column = 2).value == 2:  
           # for j in range (6, 8):          
                  #c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
                  #c = ws3.cell(row = aux_ws3, column = 8).value
                  ws2_ena_n.cell(row = i, column = 6).value = s_1
                  ws2_ena_n.cell(row = i, column = 7).value = s_2
            #aux +=1

      if ws2_cadastro.cell(row = 3, column = 2).value == 3:   
            #for j in range (6, 9):          
                  #c = ws3.cell(row = aux_ws3, column = 8).value
                  #c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
                  #c_2 = ws3.cell(row = aux_ws3 - 2, column = 8).value
                  ws2_ena_n.cell(row = i, column = 6).value = s_1
                  ws2_ena_n.cell(row = i, column = 7).value = s_2
                  ws2_ena_n.cell(row = i, column = 8).value = s_3
            #aux += 1 
            
      if ws2_cadastro.cell(row = 3, column = 2).value == 4:   
#           # for j in range (6, 10):
#                 c = ws3.cell(row = aux_ws3, column = 8).value
#                  c_1 = ws3.cell(row = aux_ws3 - 1, column = 8).value
#                  c_2 = ws3.cell(row = aux_ws3 - 2, column = 8).value
#                  c_3 = ws3.cell(row = aux_ws3 - 3, column = 8).value
                  ws2_ena_n.cell(row = i, column = 6).value = s_1
                  ws2_ena_n.cell(row = i, column = 7).value = s_2
                  ws2_ena_n.cell(row = i, column = 8).value = s_3
                  ws2_ena_n.cell(row = i, column = 9).value = s_4
                  
                               
 arquivo_excel_rodada.save(str(caminho))

####################################################################################################################################################################
####################################################################################################################################################################

# A funcao salva uma copia do calculo mensal intersemanal no caminho da planilha criacao de estudo


def salvar_copia ():
 excel_caminho = r"J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Calculo_Mensal_VariosEstudos_Intersemanal.xlsx"
 caminho_arquivo_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Criacao_Estudos_intersemanal_auto.xlsm'
 #caminho do arquivo base onde estão registrados os casos
 aplicacao_excel = client.DispatchEx("Excel.Application")
 aplicacao_excel.Visible = True
 arquivo_excel = aplicacao_excel.Workbooks.Open(Filename =caminho_arquivo_base)
 formulario = arquivo_excel.Worksheets('formulario')
 caminho_rodada = formulario.Cells(7, 14).Value #caminho da rede onde a rodada será armazenada
 #print(caminho_rodada)
 shutil.copy2(Path(excel_caminho),
               Path(caminho_rodada))