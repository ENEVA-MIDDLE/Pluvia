from win32com import client
from openpyxl import load_workbook

caminho_arquivo_cria_estudo = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Testes\Criacao_Estudos_intersemanal_auto_teste.xlsm'
caminho_arquivo_ENA = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Testes\Base_ENA_Pluvia_teste.xlsx'

arquivo_excel = client.DispatchEx("Excel.Application")
wb_study = arquivo_excel.Workbooks.Open(Filename = caminho_arquivo_cria_estudo)
ws_config = wb_study.Worksheets('configuracoes')
wb_ena = load_workbook(caminho_arquivo_ENA)
ws_estat = wb_ena['estatisticas']

linha = 7
while ws_config.Cells(linha, 6).Value is not None:
    estatistica = ws_config.Cells(linha, 8).Value
    submercado = ws_config.Cells(linha, 9).Value
    periodo = ws_config.Cells(linha, 10).Value
    #semana = ws_config.Cells(linha, 11).Value
    mes_ref = ws_config.Cells(linha, 12).Value

    ano = str(mes_ref.year)
    mes = '%02d' % mes_ref.month
    ano_mes = ano + mes
    if estatistica == 'Menor':
        estat_ID = 1
    elif estatistica == 'P25':
        estat_ID = 2
    elif estatistica == 'Mediana':
        estat_ID = 3
    elif estatistica == 'P75':
        estat_ID = 4
    elif estatistica == 'Maior':
        estat_ID = 5

    linha_ENA = 3
    while ws_estat.cell(linha_ENA, 2).value is not None:
        if ws_estat.cell(linha_ENA, 2).value == ano_mes and ws_estat.cell(linha_ENA, 3).value == submercado:
            membro = ws_estat.cell(linha_ENA, 3 + estat_ID).value
            ws_config.Cells(linha, 13).NumberFormat = '@'
            ws_config.Cells(linha, 13).Value = membro
            print (membro)
        linha_ENA = linha_ENA + 1

    linha = linha + 1

wb_ena.close()
wb_study.Save()
wb_study.Close(False)
arquivo_excel.Application.Quit()

