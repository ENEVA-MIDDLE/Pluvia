from openpyxl import load_workbook
import datetime
from Funcoes_API_Pluvia import percentile

caminho_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\2020\11\16\05.ENA\Base_ENA_Pluvia.xlsx'

arquivo_excel = load_workbook(caminho_base)
arquivo_excel.create_sheet('estatisticas')
ws_base = arquivo_excel['pluvia_definitivo']
ws_estat = arquivo_excel['estatisticas']
ws_base.cell(1, 12).value = 'mês_ref'
ws_base.cell(1, 13).value = 'semana operativa'
ws_base.cell(1, 14).value = 'ano_mes_semana'
ws_base.cell(1, 15).value = 'ano_mes'
ws_estat.cell(1, 2).value = 'TABELA DE MEMBROS DE ACORDO COM PERCENTIL'
ws_estat.cell(1, 13).value = 'TABELA DE ENA (MWm) DE ACORDO COM PERCENTIL'
ws_estat.cell(2, 2).value = 'Período'
ws_estat.cell(2, 13).value = 'Período'
ws_estat.cell(2, 3).value = 'Submercado'
ws_estat.cell(2, 14).value = 'Submercado'
ws_estat.cell(2, 4).value = 'Minimo'
ws_estat.cell(2, 15).value = 'Minimo'
ws_estat.cell(2, 5).value = 'P25'
ws_estat.cell(2, 16).value = 'P25'
ws_estat.cell(2, 6).value = 'P50'
ws_estat.cell(2, 17).value = 'P50'
ws_estat.cell(2, 7).value = 'P75'
ws_estat.cell(2, 18).value = 'P75'
ws_estat.cell(2, 8).value = 'Máximo'
ws_estat.cell(2, 19).value = 'Máximo'

linha = 2

while ws_base.cell(linha,1).value is not None:

    mes = str(ws_base.cell(linha, 8).value[6:8])
    ano = str(ws_base.cell(linha, 8).value[2:6])
    sem_op = 'RV' + str((int(ws_base.cell(linha, 8).value[12:13]) - 1))

    ws_base.cell(linha, 12).value = datetime.datetime(int(ano), int(mes), 1)
    ws_base.cell(linha, 12).number_format = 'DD/MM/YYYY'
    ws_base.cell(linha, 13).value = sem_op
    ws_base.cell(linha, 14).value = ano + mes + sem_op[-1:]
    ws_base.cell(linha, 15).value = ano + mes

    linha = linha + 1
arquivo_excel.save(caminho_base)
#linha = 2
#lista_sin = []
submercados = ['SIN', 'SUDESTE', 'SUL', 'NORDESTE', 'NORTE']
lista_periodos = []
lista_meses = []
lista_membros = []
percent_list = [0, 25, 50, 75, 100]
linha = 2
while ws_base.cell(linha, 14).value is not None:
    lista_membros.append(ws_base.cell(linha, 4).value)
    lista_periodos.append(ws_base.cell(linha, 14).value)
    lista_meses.append(ws_base.cell(linha, 15).value)
    linha = linha + 1

lista_membros = list(set(lista_membros))
val_remove = ['00', 'ENSEMBLE']
for item in val_remove:
    if item in lista_membros:
        lista_membros.remove(item)
#lista_membros.remove('ENSEMBLE')
lista_periodos = list(set(lista_periodos))
lista_meses = list(set(lista_meses))
lista_membros.sort()
lista_periodos.sort()
lista_meses.sort()

for submercado in submercados:
    for periodo in lista_periodos:
        linha = 2
        lista_sin = []
        while ws_base.cell(linha, 6).value is not None:
            if ws_base.cell(linha, 6).value == 'Submercado' and ws_base.cell(linha, 7).value == submercado and ws_base.cell(linha, 14).value == periodo and ws_base.cell(linha, 4).value != '00' and ws_base.cell(linha, 4).value != 'ENSEMBLE':
                lista_sin.append((ws_base.cell(linha, 4).value, float(str(ws_base.cell(linha, 11).value).replace(',','.'))))
            linha = linha + 1

        pri_lin_vaz = len(ws_estat['B']) + 1
        ws_estat.cell(pri_lin_vaz, 2).value = periodo
        ws_estat.cell(pri_lin_vaz, 3).value = submercado
        ws_estat.cell(pri_lin_vaz, 13).value = periodo
        ws_estat.cell(pri_lin_vaz, 14).value = submercado
        for p in range(len(percent_list)):
            resultado = percentile(lista_sin, percent_list[p])
            ws_estat.cell(pri_lin_vaz, p + 4).value = resultado[0]
            ws_estat.cell(pri_lin_vaz, p + 15).value = resultado[1]
            ws_estat.cell(pri_lin_vaz, p + 15).number_format = '#,##0.00'

for submercado in submercados:
    for mes in lista_meses:
        lista_sin = []
        for membro in lista_membros:
            dados_membro_mes = []
            linha = 2
            while ws_base.cell(linha, 6).value is not None:
                if ws_base.cell(linha, 6).value == 'Submercado' and ws_base.cell(linha, 7).value == submercado and ws_base.cell(linha, 15).value == mes and ws_base.cell(linha, 4).value == membro and ws_base.cell(linha, 4).value != '00' and ws_base.cell(linha, 4).value != 'ENSEMBLE':
                    dados_membro_mes.append(float(str(ws_base.cell(linha, 11).value).replace(',','.')))

                linha = linha + 1
            media = sum(dados_membro_mes)/ len(dados_membro_mes)
            lista_sin.append((membro, media))
        pri_lin_vaz = len(ws_estat['B']) + 1
        ws_estat.cell(pri_lin_vaz, 2).value = mes
        ws_estat.cell(pri_lin_vaz, 3).value = submercado
        ws_estat.cell(pri_lin_vaz, 13).value = mes
        ws_estat.cell(pri_lin_vaz, 14).value = submercado
        #p = ''
        for p in range(len(percent_list)):
            resultado = percentile(lista_sin, percent_list[p])
            ws_estat.cell(pri_lin_vaz, p + 4).value = resultado[0]
            ws_estat.cell(pri_lin_vaz, p + 15).value = resultado[1]
            ws_estat.cell(pri_lin_vaz, p + 15).number_format = '#,##0.00'

arquivo_excel.save(caminho_base)
arquivo_excel.close()