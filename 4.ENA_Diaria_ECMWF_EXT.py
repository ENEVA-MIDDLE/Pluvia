from openpyxl import load_workbook
import datetime
from Funcoes_API_Pluvia import getForecasts
from Funcoes_API_Pluvia import authenticatePluvia
from Funcoes_API_Pluvia import downloadForecast
from Funcoes_API_Pluvia import cria_pasta_local
from pathlib import Path
import os
import shutil
import warnings
from zipfile import ZipFile
from Funcoes_API_Pluvia import le_ena_pasta
from Funcoes_API_Pluvia import salva_ENA_base
from dateutil.relativedelta import relativedelta
import json
warnings.filterwarnings('ignore')
base_pluvia = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\11.HIDROLOGIA\01.PLUVIA\API_Pluvia\Dados_API_Pluvia.xlsx' #arquivo que configura o download das informçãoes diárias
pathResult = Path(r'C:\Users\fernando.fidalgo\OneDrive - Eneva S.A\03. Eneva\14. Comercializadora\05. Update_ONS\01.Pluvia') #caminho de download local dos arquivos
caminho_rede = Path(r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\11.HIDROLOGIA\01.PLUVIA') # caminho raiz do pluvia
authenticatePluvia () #autentica
excel = load_workbook(filename= base_pluvia, data_only=True)
ws = excel['ENA_diaria']
ws_aux = excel['aux']

hoje = datetime.datetime.today().strftime('%d/%m/%Y') #datetime.datetime(2020, 10, 19)#datetime.datetime.today().strftime('%d/%m/%Y')


print('------------------------------------------------------------------------------------------------------')
print('iniciando download de projeções para o dia', hoje)
linha = 8

while ws.cell(linha, 4).value != None and ws.cell(linha, 4 ).value != '':
    download_arquivo = False
    if ws.cell(linha, 4).value == 'ECMWF_ENS_EXT':
        forecastDate = hoje#.strftime('%d/%m/%Y')
        x = 4
        while ws_aux.cell(x,7).value != None:
            if ws_aux.cell(x,7).value == ws.cell(linha, 4 ).value:
                precipitationDataSources = []
                precipitationDataSources.append(ws_aux.cell(x,6).value)
            x = x + 1
        #precipitationDataSources = [12]
        x = 4
        while ws_aux.cell(x, 2).value != None:
            if ws_aux.cell(x, 2).value == ws.cell(linha, 5).value:
                forecastModels = []
                forecastModels.append(ws_aux.cell(x, 1).value)
            x = x + 1
        #forecastModels = [2]
        bias = ws.cell(linha, 6).value  # True / False
        if ws.cell(linha, 7).value == 'Definitivo':
            preliminary = 'False'
        elif ws.cell(linha, 7).value == 'Preliminar':
            preliminary = 'True'
        years = [int(ws.cell(linha, 8).value)]
        members = ''
        if ws.cell(linha, 9).value == None:
            members = ''
            #print('membro_none')
        else:
            members = [ws.cell(linha, 9).value]

        forecasts = getForecasts(forecastDate, precipitationDataSources, forecastModels, bias, preliminary, years,
                                 members)
        if forecasts == []:
            print('Sem previsões para o Mapa ', ws.cell(linha, 4).value)
        else:
            for forecast in forecasts:

                ano = forecastDate[6:]
                mes = forecastDate[3:5]
                dia = forecastDate[:2]
                if forecast['preliminar'] == False:
                    prelim = 'definitiva'
                elif forecast['preliminar'] == True:
                    prelim = 'preliminar'
                else:
                    prelim = 'ERRO_PRELIM'
                    print('erro de código')

                nome_ENA = forecast['nome'] + '-' + forecast['membro'] + '-' + prelim + '-' + ano + mes + dia + '-ENA.zip'
                pathForecastDay = pathResult.joinpath(ano + '-' + mes + '-' + dia)
                cria_pasta_local(pathForecastDay)
                if forecast['enaDisponivel']:
                    downloadForecast(forecast['enaId'], pathForecastDay, nome_ENA)
                    download_arquivo = True
                else:
                    print('ENA não disponível para o seguinte forecast-->', 'Mapa:', forecast['mapa'], ') - Modelo:',forecast['modelo'])

            print('Mapa:', forecast['mapa'], '- Modelo:',forecast['modelo'], '- VNA disponívels:', forecast['vnaDisponivel'], '- ENA disponível:', forecast['enaDisponivel'],
                  '- STR disponível:', forecast['strDisponivel'], '- PREVS disponível:', forecast['prevsDisponivel'])

    linha = linha + 1

excel.close()
if download_arquivo: #só verifica a pasta se houve download de arquivos
    for arquivo in os.listdir(pathForecastDay): #lê os arquivos que estão na pasta de download
        if arquivo.upper().endswith('ENA.ZIP'): #exibe somente os arquvos de ENA
            caminho_arquivo = os.path.realpath(pathForecastDay) + '\\' + arquivo
            caminho_pasta = caminho_arquivo.replace('-ENA.zip', '')
            with ZipFile(caminho_arquivo, 'r') as zipObj: #descompacta o zip na mesma pasta do zip
                zipObj.extractall(caminho_pasta)
                print('arquivo descompactado:', arquivo)
            tabela = le_ena_pasta (caminho_pasta) #função que lê o arquivo CSV de ENA dentro da pasta e retorna uma lista onde cada item da lista é uma linha, e cada linha é uma nova lista, onde cada item é uma coluna
            salva_ENA_base (tabela) #função que lê a tabela gerada do arquivo de ENA lido e salva na base em excel

            os.makedirs(Path.joinpath(caminho_rede, ano, ano + '-' + mes, dia), exist_ok=True) #cria pasta na rede caso não exista
            shutil.move(Path.joinpath(Path(pathForecastDay), arquivo), #move os arquivos ZIP para a rede
                        Path.joinpath(caminho_rede, ano, ano + '-' + mes, dia, arquivo))
            print('Arquivo movido para a rede:', arquivo)
            shutil.rmtree(caminho_pasta)
            print('Pasta local deletada: ', caminho_pasta)
