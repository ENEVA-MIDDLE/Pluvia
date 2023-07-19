import pandas as pd
from datetime import datetime, timedelta
import numpy as np
import openpyxl 
from dateutil.relativedelta import relativedelta,FR
from openpyxl import load_workbook
import psycopg2
import shutil
#db-dtypes

# Script acessa tabela de 'marcacao' do Banco de dados Postgres, faz tratamento de dados e coloca dados do preço de mercado mes a mes A0 e A1 em excel (no formato adequado para o Dashboard Semanal)

#alterar caminho do json do Bigquery/Postegres = connection 
#alterar caminho do onedrive = path_onedrive

#############################################################################
compila_cmo = pd.DataFrame()
data_marcacao_cadastroExcel=[]

# Pegar data da ultima sexta, se for sexta pega a sexta da semana passada
if datetime.today().weekday()==4:
    lastFriday = datetime.now() + relativedelta(weekday=FR(-2))
else:
    lastFriday = datetime.now() + relativedelta(weekday=FR(-1))


lastFriday.replace(hour=0,minute=0,second=0,microsecond=0)

###################################################################
#Exportando dados Postgres em formato dataframe
connection = psycopg2.connect(user="focus.energia",
                              password="Ppn2Vks4bZeTgCXAqyYRqTDFi",
                              host="postgres.focusenergia.com.br",
                              port="5432",
                              database="DadosMercado")#"postgres")
cursor = connection.cursor()
cursor.execute('''select * from public.marcacao_mercado''')
data=cursor.fetchall()

#nome das colunas
cols=[]
for elt in cursor.description:
    cols.append(elt[0])

#dataframe da base do postgres
df_results=pd.DataFrame(data=data, columns=cols)


#deletando colunas que nao tem no BigQuery
#tipo_preco='Fixo'
#data_criacao
#data_atualizacao
tipopreco = ['Fixo']
df_results=df_results[df_results.tipo_preco.isin(tipopreco)]
df_results=df_results.drop(['data_criacao', 'data_atualizacao','tipo_preco'], axis=1)
#df_results.to_excel('MarksData.xlsx')

#############################################################################################
#Aplicando filtros no dataframe exportado do Postgres
#Preco SE
#Convencional
#data_marcacao apenas dos últimos dias
#Produtos necessarios para curva de mercado A0 e A1 mes a mes


#Filtrando os ultimos 7 dias
data_fim = max(df_results['data_marcacao'])
data_inic = data_fim - timedelta(days=9)

df=df_results[(df_results['data_marcacao'] > data_inic) & (df_results['data_marcacao'] < data_fim + timedelta(days=1))]

#Filtrando Submercado
submercado = ['SE/CO']
df=df[df.submercado.isin(submercado)]
#Filtrando Produto Convencional
fonte=['Convencional']
df=df[df.fonte.isin(fonte)]

#Filtrando Produto
fonte=['A0','H1/A0','H2/A0','Q1/A0','Q2/A0','Q3/A0','Q4/A0','M0','M1','M2','M3','M4','M5','M6','M7','M8','M9','M10','M11','M12','M13','M14','M15','M16','M17','M18','M19','M20','M21','M22','M23','A1','H1/A1','H2/A1','Q1/A1','Q2/A1','Q3/A1','Q4,A1']
df=df[df.produto.isin(fonte)]

#ordenar
df=df.sort_values(['data_marcacao', 'inicio_suprimento'])

# exporta para excel
#df['inserido_em'] = df['inserido_em'].dt.tz_localize(None) #necessário alteracao de formato para salvar excel
#df.to_excel('MarksData.xlsx')

dff=df # dff se mantem a base inteira

#####################################################################
#Definicao das datas de marcacao que serao usada para exibir no dash
#D : ultima data_marcacao disponivel na base
#D-1: penultima data_marcacao disponivel na base
#D-7: D-7-> foi necessario tratamento para caso de D-7 ser um feriado

#D-1 -> considera a segunda maior data da base de dados
data_fim2=dff['data_marcacao'].unique()[len(dff['data_marcacao'].unique())-2] # penultima data na base
#data_fim2=data_fim2.astype('datetime64[s]').tolist()#formato datetime


#D-7 -> Pega pela data D-7 mesmo
#tratamento feito para evitar problema de feriado, se D-7 nao tiver dado pega D-8 ...
if data_fim - timedelta(days=7) in dff['data_marcacao'].unique(): # procura uma semana atras
    data_fim3=data_fim - timedelta(days=7)
elif data_fim - timedelta(days=8) in dff['data_marcacao'].unique(): # se nao acha pq foi fds ou feriado pega o dia anterior
    data_fim3=data_fim - timedelta(days=8)
elif data_fim - timedelta(days=9) in dff['data_marcacao'].unique():
    data_fim3=data_fim - timedelta(days=9)
elif data_fim - timedelta(days=10) in dff['data_marcacao'].unique():
    data_fim3=data_fim - timedelta(days=10)
elif data_fim - timedelta(days=11) in dff['data_marcacao'].unique():
    data_fim3=data_fim - timedelta(days=11)

######################################################################################
#For para montar base de preços no formato necessario
curva_preco=[data_fim,data_fim2, data_fim3]
id_curva_preco=[1,2,3]
aux=0

for m in curva_preco:
    print(aux)
    df=dff[(dff['data_marcacao'] == m)]
    df['MES_inicio_suprimento']=pd.DatetimeIndex(df['inicio_suprimento']).month
    df['MES_fim_suprimento']=pd.DatetimeIndex(df['fim_suprimento']).month
    df['#MESES_inicio_fim_suprimento']=df['MES_fim_suprimento']-df['MES_inicio_suprimento']
    df=df.sort_values('#MESES_inicio_fim_suprimento', ascending=False) # ordem do maior para o menor
    df=df.reset_index() 



    ###########################################################################
    #Defini qual o primeiro mês de suprimento com base na tabela marcacao do BigQuery coluna 'produto'= 'M0'
    #o ultimo mes do suprimento será dez do ano seguinte
    m0=df[(df['produto'] == 'M0')]
    m0=m0.reset_index()
    produto_data_inic= m0.at[0,'inicio_suprimento']
    produto_ano_fim=produto_data_inic.year+1

    #Cria nova base com todos os meses de suprimento
    date_range = pd.DataFrame({'date': pd.date_range(produto_data_inic, datetime(produto_ano_fim,12,1), freq='MS')})
    date_range['MERCADO']=0
    date_range['suprimento']=date_range['date']
    date_range=date_range.set_index('date')

    ###########################################################################
    #Preenchendo nova base - considera valor ano --> semestre --> trimestre --> mês
    for i in range(0,len(df)): #para cada linha do df
        if df.at[i,'#MESES_inicio_fim_suprimento']==0:
            dte=df.at[i,'inicio_suprimento']
            date_range.at[dte.strftime('%Y-%m-%d'), 'MERCADO'] = df.at[i,'preco_ask']
        else:
            dte=df.at[i,'inicio_suprimento']
            for j in range(0,df.at[i,'#MESES_inicio_fim_suprimento']+1): #para cada periodo de suprimento
                dte2=df.at[i,'inicio_suprimento']+relativedelta(months=+j)
                date_range.at[dte2.strftime('%Y-%m-%d'), 'MERCADO'] = df.at[i,'preco_ask']

    #deleta linha com data suprimento NA
    date_range=date_range[date_range.suprimento.notnull()]

    #Cria colunas necessarias na planilha em excel
    date_range['mes_suprimento']=pd.DatetimeIndex(date_range['suprimento']).month
    date_range['ano_suprimento']=pd.DatetimeIndex(date_range['suprimento']).year
    date_range['ID Estudo']= id_curva_preco[aux]
    date_range['Sensibilidade']='Original'
    date_range['Deck']='DC'+ date_range['ano_suprimento'].astype(str).str.zfill(4)+date_range['mes_suprimento'].astype(str).str.zfill(2)+'-sem1_s1'
    date_range['MEN=0-SEM=1']='0'
    date_range['SUDESTE']=date_range['MERCADO']
    date_range['SUL']= float("NAN")
    date_range['NORDESTE']=float("NAN")
    date_range['NORTE']=float("NAN")

    #Cria dataframe apenas com as colunas necessaria na planilha
    date_range2 = date_range[['ID Estudo', 'Sensibilidade', 'Deck','MEN=0-SEM=1','SUDESTE','SUL','NORDESTE','NORTE']].copy()
    date_range2['SUDESTE'] =date_range2['SUDESTE'].astype(str) # altera tipo pq o formato do excel é esse

   #Concatena e empilha resultados das diferentes datas de marcacao
    compila_cmo=pd.concat([compila_cmo,date_range2])
    data_marcacao_cadastroExcel.append(m)
    aux=aux+1



###########################################################
#Preenche a planilha do excel
path=r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_PrecoMercado.xlsx'
book = load_workbook(path)
#remove sheet compila_cmo_medio
if 'compila_cmo_medio' in book.sheetnames:
    book.remove(book['compila_cmo_medio'])
#Altera dados da sheet 'Cadastro_Estudos'
ws = book.worksheets[0]
ws['B2'] = lastFriday.strftime("%d/%m/%Y") # Preenche a data das rodadas como sendo a última sexta
ws['AA2'] = data_marcacao_cadastroExcel[0].strftime("%d/%m/%Y")
ws['B3'] = lastFriday.strftime("%d/%m/%Y") # Preenche a data das rodadas como sendo a última sexta
ws['AA3'] = data_marcacao_cadastroExcel[1].strftime("%d/%m/%Y")
ws['B4'] = lastFriday.strftime("%d/%m/%Y") # Preenche a data das rodadas como sendo a última sexta
ws['AA4'] = data_marcacao_cadastroExcel[2].strftime("%d/%m/%Y")
#escreve a compila_cmo_medio
with open(path, 'wb') as output3:
    writer3 = pd.ExcelWriter(output3, engine='openpyxl')
    writer3.book = book
    compila_cmo.to_excel(writer3, sheet_name='compila_cmo_medio', index=False)
    writer3.save()

#Copia planilha da rede para o onedrive
# Source path
path_rede=path
# Destination path
path_onedrive = r"C:\Users\Middle\OneDrive - Eneva S.A\power_automate\1.PROSPEC\Base_PrecoMercado.xlsx"
shutil.copy(path_rede, path_onedrive)
