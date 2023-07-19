from openpyxl import load_workbook
import datetime
from Funcoes_API_Pluvia import getForecasts
from Funcoes_API_Pluvia import authenticatePluvia
from Funcoes_API_Pluvia import downloadForecast
from Funcoes_API_Pluvia import cria_pasta_local
from pathlib import Path
import os
import shutil
import warnings
from zipfile import ZipFile
from Funcoes_API_Pluvia import le_ena_pasta_bacias
from Funcoes_API_Pluvia import salva_ENA_base_bacias
from dateutil.relativedelta import relativedelta
import json
from Funcoes_API_Pluvia import atualiza_imprime_relatorio_previsao_ENA
from Funcoes_API_Pluvia import deleta_linhas_duplicadas_data
from Funcoes_API_Pluvia import verifica_mapas_ONS_baixados
from Funcoes_API_Pluvia import deleta_linhas_duplicadas
from Funcoes_API_prospec import envia_email_python
from Funcoes_API_Pluvia import cria_pasta_local_temporaria
import time

warnings.filterwarnings('ignore')
base_pluvia =r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\11.HIDROLOGIA\01.PLUVIA\API_Pluvia\Dados_API_Pluvia.xlsx' #arquivo que configura o download das informçãoes diárias
pathResult = Path(cria_pasta_local_temporaria()) #Path(r'C:\Users\fernando.fidalgo\OneDrive - Eneva S.A\03. Eneva\14. Comercializadora\05. Update_ONS\01.Pluvia') #caminho de download local dos arquivos
caminho_rede = Path(r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\11.HIDROLOGIA\05.PLUVIA_BACIAS') # caminho raiz do pluvia
caminho_base_pluvia =r'C:\Users\alex.lourenco\Documents\Pluvia\\Base_ENA_Pluvia_bacias.xlsx'
    #r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Base_ENA_Pluvia_bacias.xlsx'
caminho_onedrive=r'C:\Users\alex.lourenco\Documents\Pluvia\Base_ENA_Pluvia_bacias.xlsx'
caminho_local_raiz=r'C:\Users\alex.lourenco\Documents\Pluvia'
print(pathResult)
authenticatePluvia () #autentica
excel_ena_diaria = load_workbook(filename=base_pluvia, data_only=True)
ws = excel_ena_diaria['ENA_diaria']
ws_aux = excel_ena_diaria['aux']

#hoje = datetime.datetime(2022, 6, 21).strftime('%d/%m/%Y')
hoje = datetime.datetime.today().strftime('%d/%m/%Y') #datetime.datetime(2020, 10, 19)#datetime.datetime.today().strftime('%d/%m/%Y')
agora = datetime.datetime.now()
hora_limite_definitivo = agora.replace(hour=15)
print('------------------------------------------------------------------------------------------------------')
print('iniciando download de projeções para o dia', hoje, 'às ', datetime.datetime.now().strftime(('%H:%M:%S')))
linha = 8
download_arquivo = False
if agora >= hora_limite_definitivo:
    preliminary = 'definitiva'
    envia_email = False
else:
    preliminary = 'preliminar'
    envia_email = True
while ws.cell(linha, 4).value != None and ws.cell(linha, 4 ).value != '':
    if ws.cell(linha, 2).value == 'Habilitado':
        forecastDate = hoje #.strftime('%d/%m/%Y')
        x = 4
        while ws_aux.cell(x,7).value != None:
            if ws_aux.cell(x,7).value == ws.cell(linha, 4 ).value:
                precipitationDataSources = []
                precipitationDataSources.append(ws_aux.cell(x,6).value)
            x = x + 1
        #precipitationDataSources = [12]
        x = 4
        while ws_aux.cell(x, 2).value != None:
            if ws_aux.cell(x, 2).value == ws.cell(linha, 5).value:
                forecastModels = []
                forecastModels.append(ws_aux.cell(x, 1).value)
            x = x + 1
        #forecastModels = [2]
        bias = ws.cell(linha, 6).value  # True / False
        #if agora >= hora_limite_definitivo:
        #    preliminary = 'False'
        #    envia_email = False
        #else:
        #    preliminary = 'True'
        #    envia_email = True
        #if ws.cell(linha, 7).value == 'Definitivo':
        #    preliminary = 'False'
        #elif ws.cell(linha, 7).value == 'Preliminar':
        #    preliminary = 'True'
        years = [int(ws.cell(linha, 8).value)]
        members = ''
        if ws.cell(linha, 9).value == None:
            members = ''
        else:
            members = [ws.cell(linha, 9).value]
        modes='0'
        #print(precipitationDataSources)
        forecasts = getForecasts(forecastDate, precipitationDataSources, forecastModels, bias, preliminary, years,
                                members, modes)
        #print(forecasts)
        if forecasts==[]:
            print('Sem previsões para o Mapa ', ws.cell(linha, 4).value)
        else:
            for forecast in forecasts:
                ano = forecastDate[6:]
                mes = forecastDate[3:5]
                dia = forecastDate[:2]
                if forecast['rodada'] == 'definitiva':
                    prelim = 'definitiva'
                elif forecast['rodada'] == 'preliminar':
                    prelim = 'preliminar'
                else:
                    prelim = 'ERRO_PRELIM'
                    print('erro de código')
                nome_ENA = forecast['nome'] + '-' + forecast['membro'] + '-' + prelim + '-' + ano + mes + dia + '-ENA.zip'
                pathForecastDay = pathResult
                for i in range(1,7):
                        if forecast['resultados'][i - 1]['nome']!= None and forecast['resultados'][i - 1]['nome']=='ENA':
                         #print(forecast['resultados'][i - 1]['id'])
                         downloadForecast(forecast['resultados'][i - 1]['id'], pathForecastDay, nome_ENA)
                         download_arquivo=True
                        # elif forecast['resultados'][i - 1]['nome']!='ENA':
                        #     print('ENA não disponível para o seguinte forecast-->', 'Mapa:',
                        #           forecast['mapa'],') - Modelo:', forecast['modelo'])
                # else:
                #          print('ENA não disponível para o seguinte forecast-->', 'Mapa:', forecast['mapa'],
                #                ') - Modelo:', forecast['modelo'])
             # print('Mapa:', forecast['mapa'], '- Modelo:',forecast['modelo'], '- VNA disponívels:', forecast['vnaDisponivel'], '- ENA disponível:', forecast['enaDisponivel'],
             #         '- STR disponível:', forecast['strDisponivel'], '- PREVS disponível:', forecast['prevsDisponivel'])
    linha = linha + 1
excel_ena_diaria.close()
if download_arquivo: #só verifica a pasta se houver download de arquivos
    #descobre o nome do arquivo e cria um nome temporário para a cópia local
    caminho_base_pluvia_realpath = os.path.realpath(caminho_base_pluvia)
    caracter_ena = caminho_base_pluvia_realpath.rfind('\\') + 1
    nome_arquivo_ena = caminho_base_pluvia_realpath[caracter_ena:]
    nome_arquivo_temp = nome_arquivo_ena[:-5] + '_temp' + '.xlsx'
    caminho_arquivo_temp = os.path.join(Path(caminho_local_raiz), nome_arquivo_temp)
    print('Inciando cópia de arquivo original da rede no caminho: ', caminho_base_pluvia_realpath)
    shutil.copy2(caminho_base_pluvia, caminho_arquivo_temp)
    print('Arquivo temporário copiado para pasta local: ', nome_arquivo_temp)
    time.sleep(5)
    if preliminary == 'False':
        deleta_linhas_duplicadas_data(hoje, caminho_arquivo_temp)
    for arquivo in os.listdir(pathForecastDay): #lê os arquivos que estão na pasta de download
        if arquivo.upper().endswith('ENA.ZIP'): #exibe somente os arquvos de ENA
            caminho_arquivo = os.path.realpath(pathForecastDay) + '\\' + arquivo
            caminho_pasta = caminho_arquivo.replace('-ENA.zip', '')
            with ZipFile(caminho_arquivo, 'r') as zipObj: #descompacta o zip na mesma pasta do zip
                zipObj.extractall(caminho_pasta)
                print('arquivo descompactado:', arquivo)
            #print(caminho_pasta)
            tabela = le_ena_pasta_bacias(caminho_pasta) #função que lê o arquivo CSV de ENA dentro da pasta e retorna uma lista onde cada item da lista é uma linha, e cada linha é uma nova lista, onde cada item é uma coluna
            #print(tabela)
            data_mapa = tabela[0][0]
            mapa_atual = tabela[0][1]
            modelo = tabela[0][2]
            membro = tabela[0][3]
            deleta_linhas_duplicadas(data_mapa, mapa_atual, modelo, membro, caminho_arquivo_temp)
            salva_ENA_base_bacias(tabela, caminho_salvar_base_pluvia=caminho_arquivo_temp)  #função que lê a tabela gerada do arquivo de ENA lido e salva na base em excel
            os.makedirs(Path.joinpath(caminho_rede, ano, ano + '-' + mes, dia), exist_ok=True) #cria pasta na rede caso não exista
            shutil.move(Path.joinpath(Path(pathForecastDay), arquivo), #move os arquivos ZIP para a rede
                        Path.joinpath(caminho_rede, ano, ano + '-' + mes, dia, arquivo))
            print('Arquivo movido para a rede:', arquivo)
            shutil.rmtree(caminho_pasta)
            print('Pasta local deletada: ', caminho_pasta)
    shutil.copy2(caminho_arquivo_temp, caminho_onedrive)
    shutil.copy2(caminho_arquivo_temp, caminho_base_pluvia)
    print('Arquivo de base de dados Pluvia, movido para a rede')
    shutil.rmtree(pathForecastDay)
    print('Pasta temporária deletada:', pathForecastDay)
if envia_email:
#    resposta = atualiza_imprime_relatorio_previsao_ENA()
    assunto = 'RELATÓRIO DE PREVISÃO DIÁRIA DE ENA BACIAS - ' + hoje
    corpo = 'Prezados,\n\nSegue anexado o Relatório contendo a previsão de ENA baseada na previsão preliminar de hoje disponibilizada pelo Pluvia.'
#    arquivo_pdf = resposta[0]
#    caminho_pdf = resposta[1]
#    #caminho_gif = verifica_mapas_ONS_baixados ()
    destinatario = ['alex.lourenco@eneva.com.br'] #todosenevacom

#    #envia_email_pdf_gif (assunto, corpo, caminho_pdf, caminho_gif, destinatario)
    #envia_email_python(assunto, corpo,caminho_anexo=None, anexo=None, destinatario=destinatario)
#    #print('envio de e-mail desabilitado temporariamente')
print('Concluído o download de projeções para o dia', hoje, 'às ', datetime.datetime.now().strftime(('%H:%M:%S')))
