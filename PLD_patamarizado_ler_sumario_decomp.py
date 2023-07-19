from openpyxl import Workbook
from openpyxl.styles import Alignment

caminho_sumario = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\05. Diario\05.Arquivos Semanais\DC202108\Relatorio_Sumario-202108-sem4'
nome_sumario ='SUMARIO.RV3'
caminho_arquivo_sumario = caminho_sumario + '\\' + nome_sumario
texto_inicio = 'CUSTO MARGINAL DE OPERACAO'
texto_termino = 'Patamares de carga:'
lista_patamar = []
trig = False
with open(caminho_arquivo_sumario, 'r') as sumario:

    for line in sumario:
        if texto_inicio in line:
            trig = True
        elif texto_termino in line:
            trig = False
            break
        if trig:
            lista_patamar.append(line. replace('\n', ''))
        else:
            continue
print(lista_patamar)
#print(lista_patamar[1])
start = lista_patamar[1].find('X')
seg_x = lista_patamar[1].find('X', start + 1)
ter_x = lista_patamar[1].find('X', seg_x + 1)
qua_x = lista_patamar[1].find('X', ter_x + 1)
qui_x = lista_patamar[1].find('X', qua_x + 1)
#final = lista_patamar[1].find('X', qui_x + 1)
print(start, ' - ', seg_x, ' - ', ter_x, ' - ', qua_x, ' - ', qui_x)
carga_subm = {}
for item in range(4, len(lista_patamar)):
    dados = []
    if 'X------X' in lista_patamar[item]:
        break
    else:
        if 'Med' in lista_patamar[item][start:seg_x].replace(' ', '') and lista_patamar[item][start:seg_x].replace(' ', '') != 'Med_FC':
            subm = lista_patamar[item][start:seg_x].replace(' ', '')
            if subm == 'Med_SE':
                submercado = 'SUDESTE'
            elif subm == 'Med_S':
                submercado = 'SUL'
            elif subm == 'Med_NE':
                submercado = 'NORDESTE'
            elif subm == 'Med_N':
                submercado = 'NORTE'

            media = (lista_patamar[item][start:seg_x].replace(' ', ''), lista_patamar[item][seg_x:ter_x].replace(' ', ''))
            dados.append(media)
            #dados.append(lista_patamar[item][start:seg_x].replace(' ', ''))
            #dados.append(lista_patamar[item][seg_x:ter_x].replace(' ', ''))
            #dados.append(lista_patamar[item][ter_x:qua_x].replace(' ', ''))
            #dados.append(lista_patamar[item][qua_x:qui_x].replace(' ', ''))

            for x in range(3, 0, -1):

                crg_pat = (lista_patamar[item - x][start:seg_x].replace(' ', ''), lista_patamar[item - x][seg_x:ter_x].replace(' ', ''))
                #dados.append(lista_patamar[item - x][start:seg_x].replace(' ', ''))
                #dados.append(lista_patamar[item - x][seg_x:ter_x].replace(' ', ''))
                #dados.append(lista_patamar[item - x][ter_x:qua_x].replace(' ', ''))
                #dados.append(lista_patamar[item - x][qua_x:qui_x].replace(' ', ''))
        #dados.append(lista_patamar[item][qui_x:final].replace(' ', ''))
                dados.append(crg_pat)
            carga_subm[submercado] = dados
print(carga_subm)
wb_rascunho = Workbook()
ws_rascunho = wb_rascunho.active
lin = 2
ws_rascunho.cell(1, 1).value = 'SUBMERCADO'
ws_rascunho.cell(1, 2).value = 'PAT 1 (PES)'
ws_rascunho.cell(1, 3).value = 'PAT 2 (MED)'
ws_rascunho.cell(1, 4).value = 'PAT 3 (LEV)'
#for n in range(1, 11):
#    ws_rascunho.column_dimensions[n].width = 13.5
ws_rascunho.column_dimensions['B'].width = 13.5
ws_rascunho.column_dimensions['C'].width = 13.5
ws_rascunho.column_dimensions['D'].width = 13.5
ws_rascunho.column_dimensions['E'].width = 13.5
ws_rascunho.column_dimensions['F'].width = 13.5
ws_rascunho.column_dimensions['G'].width = 13.5
ws_rascunho.column_dimensions['H'].width = 13.5
ws_rascunho.column_dimensions['I'].width = 13.5
ws_rascunho.column_dimensions['J'].width = 13.5
ws_rascunho.column_dimensions['K'].width = 13.5
ws_rascunho.column_dimensions['L'].width = 13.5
#range_cell = ws_rascunho['B1':'D5']
#centraliza o texto nas células
for coluna in range(2,14):
    for linha in range (1,11):
        ws_rascunho.cell(linha, coluna).alignment = Alignment(horizontal='center')

for key, value in carga_subm.items():
    #print(sbm)
    for item in range(len(value)):
        print('item:', item, range(len(value)), 'Chave:', key, 'Valor:',value[item], ' - ', value[item][1] )
        ws_rascunho.cell(lin, 1).value = key
        ws_rascunho.cell(lin, item + 1).value = float(value[item][1])
    lin = lin + 1


col = 2
lin = 10
for key, value in carga_subm.items():
    #print(sbm)
    for item in range(1, len(value)):
        ws_rascunho.cell(lin - 2, col).value = key
        ws_rascunho.cell(lin - 1, col).value = value[item][0]
        ws_rascunho.cell(lin, col).value = float(value[item][1])
        col = col + 1

wb_rascunho.save(caminho_sumario + '\\' + 'rascunho_patamar.xlsx')
wb_rascunho.close()