import requests
import json
from pathlib import Path
import os
import time
import csv
import datetime
from openpyxl import load_workbook
from win32com import client
import urllib3
import datetime
import imghdr
import shutil
from zipfile import ZipFile
import warnings
from datetime import date
# -----------------------------------------------------------------------------
# Global variables | Variáveis globais
# -----------------------------------------------------------------------------
basicURL ='https://pluvia.app'

verifyCertificate = True
username = 'comercializacao.eneva'
password = 'T5Yx*CCuRM8@'
global token
caminho_base_pluvia_acumulado = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Base_ENA_Pluvia.xlsx'
# -----------------------------------------------------------------------------
# Get token | Obter token
# -----------------------------------------------------------------------------


def getToken(username_temp, password_temp):
    url = basicURL + '/api/token'

    global username
    username = username_temp
    global password
    password = password_temp

    headers = {
        'content-type': 'application/x-www-form-urlencoded'
    }

    data = {
        'grant_type': 'password',
        'username': username,
        'password': password
    }

    tokenResponse = requests.post(url, headers=headers, data=data,
                                  verify=verifyCertificate)
    token_json = tokenResponse.json()
    token = token_json["access_token"]

    return token


# -----------------------------------------------------------------------------
# Get token | Obter token
# -----------------------------------------------------------------------------


def authenticatePluvia():
    global token
    username = 'comercializacao.eneva'
    password = 'T5Yx*CCuRM8@'
    token = getToken(username, password)
    print('Autenticado com Sucesso no Pluvia')

# -----------------------------------------------------------------------------
# Cria pasta local para salvar os arquivos
# -----------------------------------------------------------------------------

def cria_pasta_local (caminho):
    os.makedirs(caminho, exist_ok=True)

# -----------------------------------------------------------------------------
# Get file from REST API | Obter arquivo via REST API
# -----------------------------------------------------------------------------


def getFileFromAPI(*args):

    if len(args) < 3:
        print('Sao necessarios ao menos tres argumentos: token e apiFunction')
        return ''
    elif len(args) == 3:
        token = args[0]
        apiFunction = args[1]
        fileName = args[2]
        pathToDownload = ''
    elif len(args) == 4:
        token = args[0]
        apiFunction = args[1]
        fileName = args[2]
        pathToDownload = args[3]
    else:
        print('Sao aceitos no máximo quatro argumentos: token, apiFunction e'
              ' parametros')
        return ''

    # Specify URL | Especificar URL
    url = basicURL + apiFunction

    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }

    # Call REST API | Chamar REST API
    response = requests.get(url, headers=headers, stream=True,
                            verify=verifyCertificate)

    print(response.status_code)
    #print(response.text)

    if (response.status_code == 401):
        token = getToken(username, password)

        headers = {
            'Authorization': 'Bearer ' + token,
            "Content-Type": "application/json"
        }

        response = requests.get(url, headers=headers, stream=True,
                                verify=verifyCertificate)

        print(response.status_code)

    if (response.status_code == 200):
        try:
            with open(pathToDownload.joinpath(fileName), 'wb') as file:
                for chunk in response.iter_content(chunk_size=1024):
                    if chunk:  # filter out keep-alive new chunks
                        file.write(chunk)
            print('Arquivo baixado: ', fileName)
        except:
            print('Não foi possível salvar o arquivo', str(pathToDownload.joinpath(fileName)))

    return ''

# -----------------------------------------------------------------------------
# Get JSON from REST API | Obter JSON via REST API
# -----------------------------------------------------------------------------


def getInfoFromAPI(*args):

    if len(args) < 2:
        print('Sao necessarios ao menos dois argumentos: token e apiFunction')
        return ''
    elif len(args) == 2:
        token = args[0]
        apiFunction = args[1]
    elif len(args) == 3:
        token = args[0]
        apiFunction = args[1]
        params = args[2]
    else:
        print('Sao necessarios ao menos dois argumentos: token, apiFunction e'
              ' parametros')
        return ''

    # Specify URL | Especificar URL
    url = basicURL + apiFunction

    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }

    # Call REST API | Chamar Rest API
    if 'params' in locals():
        response = requests.get(url, headers=headers, params=params,
                                verify=verifyCertificate)
    else:
        response = requests.get(url, headers=headers, verify=verifyCertificate)

    #print(response.status_code)

    if (response.status_code == 401):
        token = getToken(username, password)

        headers = {
            'Authorization': 'Bearer ' + token,
            "Content-Type": "application/json"
        }

        if 'params' in locals():
            response = requests.get(url, headers=headers, params=params,
                                    verify=verifyCertificate)
        else:
            response = requests.get(url, headers=headers,
                                    verify=verifyCertificate)

        #print(response.status_code)

    if (response.status_code == 200):
        return response.json()
    return ''





def downloadForecast(idForecast, pathToDownload, fileName):
    response = getFileFromAPI(token, '/api/resultados/' + str(idForecast), fileName, pathToDownload)


# -----------------------------------------------------------------------------
# Get list of Forecast Modelos | Obter lista dos Modelos de Previsão
# -----------------------------------------------------------------------------

def getIdsOfForecastModels():
    return getInfoFromAPI(token, '/api/valoresParametros/modelos')

def getIdOfForecastModel(forecastModel):
    "Possible values: IA, IA+SMAP or SMAP "
    return next(item for item in getIdsOfForecastModels() if item["descricao"] == forecastModel)['id']

# -----------------------------------------------------------------------------
# Get list of Precitation Data Source | Obter lista de Fonte de Dados de Precipitação
# -----------------------------------------------------------------------------


def getIdsOfPrecipitationsDataSource():
    return getInfoFromAPI(token, '/api/valoresParametros/mapas')


def getIdOfPrecipitationDataSource(precipitationDataSource):
    "Possible values: MERGE, ETA, GEFS, CFS, ONS NT0156, Usuário, Prec. Zero, ONS, ONS SOMBRA, ECMWF_ENS or ECMWF_ENS_EXT"
    return next(item for item in getIdsOfPrecipitationsDataSource() if item["descricao"] == precipitationDataSource)['id']


# -----------------------------------------------------------------------------
# Get list of Precitation Data Source | Obter lista de Fonte de Dados de Precipitação
# -----------------------------------------------------------------------------


def getForecasts(forecastDate, forecastSources, forecastModels, bias, preliminary, years, members):
    "forecastDate mandatory"

    if forecastDate == '':
        print('Data de Previsão não pode ser nula')
        return []

    params = "dataPrevisao=" + forecastDate

    for forecastSource in forecastSources:
        params += "&mapas=" + str(forecastSource)

    for forecastModel in forecastModels:
        params += "&modelos=" + str(forecastModel)

    if bias != '':
        params += "&semVies=" + str.lower(bias)

    if preliminary != '':
        params += "&preliminar=" + str.lower(preliminary)

    for year in years:
        params += "&anos=" + str(year)

    for member in members:
        params += "&membros=" + str(member)

    #return params
    return getInfoFromAPI(token, '/api/previsoes?' + params)

# -----------------------------------------------------------------------------
# Lê arquivo de ENA em uma pasta a partir de um caminho tradicional do wildown
# -----------------------------------------------------------------------------

def le_ena_pasta (caminho_pasta_real_path):
    #falta inserir o nome do
    dia = int(caminho_pasta_real_path[-2:])
    mes = int(caminho_pasta_real_path[-4:][:2])
    ano = int(caminho_pasta_real_path[-8:][:4])
    data = datetime.datetime(ano, mes, dia).strftime('%d/%m/%Y')

    dados_basicos_mapa = caracteristicas_mapa (caminho_pasta_real_path)
    dados_basicos_mapa.insert(0, data)
    tabela = []
    for arquivo in os.listdir(caminho_pasta_real_path):
        if arquivo.upper().endswith('-ENA.CSV'):
            with open(caminho_pasta_real_path + '\\' + arquivo) as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=';')
                line_count = 0
                pula_linha = 0
                for row in csv_reader:
                    linha = []
                    if row == ['### Resultados em Formato de Base de Dados']:
                        pula_linha = 1
                        continue
                    elif pula_linha == 1:
                        pula_linha = 2
                        continue
                    elif pula_linha == 2:
                        if row != [] and row[0] == 'Submercado' and row[3]== 'MEDIA':
                            linha.extend(dados_basicos_mapa)
                            linha.extend(row)
                            tabela.append(linha)
                            #print(linha)
    #print(tabela)
    return tabela

def le_ena_pasta_bacias (caminho_pasta_real_path):
    #falta inserir o nome do
    dia = int(caminho_pasta_real_path[-2:])
    mes = int(caminho_pasta_real_path[-4:][:2])
    ano = int(caminho_pasta_real_path[-8:][:4])
    data = datetime.datetime(ano, mes, dia).strftime('%d/%m/%Y')

    dados_basicos_mapa = caracteristicas_mapa (caminho_pasta_real_path)
    dados_basicos_mapa.insert(0, data)
    tabela = []
    for arquivo in os.listdir(caminho_pasta_real_path):
        if arquivo.upper().endswith('-ENA.CSV'):
            with open(caminho_pasta_real_path + '\\' + arquivo) as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=';')
                line_count = 0
                pula_linha = 0
                for row in csv_reader:
                    linha = []
                    if row == ['### Resultados em Formato de Base de Dados']:
                        pula_linha = 1
                        continue
                    elif pula_linha == 1:
                        pula_linha = 2
                        continue
                    elif pula_linha == 2:
                        if row != [] and (row[0] == 'Submercado' or row[0] == 'Bacia'):
                            linha.extend(dados_basicos_mapa)
                            linha.extend(row)
                            tabela.append(linha)
                            #print(linha)
    #print(tabela)
    return tabela
# -----------------------------------------------------------------------------
# Pega as características do mapa de acordo com o nome da pasta (em formato windows), retorna uma lista
# -----------------------------------------------------------------------------

def caracteristicas_mapa (caminho_pasta_real_path):
    nome_pasta = caminho_pasta_real_path[caminho_pasta_real_path.rfind('\\') + 1:]
    mapa = nome_pasta[:nome_pasta.find('-')]
    nome_pasta_sem_mapa = nome_pasta.replace(mapa + '-', '')
    if nome_pasta_sem_mapa[:nome_pasta_sem_mapa.find('-', )] == 'Preliminar':
        prelim = nome_pasta_sem_mapa[:nome_pasta_sem_mapa.find('-', )]
        nome_pasta_sem_prelim = nome_pasta_sem_mapa.replace(prelim + '-', '')
        modelo = nome_pasta_sem_prelim[:nome_pasta_sem_prelim.find('-', )]
        nome_pasta_sem_modelo = nome_pasta_sem_prelim.replace(modelo + '-', '')
    else:
        modelo = nome_pasta_sem_mapa[:nome_pasta_sem_mapa.find('-', )]
        nome_pasta_sem_modelo = nome_pasta_sem_mapa.replace(modelo + '-', '')
    membro = nome_pasta_sem_modelo[:nome_pasta_sem_modelo.find('-')]
    nome_pasta_sem_membro = nome_pasta_sem_modelo.replace(membro + '-', '')
    tipo_previsao = nome_pasta_sem_membro[:nome_pasta_sem_membro.find('-')]
    print('Nome da Pasta:', nome_pasta, ' - Mapa:', mapa, ' - Modelo:', modelo, ' - Membro:', membro, 'Tipo de Previsão:', tipo_previsao)
    return [mapa, modelo, membro, tipo_previsao]

# -----------------------------------------------------------------------------
# Pega uma lista de listas e salva em um arquivo excel - 1 lista tem uma lista de linhas, cada elemento é uma lista de colunas
# -----------------------------------------------------------------------------


def salva_ENA_base (tabela, **kwargs):

    caminho_base_pluvia_funcao = kwargs.get('caminho_salvar_base_pluvia')
    data = tabela[0][0]
    mapa = tabela[0][1]
    if caminho_base_pluvia_funcao:
        caminho_base_pluvia = caminho_base_pluvia_funcao
    else:
        caminho_base_pluvia = caminho_base_pluvia_acumulado
    #descobre o nome do arquivo e cria um nome temporário para a cópia local
    caminho_base_pluvia_realpath = os.path.realpath(caminho_base_pluvia)
    #caracter_ena = caminho_base_pluvia_realpath.rfind('\\') + 1
    #nome_arquivo_ena = caminho_base_pluvia_realpath[caracter_ena:]
    #nome_arquivo_temp = Path(nome_arquivo_ena[:-5] + '_temp' + '.xlsm')
    #print('Inciando cópia de arquivo original da rede no caminho: ', caminho_base_pluvia_realpath)
    #shutil.copy2(caminho_base_pluvia, nome_arquivo_temp)
    #print('Arquivo temporário copiado para pasta local: ', nome_arquivo_temp)
    #time.sleep(5)
    #arquivo_excel_base_pluvia = load_workbook(caminho_base_pluvia)
    #ws_base = arquivo_excel_base_pluvia['pluvia_definitivo']
    arquivo_excel_base_pluvia = client.DispatchEx("Excel.Application")
    wb_base_ena = arquivo_excel_base_pluvia.Workbooks.Open(Filename= caminho_base_pluvia )
    ws_base_ena = wb_base_ena.Worksheets('pluvia_definitivo')
    #lin = ws_base.max_row + 1
    lin = ws_base_ena.UsedRange.Rows.Count + 1
    for linha in tabela:
        for coluna in range(len(linha)):
            if (coluna + 1) == 1:
                mes_correto = linha[coluna][3:6]
                mes_dia_ano = mes_correto + linha[coluna].replace(linha[coluna][3:6], '')
                #print('Mês tabela:', linha[coluna], type(linha[coluna]), "Mês correto:", mes_correto, 'Mês Corrigido:', mes_dia_ano, type(mes_dia_ano) )

                ws_base_ena.Cells(lin, coluna + 1).Value = mes_dia_ano #mes_dia_ano
            elif (coluna + 1) == 10 or (coluna + 1) == 11:
                ws_base_ena.Cells(lin, coluna + 1).Value = float(str(linha[coluna]).replace(',','.'))

            else:
                ws_base_ena.Cells(lin, coluna + 1).Value = linha[coluna]
        lin = lin + 1
    wb_base_ena.Close(True)
    time.sleep(10)
    arquivo_excel_base_pluvia.Quit()
    del arquivo_excel_base_pluvia
    #arquivo_excel_base_pluvia.save(caminho_base_pluvia)
    #time.sleep(10)
    #arquivo_excel_base_pluvia.close()

def salva_ENA_base_bacias (tabela, **kwargs):

    caminho_base_pluvia_funcao = kwargs.get('caminho_salvar_base_pluvia')
    data = tabela[0][0]
    mapa = tabela[0][1]
    if caminho_base_pluvia_funcao:
        caminho_base_pluvia = caminho_base_pluvia_funcao
    else:
        caminho_base_pluvia = caminho_base_pluvia_acumulado
    #descobre o nome do arquivo e cria um nome temporário para a cópia local
    caminho_base_pluvia_realpath = os.path.realpath(caminho_base_pluvia)
    #caracter_ena = caminho_base_pluvia_realpath.rfind('\\') + 1
    #nome_arquivo_ena = caminho_base_pluvia_realpath[caracter_ena:]
    #nome_arquivo_temp = Path(nome_arquivo_ena[:-5] + '_temp' + '.xlsm')
    #print('Inciando cópia de arquivo original da rede no caminho: ', caminho_base_pluvia_realpath)
    #shutil.copy2(caminho_base_pluvia, nome_arquivo_temp)
    #print('Arquivo temporário copiado para pasta local: ', nome_arquivo_temp)
    #time.sleep(5)
    #arquivo_excel_base_pluvia = load_workbook(caminho_base_pluvia)
    #ws_base = arquivo_excel_base_pluvia['pluvia_definitivo']
    arquivo_excel_base_pluvia = client.DispatchEx("Excel.Application")
    wb_base_ena = arquivo_excel_base_pluvia.Workbooks.Open(Filename= caminho_base_pluvia )
    ws_base_ena = wb_base_ena.Worksheets('pluvia_definitivo')
    #lin = ws_base.max_row + 1
    lin = ws_base_ena.UsedRange.Rows.Count + 1
    for linha in tabela:
        for coluna in range(len(linha)):
            if (coluna + 1) == 1:
                #mes_correto = linha[coluna][3:6]
                #mes_dia_ano = mes_correto + linha[coluna].replace(linha[coluna][3:6], '')
                anocorreto=linha[coluna][6:10]
                mescorreto=linha[coluna][3:5]
                diacorreto=linha[coluna][0:2]

                ws_base_ena.Cells(lin, coluna + 1).Value = "'"+diacorreto+"/"+mescorreto+"/"+anocorreto #mes_dia_ano 
            elif (coluna + 1) == 9:##ADD contornar erro data, as vezes ficava d/m/a, as vezes m/d/a
                if linha[coluna]=='MEDIA':
                    valor='MEDIA'
                    ws_base_ena.Cells(lin, coluna + 1).Value =valor ##ADD                    
                else:
                    anocerto=linha[coluna][6:10]
                    mescerto=linha[coluna][3:5]
                    diacerto=linha[coluna][0:2]
                    #mes_certo = linha[coluna][3:6]
                    #valor =  str(mes_certo + linha[coluna].replace(linha[coluna][3:6], ''))     #mes_dia_ano               
                    ws_base_ena.Cells(lin, coluna + 1).Value ="'"+diacerto+"/"+mescerto+"/"+anocerto
                #ws_base_ena.Cells(lin, coluna + 1).Value =valor ##ADD
            elif (coluna + 1) == 10 or (coluna + 1) == 11:
                ws_base_ena.Cells(lin, coluna + 1).Value = float(str(linha[coluna]).replace(',','.'))

            else:
                ws_base_ena.Cells(lin, coluna + 1).Value = linha[coluna]
        lin = lin + 1
        print(lin)
    wb_base_ena.Close(True)
    time.sleep(10)
    arquivo_excel_base_pluvia.Quit()
    del arquivo_excel_base_pluvia
    #arquivo_excel_base_pluvia.save(caminho_base_pluvia)
    #time.sleep(10)
    #arquivo_excel_base_pluvia.close()

def atualiza_imprime_relatorio_previsao_ENA ():
    caminho_base_relatorio = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES'
    nome_arquivo = 'Relatorio_previsao_ENA.xlsm'
    caminho_arquivo_relatorio = caminho_base_relatorio + '\\' + nome_arquivo
    print('Abrindo arquivo base de relatório:', nome_arquivo)
    excel_macro_rel_ena = client.DispatchEx("Excel.Application")
    wb_rel_ena = excel_macro_rel_ena.Workbooks.Open(Filename= caminho_arquivo_relatorio )
    ws_rascunho_rel_ena = wb_rel_ena.Worksheets('rascunho')
    excel_macro_rel_ena.Visible = True
    print('Atualizando conexões do arquivo base de relatório:', nome_arquivo)
    wb_rel_ena.RefreshAll()
    #parâmetros da macro que imprime relatório
    nome_macro = 'imprimir_previsao_ENA_todas_abas'
    num_modulo = 1
    macro = 'Módulo' + str(num_modulo) + '.' + nome_macro
    wb_rel_ena.Application.Run("\'" + nome_arquivo + "\'" + '!' + macro)

    anexo = str(ws_rascunho_rel_ena.Cells(8, 5).Value)
    caminho_anexo = str(ws_rascunho_rel_ena.Cells(6, 5).Value) + str('\\') + str(anexo)
    wb_rel_ena.Close (False)
    excel_macro_rel_ena.Quit()
    del excel_macro_rel_ena

    return [anexo, caminho_anexo]


def verifica_mapas_ONS_baixados ():
    caminho_base_mapas_ONS = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\09.METEOROLOGIA'
    hoje = datetime.datetime.today()
    dia = '%02d' % hoje.day
    mes = '%02d' % hoje.month
    ano = str(hoje.year)
    pasta_dia = ano + '\\' + ano + '-' + mes + '\\' + dia
    caminho_dia = caminho_base_mapas_ONS + '\\' + pasta_dia
    lista_mapas = []
    if os.path.exists(caminho_dia):
        for pasta in os.listdir(caminho_dia):
            pasta_mapa = caminho_dia + '\\' + pasta
            for mapa_animado in  os.listdir(pasta_mapa):
                if mapa_animado.endswith('animado.gif'):
                    caminho_mapa = pasta_mapa + '\\' + mapa_animado
                    lista_mapas.append(caminho_mapa)
    return lista_mapas




def deleta_linhas_duplicadas(data_str, mapa, modelo, membro, caminho_base_pluvia_realpath):
    #data = '05/11/2020'
    #mapa = 'CFSv2'
    caminho_base_pluvia_realpath = os.path.realpath(caminho_base_pluvia_realpath)
    arquivo_excel = load_workbook(filename=caminho_base_pluvia_realpath)
    print('Base de dados em excel aberta')
    ws = arquivo_excel['pluvia_definitivo']
    linha = 2
    linhas_deletadas = 0

    while ws.cell(linha, 1).value is not None:
        if ws.cell(linha, 1).value == data_str and ws.cell(linha, 2).value == mapa and ws.cell(linha, 3).value == modelo and ws.cell(linha, 4).value == membro:
            linha_inicial = linha
            x = True
            y = 0
            while ws.cell(linha + y, 1).value == data_str and ws.cell(linha + y, 2).value == mapa and ws.cell(linha + y, 3).value == modelo and ws.cell(linha + y, 4).value == membro:
                y = y + 1

            ws.delete_rows(linha, y)
            linhas_deletadas = y
        else:
            linha = linha + 1
        #print(linha)
    arquivo_excel.save(caminho_base_pluvia_realpath)
    arquivo_excel.close()
    print('Foram deletadas ', linhas_deletadas, 'linhas.')



def deleta_linhas_duplicadas_data(data_str, caminho_base_pluvia_realpath):
    data_deletar = data_str #datetime.datetime.strptime(data_str, '%d/%m/%Y')
    print('Data que terão dados deletados da base:', data_deletar)
    #print(type(data_datetime))
    caminho_base_pluvia_realpath = os.path.realpath(caminho_base_pluvia_realpath)

    arquivo_excel_base_pluvia = client.DispatchEx("Excel.Application")
    wb_base_ena = arquivo_excel_base_pluvia.Workbooks.Open(Filename= caminho_base_pluvia_realpath )

    #arquivo_excel = load_workbook(filename=caminho_base_pluvia_realpath)
    print('Base de dados em excel aberta')
    ws_ena_del = wb_base_ena.Worksheets('pluvia_definitivo')
    linha = 2000
    linhas_deletadas = 0

    while ws_ena_del.Cells(linha, 1).Value is not None: # ws_ena_del.Cells(linha, 1).Value != '' or
        #print(ws_ena_del.Cells(linha, 1).Value)

        data_arquivo = ws_ena_del.Cells(linha, 1).Value
        #print('linha:', linha, data_arquivo, type(data_arquivo))
        if type(data_arquivo) is not str and (data_arquivo is not None):
            #data_arquivo = datetime.datetime.strptime(str(data_arquivo).rstrip("+00:00").replace(' ', ''), '%Y-%m-%d')
            data_arquivo = datetime.datetime.strftime(data_arquivo, '%d/%m/%Y')
            #print(data_arquivo)
            #data_arquivo = data_arquivo.strftime('%d/%m/%Y')
        if data_arquivo == data_deletar:
            print('Data igual na linha:', linha)
            y = linha
            nova_data_arquivo = data_arquivo
            while nova_data_arquivo == data_deletar:
                #print('entrou no while')
                data_arquivo = ws_ena_del.Cells(y, 1).Value

                if type(data_arquivo) is not str and (data_arquivo is not None):
                    #nova_data = datetime.datetime.strptime(str(ws_ena_del.Cells(y, 1).Value).rstrip("+00:00").replace(' ', ''), '%Y-%m-%d')
                    nova_data_arquivo = datetime.datetime.strftime(ws_ena_del.Cells(y, 1).Value, '%d/%m/%Y')
                else:
                    nova_data_arquivo = ws_ena_del.Cells(y, 1).Value
                y = y + 1
            ws_ena_del.Range(ws_ena_del.Cells(linha, 1), ws_ena_del.Cells(y-2, 1)).EntireRow.Delete()
            linhas_deletadas = y - 1 - linha
            print('Foram deletadas ', linhas_deletadas, 'linhas.')
        else:
            linha = linha + 1
        #print(linha)
    wb_base_ena.Close(True)
    time.sleep(10)
    arquivo_excel_base_pluvia.Quit()
    del arquivo_excel_base_pluvia
    print('Base de dados do Plúvia foi salva')


# -----------------------------------------------------------------------------
# Pega as configurações de ENA do arquivo base e retorna uma lista com cada mapa/modelo que dever[a ser baixado no pluvia
# -----------------------------------------------------------------------------

def le_configuracoes_ena(data_mapas, **kwargs):
    if kwargs.get('Diario'):
        caminho_dados_ENA_prevs = r'C:\Users\alex.lourenco\OneDrive - Eneva S.A\Documentos\processos_alex\diario\Dados_API_Pluvia_Diario.xlsx'
            #r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\\05. Diario\Dados_API_Pluvia_Diario.xlsx'
        print('Iniciando leitura das Configurações de PREVS diárias')
    else:
        caminho_dados_ENA_prevs = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Dados_API_Pluvia_Intersemanal.xlsx'
        print('Iniciando leitura das Configurações de ENA para o Intersemanal')


    excel = load_workbook(filename=caminho_dados_ENA_prevs, data_only=True)
    ws = excel['ENA_diaria']
    ws_aux = excel['aux']
    hoje = data_mapas #'26/11/2020'#datetime.datetime.today().strftime('%d/%m/%Y')
    lista_mapas = []
    linha = 8
    while ws.cell(linha, 2).value is not None:
        if ws.cell(linha, 2).value == 'Habilitado':
            forecastDate = hoje#.strftime('%d/%m/%Y')
            # Faz a tradução do mapa preenchido na planilha para o código do plúvia
            x = 4
            while ws_aux.cell(x,7).value != None:
                if ws_aux.cell(x,7).value == ws.cell(linha, 4 ).value:
                    precipitationDataSources = []
                    precipitationDataSources.append(ws_aux.cell(x,6).value)
                    nome_mapa = ws.cell(linha, 4 ).value
                x = x + 1
            # Faz a tradução do modelo preenchido na planilha para o código do plúvia
            x = 4
            while ws_aux.cell(x, 2).value is not None:
                if ws_aux.cell(x, 2).value == ws.cell(linha, 5).value:
                    forecastModels = []
                    forecastModels.append(ws_aux.cell(x, 1).value)
                x = x + 1
            #salva o dado da variável bias (Sem viés)
            bias = ws.cell(linha, 6).value  # True / False
            #faz a tradução do de definitivo/preliminar para False/true na variável preliminary
            if ws.cell(linha, 7).value == 'Definitivo':
                preliminary = 'False'
            elif ws.cell(linha, 7).value == 'Preliminar':
                preliminary = 'True'
            #salva variável years
            years = [int(ws.cell(linha, 8).value)]
            #define a variáel membro como vazia, para o caso de dar erro na leitura do campo no excel.
            members = ''
            #caso a célula do excel seja deixado em branco na coluna de membro, ele buscará todos o membros disponívels daquele mapa e modelo
            #caso contrário o excel deve ser preenchido com os membros entre aspas simples e separados por virgula Exemplo: '00', 'ENSEMBLE'
            if ws.cell(linha, 9).value == None:
                members = ''
            else:
                members = [ws.cell(linha, 9).value]
            forecast = {}
            forecast = {'forecastDate':forecastDate, 'mapa': precipitationDataSources, 'nome mapa': nome_mapa, 'modelo': forecastModels, 'bias': bias, 'preliminary': preliminary, 'years': years, 'members': members}

            lista_mapas.append(forecast)
        linha = linha + 1
    excel.close()
    return (lista_mapas, hoje)

# ---------------------------------------------------------------------------------------------------------------
# copia arquivo base de ENA para a pasta da rodada e retorna o caminho do arquivo base
# ---------------------------------------------------------------------------------------------------------------

def copia_arquivo_base_ena (data_str):
    caminho_base_pluvia_original = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\11.HIDROLOGIA\01.PLUVIA\API_Pluvia\Base_ENA_Pluvia_original.xlsx'
    caminho_base_rodada = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS'
    ano = data_str[-4:]
    mes = data_str[:5][-2:]
    dia = data_str[:2]
    caminho_ENA_rodada = caminho_base_rodada + '\\' + ano + '\\' + mes + '\\' + dia + '\\' + '05.ENA'
    nome_arquivo = 'Base_ENA_Pluvia.xlsx'
    os.makedirs(Path(caminho_ENA_rodada), exist_ok = True)
    shutil.copy2(Path(caminho_base_pluvia_original),Path.joinpath(Path(caminho_ENA_rodada), nome_arquivo))
    print('Base Pluvia Original copiada para o caminho da rodada')
    return caminho_ENA_rodada + '\\' + nome_arquivo

def percentile(data, percentile):
    size = len(data)
    data.sort(key=lambda x:x[1])
    data_id = int(round((size * percentile) / 100)) - 1
    #print('data_id: ', data_id)
    if data_id < 0:
        data_id = 0

    return data[data_id]


def baixa_prevs_padrao(lista_mapas, data_ec_ext, **kwargs):

    pathResult = Path(
        r'C:\Users\fernando.fidalgo\OneDrive - Eneva S.A\03. Eneva\14. Comercializadora\05. Update_ONS\01.Pluvia')  # caminho de download
    caminho_base_rodada = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS'
    ano = data_ec_ext[-4:]
    mes = data_ec_ext[:5][-2:]
    dia = data_ec_ext[:2]
    caminho_PREVS_rodada = caminho_base_rodada + '\\' + ano + '\\' + mes + '\\' + dia + '\\' + '02.Prevs'

    #02.Prevs'
    if kwargs.get('membros_EC_EXT'):
        membros_EC_EXT = kwargs.get('membros_EC_EXT')
        lista_mapas.append(('ECMWF_ENS_EXT', membros_EC_EXT))
        lista_membros = []
        for m in membros_EC_EXT:
            lista_membros.append(m[0])
    #if kwargs.get('caminho_rodada'):
    #    caminho_rodada = kwargs.get('caminho_rodada')
    print('Lista Membros do baixa_prevs_padrao:', lista_membros)
    authenticatePluvia()

    #hoje = datetime.datetime.today().strftime('%d/%m/%Y')
    forecastDate = data_ec_ext #'23/11/2020'  # hoje  # '29/09/2020'#hoje

    for mapa in lista_mapas:
        print (mapa)
        if mapa[0] == 'ONS':
            precipitationDataSources = [12]
            forecastModels = [2]
            bias = 'True'  # True (sem viés) / False (com viés)
            preliminary = 'False'  # True(preliminar) / False (definitivo)
            years = [2021]
            members = ''  # ['00', 'ENSEMBLE']
        elif mapa[0] == 'ECMWF_ENS':
            precipitationDataSources = [10]
            forecastModels = [2]
            bias = 'False'  # True / False
            preliminary = 'False'  # True / False
            years = [2021]
            members = ['ENSEMBLE']  # , 'ENSEMBLE'
        elif mapa[0] == 'ECMWF_ENS_EXT':
            precipitationDataSources = [11]
            forecastModels = [2]
            bias = 'False'  # True / False
            preliminary = 'False'  # True / False
            years = [2021]
            members = lista_membros
        elif mapa[0] == 'ECMWF_ENS_EXT-00':
            precipitationDataSources = [11]
            forecastModels = [2]
            bias = 'False'  # True / False
            preliminary = 'False'  # True / False
            years = [2021]
            members = ['00']  # , 'ENSEMBLE'
        elif mapa[0] == 'ECMWF_ENS_EXT-ENSEMBLE':
            precipitationDataSources = [11]
            forecastModels = [2]
            bias = 'False'  # True / False
            preliminary = 'False'  # True / False
            years = [2021]
            members = ['ENSEMBLE']
        elif mapa[0] == 'GEFS':
            precipitationDataSources = [3]
            forecastModels = [2]
            bias = 'False'  # True / False
            preliminary = 'False'  # True / False
            years = [2021]
            members = ['ENSEMBLE']
        elif mapa[0] == 'GEFS_00':
            precipitationDataSources = [3]
            forecastModels = [2]
            bias = 'False'  # True / False
            preliminary = 'False'  # True / False
            years = [2021]
            members = ['00']
        elif mapa[0] == 'CFS':
            precipitationDataSources = [4]
            forecastModels = [2]
            bias = 'False'  # True / False
            preliminary = 'False'  # True / False
            years = [2021]
            members = ['ENSEMBLE']

        pathForecastDay = pathResult.joinpath(forecastDate[6:] + '-' + forecastDate[3:5] + '-' + forecastDate[:2])
        cria_pasta_local(pathForecastDay)
        forecasts = getForecasts(forecastDate, precipitationDataSources, forecastModels, bias, preliminary, years,
                                 members)
        for forecast in forecasts:
            nome_prevs = forecast['nome'] + ' - ' + forecast['membro'] + ' - Prevs.zip'
            downloadForecast(forecast['prevsId'], pathForecastDay, nome_prevs)
            print(forecast)
            print(forecast['nome'], ' - ', 'PrevsId: ', forecast['prevsId'])
            novo_nome_prevs = nome_prevs.replace(' - Prevs.zip', '')
            membro = novo_nome_prevs[-2:]
            with ZipFile(os.path.join(pathForecastDay, nome_prevs), 'r') as zipObj:
                zipObj.extractall(os.path.join(pathForecastDay, novo_nome_prevs))
            time.sleep(10)
            for arquivo in os.listdir(os.path.join(pathForecastDay, novo_nome_prevs)):
                novo_nome = arquivo[:12] + arquivo[-4:]
                os.rename(os.path.join(pathForecastDay, novo_nome_prevs, arquivo),
                          os.path.join(pathForecastDay, novo_nome_prevs, novo_nome))
            completa_prevs (os.path.join(pathForecastDay, novo_nome_prevs))
            os.remove(os.path.join(pathForecastDay, nome_prevs))

            if mapa[0] == 'ECMWF_ENS_EXT':
                print('Mapa é ECMWF_ENS_EXT')
                for previsao in membros_EC_EXT:

                    if previsao[0] == membro:
                        print('previsao[0]:', previsao[0])
                        print('Membro:', membro)
                        pasta_destino = previsao[1]
                        for file in os.listdir(os.path.join(pathForecastDay, novo_nome_prevs)):
                            shutil.copy2(os.path.join(pathForecastDay, novo_nome_prevs, file),
                                        caminho_PREVS_rodada + '\\' + pasta_destino)
                shutil.rmtree(os.path.join(pathForecastDay, novo_nome_prevs))
            else:
                pasta_destino = mapa[1]
                for file in os.listdir(os.path.join(pathForecastDay, novo_nome_prevs)):
                    shutil.copy2(os.path.join(pathForecastDay, novo_nome_prevs, file),
                                caminho_PREVS_rodada + '\\' + pasta_destino)
                shutil.rmtree(os.path.join(pathForecastDay, novo_nome_prevs))



# ---------------------------------------------------------------------------------------------------------------
# lê o arquivo de configuração de mapas do intersemanal (para o caso de querer incluir o CFS, ainda a configurar)
# ---------------------------------------------------------------------------------------------------------------

#primeira função do pacote do intersemanal
#a princípio essa função baixa apenas os membros do EC estentidido, a ideia é poder utilizar o CFS no futuro também
#necessário incluir uma forma de baixar a previsão do dia que quiser, hoje está configurado para para baixar do próprio dia
#retorna o caminho raiz onde os arquivos baixados foram salvos

def baixa_ENA_configurada (data_mapas):
    warnings.filterwarnings('ignore')  # ignora avisos

    #pathResult = Path(
    #    r'C:\Users\fernando.fidalgo\OneDrive - Eneva S.A\03. Eneva\14. Comercializadora\05. Update_ONS\01.Pluvia')  # caminho de download local dos arquivos
    pathResult = Path(cria_pasta_local_temporaria())
    download_arquivo = False
    # traz a relação de ENAS a serem baixadas
    resposta = le_configuracoes_ena(data_mapas)
    lista_mapas = resposta[0]
    data_mapa = resposta[1]
    if lista_mapas == []:
        print('Nenhum mapa foi definido para download')
        quit()
    authenticatePluvia()

    # este looping faz o download de cada uma das previsões da lista de mapas
    for previsao in lista_mapas:
        forecasts = getForecasts(previsao['forecastDate'], previsao['mapa'], previsao['modelo'], previsao['bias'],
                                 previsao['preliminary'], previsao['years'], previsao['members'])
        if forecasts == []:
            print('Sem previsões para o Mapa ', previsao['mapa'])
        else:
            # looping em cada membro dentro de uma previsão específica
            for forecast in forecasts:
                # armazena variáveis de tempo para criar nome de pasta e nome do arquivo
                ano = previsao['forecastDate'][6:]
                mes = previsao['forecastDate'][3:5]
                dia = previsao['forecastDate'][:2]
                # transforma o valor da variável como o pluvia enxerga (false/true) para definitiva/preliminar para ser usada no nome do arquivo
                if forecast['preliminar'] == False:
                    prelim = 'definitiva'
                elif forecast['preliminar'] == True:
                    prelim = 'preliminar'
                else:
                    prelim = 'ERRO_PRELIM'
                    print('erro de código')
                # define o nome do arquivo de ENA que será baixado
                nome_mapa = forecast['nome'].replace('-' + forecast['membro'], '')
                nome_ENA = nome_mapa + '-' + forecast['membro'] + '-' + prelim + '-' + ano + mes + dia + '-ENA.zip'
                # define o caminho da pasta local para onde o arquivo de ENA será baixado
                pathForecastDay = pathResult #pathResult.joinpath(ano + '-' + mes + '-' + dia)
                # cria pasta local baseada nas variáveis de data da previsão
                #cria_pasta_local(pathForecastDay)
                # verifica se a ENA está disponível, caso disponível, ele faz o download daquela ENA
                if forecast['enaDisponivel'] and forecast['prevsDisponivel']:
                    downloadForecast(forecast['enaId'], pathForecastDay, nome_ENA)
                    # muda a variável para True, isso indicará mais na frente que algum arquivo de ENA foi baixado dentro de todas as previsões, para que o código possa fazer o tratamento dessas informações
                    download_arquivo = True
                else:
                    print('ENA não disponível para o seguinte forecast-->', 'Mapa:', forecast['mapa'], ') - Modelo:',
                          forecast['modelo'])
    return (pathForecastDay, data_mapa)

def baixa_prevs_configurada (lista_mapas):
    warnings.filterwarnings('ignore')  # ignora avisos
    #MFLBpathResult = Path(
    #    r'C:\Users\fernando.fidalgo\OneDrive - Eneva S.A\03. Eneva\14. Comercializadora\05. Update_ONS\01.Pluvia')
    pathResult = Path(cria_pasta_local_temporaria())
    if lista_mapas == []:
        print('Nenhum mapa foi definido para download')
        quit()
    authenticatePluvia()

    # este looping faz o download de cada uma das previsões da lista de mapas
    for previsao in lista_mapas:
        forecasts = getForecasts(previsao['forecastDate'], previsao['mapa'], previsao['modelo'], previsao['bias'],
                                 previsao['preliminary'], previsao['years'], previsao['members'])
        #print('forecasts:',forecasts)
        if forecasts == []:
            print('Sem previsões para o Mapa ', previsao['mapa'])
        else:
            # looping em cada membro dentro de uma previsão específica
            for forecast in forecasts:
                #print('forecast:', forecast)
                # armazena variáveis de tempo para criar nome de pasta e nome do arquivo
                ano = previsao['forecastDate'][6:]
                mes = previsao['forecastDate'][3:5]
                dia = previsao['forecastDate'][:2]
                # transforma o valor da variável como o pluvia enxerga (false/true) para definitiva/preliminar para ser usada no nome do arquivo
                if forecast['preliminar'] == False:
                    prelim = 'definitiva'
                elif forecast['preliminar'] == True:
                    prelim = 'preliminar'
                else:
                    prelim = 'ERRO_PRELIM'
                    print('erro de código')
                # define o nome do arquivo de ENA que será baixado
                nome_mapa = forecast['nome'].replace('-' + forecast['membro'], '')
                nome_prevs = nome_mapa + '-' + forecast['membro'] + '-' + prelim + '-' + ano + mes + dia + '-PREVS.zip'
                # define o caminho da pasta local para onde o arquivo de ENA será baixado
                pathForecastDay = pathResult.joinpath(ano + '-' + mes + '-' + dia)
                # cria pasta local baseada nas variáveis de data da previsão
                cria_pasta_local(pathForecastDay)
                # verifica se a ENA está disponível, caso disponível, ele faz o download daquela ENA
                if forecast['prevsDisponivel']:
                    downloadForecast(forecast['prevsId'], pathForecastDay, nome_prevs)
                    # muda a variável para True, isso indicará mais na frente que algum arquivo de ENA foi baixado dentro de todas as previsões, para que o código possa fazer o tratamento dessas informações
                    download_arquivo = True
                else:
                    print('PREVS não disponível para o seguinte forecast-->', 'Mapa:', forecast['mapa'], ') - Modelo:',
                          forecast['modelo'])
    return (pathForecastDay)


# ---------------------------------------------------------------------------------------------------------------
# copia base padrão de ENA para a pasta da rodada
# ---------------------------------------------------------------------------------------------------------------

def arquiva_base_ENA (data_ec_ext, caminho_local):
    pathForecastDay = os.path.realpath(caminho_local) #r'C:\Users\fernando.fidalgo\OneDrive - Eneva S.A\03. Eneva\14. Comercializadora\05. Update_ONS\01.Pluvia\2020-11-16'
    print(pathForecastDay)
    #data_ec = '16/11/2020'
    caminho_arquivo_ENA = copia_arquivo_base_ena(data_ec_ext) #copia base padrão de ENA para a pasta da rodada baseado na data de entrada
    print(caminho_arquivo_ENA)
    for arquivo in os.listdir(pathForecastDay):  # lê os arquivos que estão na pasta de download
        print(arquivo)
        if arquivo.upper().endswith('ENA.ZIP'):  # exibe somente os arquvos de ENA
            print(arquivo)
            caminho_arquivo = os.path.realpath(pathForecastDay) + '\\' + arquivo
            caminho_pasta = caminho_arquivo.replace('-ENA.zip', '')
            with ZipFile(caminho_arquivo, 'r') as zipObj:  # descompacta o zip na mesma pasta do zip
                zipObj.extractall(caminho_pasta)
                print('arquivo descompactado:', arquivo)
            # função que lê o arquivo CSV de ENA dentro da pasta e retorna uma lista onde cada item da lista é uma linha, e cada linha é uma nova lista, onde cada item é uma coluna
            tabela = []
            tabela = le_ena_pasta(caminho_pasta)
            data_mapa = tabela[0][0]
            mapa_atual = tabela[0][1]
            modelo = tabela[0][2]
            membro = tabela[0][3]

            salva_ENA_base(tabela,
                           caminho_salvar_base_pluvia=caminho_arquivo_ENA)  # função que lê a tabela gerada do arquivo de ENA lido e salva na base em excel

            shutil.rmtree(caminho_pasta) #deleta a pasta descompactada
            os.remove(os.path.join(pathForecastDay, arquivo)) #deleta o arquivo ZIP
            # print('Pasta local deletada: ', caminho_pasta)
    shutil.rmtree(caminho_local) #deleta a pasta temporária criada para baixar as ENAS
    return caminho_arquivo_ENA

def calcula_estatistica_ENA (caminho_base):
    #caminho_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\2020\11\16\05.ENA\Base_ENA_Pluvia.xlsx'

    arquivo_excel = load_workbook(caminho_base)
    arquivo_excel.create_sheet('estatisticas')
    ws_base = arquivo_excel['pluvia_definitivo']
    ws_estat = arquivo_excel['estatisticas']
    ws_base.cell(1, 12).value = 'mês_ref'
    ws_base.cell(1, 13).value = 'semana operativa'
    ws_base.cell(1, 14).value = 'ano_mes_semana'
    ws_base.cell(1, 15).value = 'ano_mes'
    ws_estat.cell(1, 2).value = 'TABELA DE MEMBROS DE ACORDO COM PERCENTIL'
    ws_estat.cell(1, 13).value = 'TABELA DE ENA (MWm) DE ACORDO COM PERCENTIL'
    ws_estat.cell(2, 2).value = 'Período'
    ws_estat.cell(2, 13).value = 'Período'
    ws_estat.cell(2, 3).value = 'Submercado'
    ws_estat.cell(2, 14).value = 'Submercado'
    ws_estat.cell(2, 4).value = 'Minimo'
    ws_estat.cell(2, 15).value = 'Minimo'
    ws_estat.cell(2, 5).value = 'P25'
    ws_estat.cell(2, 16).value = 'P25'
    ws_estat.cell(2, 6).value = 'P50'
    ws_estat.cell(2, 17).value = 'P50'
    ws_estat.cell(2, 7).value = 'P75'
    ws_estat.cell(2, 18).value = 'P75'
    ws_estat.cell(2, 8).value = 'Máximo'
    ws_estat.cell(2, 19).value = 'Máximo'

    linha = 2

    while ws_base.cell(linha, 1).value is not None:
        mes = str(ws_base.cell(linha, 8).value[6:8])
        ano = str(ws_base.cell(linha, 8).value[2:6])
        sem_op = 'RV' + str((int(ws_base.cell(linha, 8).value[12:13]) - 1))

        ws_base.cell(linha, 12).value = datetime.datetime(int(ano), int(mes), 1)
        ws_base.cell(linha, 12).number_format = 'DD/MM/YYYY'
        ws_base.cell(linha, 13).value = sem_op
        ws_base.cell(linha, 14).value = ano + mes + sem_op[-1:]
        ws_base.cell(linha, 15).value = ano + mes

        linha = linha + 1
    arquivo_excel.save(caminho_base)
    submercados = ['SIN', 'SUDESTE', 'SUL', 'NORDESTE', 'NORTE']
    lista_periodos = []
    lista_meses = []
    lista_membros = []
    percent_list = [0, 25, 50, 75, 100]
    linha = 2
    while ws_base.cell(linha, 14).value is not None:
        lista_membros.append(ws_base.cell(linha, 4).value)
        lista_periodos.append(ws_base.cell(linha, 14).value)
        lista_meses.append(ws_base.cell(linha, 15).value)
        linha = linha + 1

    lista_membros = list(set(lista_membros))
    val_remove = ['00', 'ENSEMBLE']
    for item in val_remove:
        if item in lista_membros:
            lista_membros.remove(item)
    # lista_membros.remove('ENSEMBLE')
    lista_periodos = list(set(lista_periodos))
    lista_meses = list(set(lista_meses))
    lista_membros.sort()
    lista_periodos.sort()
    lista_meses.sort()

    for submercado in submercados:
        for periodo in lista_periodos:
            linha = 2
            lista_sin = []
            while ws_base.cell(linha, 6).value is not None:
                if ws_base.cell(linha, 6).value == 'Submercado' and ws_base.cell(linha,
                                                                                 7).value == submercado and ws_base.cell(
                        linha, 14).value == periodo and ws_base.cell(linha, 4).value != '00' and ws_base.cell(linha,
                                                                                                              4).value != 'ENSEMBLE':
                    lista_sin.append(
                        (ws_base.cell(linha, 4).value, float(str(ws_base.cell(linha, 11).value).replace(',', '.'))))
                linha = linha + 1

            pri_lin_vaz = len(ws_estat['B']) + 1
            ws_estat.cell(pri_lin_vaz, 2).value = periodo
            ws_estat.cell(pri_lin_vaz, 3).value = submercado
            ws_estat.cell(pri_lin_vaz, 13).value = periodo
            ws_estat.cell(pri_lin_vaz, 14).value = submercado
            for p in range(len(percent_list)):
                resultado = percentile(lista_sin, percent_list[p])
                ws_estat.cell(pri_lin_vaz, p + 4).value = resultado[0]
                ws_estat.cell(pri_lin_vaz, p + 15).value = resultado[1]
                ws_estat.cell(pri_lin_vaz, p + 15).number_format = '#,##0.00'

    for submercado in submercados:
        for mes in lista_meses:
            lista_sin = []
            for membro in lista_membros:
                dados_membro_mes = []
                linha = 2
                while ws_base.cell(linha, 6).value is not None:
                    if ws_base.cell(linha, 6).value == 'Submercado' and ws_base.cell(linha,
                                                                                     7).value == submercado and ws_base.cell(
                            linha, 15).value == mes and ws_base.cell(linha, 4).value == membro and ws_base.cell(linha,
                                                                                                                4).value != '00' and ws_base.cell(
                            linha, 4).value != 'ENSEMBLE':
                        dados_membro_mes.append(float(str(ws_base.cell(linha, 11).value).replace(',', '.')))

                    linha = linha + 1
                media = sum(dados_membro_mes) / len(dados_membro_mes)
                lista_sin.append((membro, media))
            pri_lin_vaz = len(ws_estat['B']) + 1
            ws_estat.cell(pri_lin_vaz, 2).value = mes
            ws_estat.cell(pri_lin_vaz, 3).value = submercado
            ws_estat.cell(pri_lin_vaz, 13).value = mes
            ws_estat.cell(pri_lin_vaz, 14).value = submercado
            # p = ''
            for p in range(len(percent_list)):
                resultado = percentile(lista_sin, percent_list[p])
                ws_estat.cell(pri_lin_vaz, p + 4).value = resultado[0]
                ws_estat.cell(pri_lin_vaz, p + 15).value = resultado[1]
                ws_estat.cell(pri_lin_vaz, p + 15).number_format = '#,##0.00'

    arquivo_excel.save(caminho_base)
    arquivo_excel.close()

def analise_resultado_estatistica(caminho_arquivo_cria_estudo, caminho_arquivo_ENA):
    #caminho_arquivo_cria_estudo = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Criacao_Estudos_intersemanal_auto.xlsm'
    #caminho_arquivo_ENA = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Testes\Base_ENA_Pluvia_teste.xlsx'

    arquivo_excel = client.DispatchEx("Excel.Application")
    wb_study = arquivo_excel.Workbooks.Open(Filename=caminho_arquivo_cria_estudo)
    ws_config = wb_study.Worksheets('configuracoes')
    wb_ena = load_workbook(caminho_arquivo_ENA)
    ws_estat = wb_ena['estatisticas']

    linha = 26
    while ws_config.Cells(linha, 6).Value is not None:
        estatistica = ws_config.Cells(linha, 8).Value
        submercado = ws_config.Cells(linha, 9).Value
        periodo = ws_config.Cells(linha, 10).Value
        # semana = ws_config.Cells(linha, 11).Value
        mes_ref = ws_config.Cells(linha, 12).Value

        ano = str(mes_ref.year)
        mes = '%02d' % mes_ref.month
        ano_mes = ano + mes
        if estatistica == 'Menor':
            estat_ID = 1
        elif estatistica == 'P25':
            estat_ID = 2
        elif estatistica == 'Mediana':
            estat_ID = 3
        elif estatistica == 'P75':
            estat_ID = 4
        elif estatistica == 'Maior':
            estat_ID = 5

        linha_ENA = 3
        while ws_estat.cell(linha_ENA, 2).value is not None:
            if ws_estat.cell(linha_ENA, 2).value == ano_mes and ws_estat.cell(linha_ENA, 3).value == submercado:
                membro = ws_estat.cell(linha_ENA, 3 + estat_ID).value
                ws_config.Cells(linha, 13).NumberFormat = '@'
                ws_config.Cells(linha, 13).Value = membro
                #print(membro)
            linha_ENA = linha_ENA + 1

        linha = linha + 1

    wb_ena.close()
    wb_study.Save()
    wb_study.Close(False)
    arquivo_excel.Application.Quit()

def compila_baixa_prevs_intersemanal (caminho_arquivo_cria_estudo, data_ec_ext):
    #caminho_arquivo_cria_estudo = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal\Criacao_Estudos_intersemanal_auto.xlsm'

    arquivo_excel = client.DispatchEx("Excel.Application")
    wb_study = arquivo_excel.Workbooks.Open(Filename=caminho_arquivo_cria_estudo)
    ws_config = wb_study.Worksheets('configuracoes')
    mapas_fixos = []
    membros_EC_ext = []
    linha = 9
    if ws_config.Cells(linha, 6).Value is None:
        print('Rodada sem Mapas fixos')
    else:
        while ws_config.Cells(linha, 6).Value is not None:
            mapas_fixos.append((ws_config.Cells(linha, 6).Value, ws_config.Cells(linha, 8).Value))
            linha = linha + 1
    linha = 26
    if ws_config.Cells(linha, 6).Value is None:
        print('Rodada sem Mapas com Membros')
    else:
        while ws_config.Cells(linha, 6).Value is not None:
            membros_EC_ext.append((str('%02d' % int(ws_config.Cells(linha, 13).Value)), ws_config.Cells(linha, 14).Value))
            linha = linha + 1


    print(mapas_fixos)
    print(membros_EC_ext)
    # caminho da rodada
    baixa_prevs_padrao(mapas_fixos, data_ec_ext, membros_EC_EXT=membros_EC_ext)

    wb_study.Close(False)
    arquivo_excel.Application.Quit()
#-------------------------------------------------------------------------------------------------
# Completa as prevs reetroativo até RV0 (no primeiro mês) e até RV4 (no ultimo mês)
#-------------------------------------------------------------------------------------------------
def completa_prevs (caminho_prevs):
    caminho_prevs = os.path.realpath(caminho_prevs)
    lista_meses = []
    for file in os.listdir(caminho_prevs): #salva todos arquivos prevs naquela pasta em uma lista de tuplas
        lista_meses.append((file, file[:6] + file[-1:]))
    print(lista_meses)
    lista_meses.sort(key=lambda  x:x[1]) #ordena a lista de tuplas pelo segundo elemento da tupla, que é a juncao de ano + mes + revisao
    prim_rev = lista_meses[0][1] #identifica a primeira semana do primeiro mes
    ultm_rev = lista_meses[-1][1] #identifica a ultima semana do ultimo mes
    print('primeira revisao:', prim_rev)
    print ('Última revisao:', ultm_rev)
    if prim_rev[-1:] != '0': #loop que copia a primeira revisão até RV0
        for sem in range(int(prim_rev[-1:])-1,-1 , -1):

            #novo_nome_prevs = nome_prevs[:len(nome_prevs) - 1]
            #novo_nome_prevs = prim_rev[:6] + '-prevs.rv' + str(sem)
            for file in lista_meses:
                if file[1] == prim_rev:
                    novo_nome_prevs = file[0][:len(file[0]) - 1] + str(sem)
                    print(novo_nome_prevs)
                    shutil.copy2(caminho_prevs + '\\' + file[0],caminho_prevs + '\\' + novo_nome_prevs)
                    print('Arquivo copiado:', file[0], ' Novo Arquivo:', novo_nome_prevs)
    if ultm_rev[-1:] != '4': #loop que copia a última revisão até RV4
        for sem in range(int(ultm_rev[-1:])+1, 5):
            novo_nome_prevs = ultm_rev[:6] + '-prevs.rv' + str(sem)
            for file in lista_meses:
                if file[1] == ultm_rev:
                    novo_nome_prevs = file[0][:len(file[0]) - 1] + str(sem)
                    print(novo_nome_prevs)
                    shutil.copy2(caminho_prevs + '\\' + file[0],caminho_prevs + '\\' + novo_nome_prevs)
                    print('Arquivo copiado:', file[0], ' Novo Arquivo:', novo_nome_prevs)


def cria_pastas_rodada ():
    hoje = datetime.datetime.today()
    caminho_padrao = Path(r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS')
    dia = '%02d' % hoje.day
    mes = '%02d' % hoje.month
    ano = str(hoje.year)
    caminho_rodada = os.path.join(caminho_padrao, ano, mes, dia)
    pastas = ['01.Decks', '02.Prevs', '03.GEVAZP', '04.Download Estudos']
    for pasta in pastas:
        os.makedirs(os.path.join(caminho_rodada, pasta), exist_ok=True)
    print('Pastas padrão criadas no caminho da rodada:', caminho_rodada)
    return caminho_rodada

def verifica_disponibilidade_PREVS (data_mapas, **kwargs):
    warnings.filterwarnings('ignore')  # ignora avisos
    if kwargs.get('Preliminar'):
        prelim_var = 'True'
    else:
        prelim_var = 'False'
    # caminho de download local dos arquivos
    #MFLBpathResult = Path(r'C:\Users\fernando.fidalgo\OneDrive - Eneva S.A\03. Eneva\14. Comercializadora\05. Update_ONS\01.Pluvia')
    pathResult = Path(cria_pasta_local_temporaria())
    download_prevs = True
    # traz a relação de ENAS a serem baixadas
    resposta = le_configuracoes_ena(data_mapas, Diario=True)
    lista_mapas = resposta[0]
    #print('lista mapas:',lista_mapas)
    data_mapa = resposta[1]
    if lista_mapas == []:
        print('Nenhum mapa foi definido para download')
        quit()
    else:
        for previsao in lista_mapas:
            previsao['preliminary'] = prelim_var
    authenticatePluvia()
    # este looping verifica se as Prevs estão disponível
    prevs_disponibilidade = {'Disponivel': [], 'Indisponivel': []}

    for previsao in lista_mapas:
        forecasts = getForecasts(previsao['forecastDate'], previsao['mapa'], previsao['modelo'], previsao['bias'],
                                 prelim_var, previsao['years'], previsao['members'])
        #print('forecasts:',forecasts)
        if forecasts == []:
            print('Sem previsões para o Mapa ', previsao['nome mapa'], ' ( ID ', previsao['mapa'][0], ') - ', 'Hora:', datetime.datetime.now().strftime('%d/%m/%Y %H:%M %S'))
            download_prevs = False
        else:
            # looping em cada membro dentro de uma previsão específica
            for forecast in forecasts:
                if forecast['prevsDisponivel']:
                    print('prevs disponível para o mapa:', forecast['mapa'])
                    prevs_disponibilidade['Disponivel'].append(forecast['mapa'])
                else:
                    print('PREVS não disponível para o seguinte forecast-->', 'Mapa:', forecast['mapa'], ') - Modelo:',
                          forecast['modelo'])
                    prevs_disponibilidade['Indisponivel'].append(forecast['mapa'])
                    download_prevs = False
    return {'download_prevs': download_prevs, 'Data Mapas': data_mapa, 'PREVS Disponiveis':prevs_disponibilidade['Disponivel'], 'PREVS Indisponiveis':prevs_disponibilidade['Indisponivel'], 'Lista de Mapas': lista_mapas}


def cria_pasta_local_temporaria ():
    caminho_download_padrao = r'C:\SCRIPTS_\Temp'
        #r'C:\Users\middle\Desktop\arquivos_download\temp'
    cont = 0
    caminho_temp = caminho_download_padrao
    while os.path.exists(caminho_temp):
        cont = cont + 1
        caminho_temp = os.path.realpath(caminho_download_padrao) + '_' + str(cont)
    os.makedirs(caminho_temp)
    return caminho_temp