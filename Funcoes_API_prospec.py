from logging import info
import requests
import json
import os
from pathlib import Path
from openpyxl import load_workbook
import datetime
import time
import shutil
from win32com import client
import certifi
import urllib3
import datetime
import smtplib
from email.message import EmailMessage
from dateutil.relativedelta import relativedelta
import csv
from zipfile import ZipFile
import imghdr
from copy import deepcopy
import pendulum
import pandas as pd
from calendar import monthrange
from datetime import date, timedelta as delta
from funcao_pld_semanal import funcao_copia_cola_semanal
from funcao_pld_semanal import arquiva_pld_semanal
from funcao_pld_semanal import pld_realizado_semanal
from funcao_pld_semanal import arquiva_valor_semanal
from funcao_pld_semanal import salva_arquivo_semanal 
from funcao_pld_semanal import criar_copia_semanal
from funcao_pld_semanal import retorna_arquivo_semanal
from funcao_pld_semanal import StudyId_semanal
from funcao_pld_semanal import ena_realizada_semanal
from funcao_pld_semanal import salvar_copia_semanal
from funcao_pld_semanal import apagar_linhas_semanal
from funcao_pld_intersemanal import arquiva_pld
from funcao_pld_intersemanal import pld_realizado
from funcao_pld_intersemanal import arquiva_valor
from funcao_pld_intersemanal import salva_arquivo 
from funcao_pld_intersemanal import criar_copia
from funcao_pld_intersemanal import retorna_arquivo
from funcao_pld_intersemanal import StudyId
from funcao_pld_intersemanal import ena_realizada
from funcao_pld_intersemanal import salvar_copia
from funcao_pld_intersemanal import apagar_linhas
from funcao_pld_intersemanal import funcao_copia_cola
from get_pld_semanal import next_friday

piso =  float(os.environ.get('PLD_PISO', '55.7'))
teto_diario = float(os.environ.get('PLD_TETO_DIARIO', '640.5')) 
teto_horario =  float(os.environ.get('PLD_TETO_HORARIO', '1314.02')) 
actual_week_historical = os.environ.get('ACTUAL_WEEK_HISTORICAL', 'TRUE').upper() == 'TRUE'

api_url_base = 'https://api.prospec.app'
#'https://api.prospec.norus.com.br' #endereço antigo da API
verifyCertificate = True #certifi.where() #True mudança realizada para corrigir o erro de SSL

chromedriver = r'C:\Users\Middle\Documents\01Chromedriver\chromedriver_win32\chromedriver'

token = ''
#------------------------------------------------------------------------------
#Autenticar no site do Prospec
#------------------------------------------------------------------------------
def autenticar_prospec ():
    global token
    api_function = '/api/Token'
    url = api_url_base + api_function
    urllib3.disable_warnings((urllib3.exceptions.InsecureRequestWarning))
    username = 'middle@eneva.com.br' #'renata.hunder@eneva.com.br'
    password = 'j0MKyBVP'
    headers = {
        'content-type': 'application/x-www-form-urlencoded'
    }

    data = {
        'grant_type': 'password',
        'username': username,
        'password': password
    }
    response = requests.post(url, headers=headers, data=data, verify=verifyCertificate)
    response_json = response.json()
    token = response_json["access_token"]
    #print(response_json)
    return token

# pega quantidade de requisições
def conta_requisicoes (token):
    api_function = '/api/Account/Requests'
    url = api_url_base + api_function
    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }
    response_api = requests.get(url, headers=headers, verify=verifyCertificate)
    number_of_request = response_api.json()
    #print('Você já usou ', number_of_request, ' requests')
    return number_of_request

# -----------------------------------------------------------------------------
# Create one study | Criar um estudo
# -----------------------------------------------------------------------------


def createStudy(title, description, idDecomp, idNewave):
    parameter = ''
    data = {
        "Title": title,
        "Description": description,
        "DecompVersionId": int(idDecomp),
        "NewaveVersionId": int(idNewave)
    }

    print("Creating study with the following configuration:")
    print(data)

    prospecStudyId = postInAPI(token, '/api/prospectiveStudies', parameter,
                               data)
    return prospecStudyId

def cria_estudo(nome_estudo):
    api_function = '/api/prospectiveStudies'
    url = api_url_base + api_function
    params =''
    data = {
        "Title": nome_estudo,
        "Description": 'Descrição do estudo',
        "DecompVersionId": 0, #int(4),
        "NewaveVersionId": 0, #int(5)
    }
    prospecStudyId = ''
    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }

    response = requests.post(url, headers=headers, params=params,
                             data=json.dumps(data), verify=verifyCertificate)
    idStudy = response.text
    print('Código do Status: ', response.status_code)
    print('ID do Estudo:' + idStudy)
    print(json.dumps(response.json(),indent=2))
    return idStudy
#TÉRMINO DA FUNÇÃO QUE CRIA O ESTUDO




# -----------------------------------------------------------------------------
# Get list of DECOMPs | Obter lista de DECOMPs
# -----------------------------------------------------------------------------


def getListOfDECOMPs():
    apiFunction = '/api/CepelModels/Decomps'
    url = api_url_base + apiFunction
    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }
    response = requests.get(url, headers=headers, verify=verifyCertificate)
    if (response.status_code == 200):
        return response.json()
    return ''


# -----------------------------------------------------------------------------
# Get list of DECOMPs and choose one | Obter lista de DECOMPs e escolher um
# -----------------------------------------------------------------------------


def getIdOfDECOMP(version):
    apiFunction = '/api/CepelModels/Decomps'

    listOfDecomps = getListOfDECOMPs()

    idDecomp = ''
    for decomp in listOfDecomps:
        if decomp['Version'] == version:
            idDecomp = decomp['Id']
            return idDecomp

    return 0

# -----------------------------------------------------------------------------
# Get list of NEWAVES | Obter lista de NEWAVES
# -----------------------------------------------------------------------------


def getListOfNEWAVES():
    apiFunction = '/api/CepelModels/Newaves'
    url = api_url_base + apiFunction
    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }
    response = requests.get(url, headers=headers, verify=verifyCertificate)
    if (response.status_code == 200):
        return response.json()
    return ''

# -----------------------------------------------------------------------------
# Send files to a study | Enviar arquivos para um estudo
# -----------------------------------------------------------------------------


def sendFileToStudy(idStudy, pathToFile, fileName):
    apiFunction = '/api/prospectiveStudies/' + str(idStudy) + '/UploadFiles'
    prospecStudy = sendFileToAPI(token, apiFunction ,
                                 pathToFile, fileName)
    return prospecStudy

# -----------------------------------------------------------------------------
# Send files to a deck of a study | Enviar arquivos para um deck de um estudo
# -----------------------------------------------------------------------------


def sendFileToDeck(idStudy, idDeck, pathToFile, fileName):
    apiFunction = '/api/prospectiveStudies/' + str(idStudy) + '/UploadFiles?deckId=' + str(idDeck)
    prospecStudy = sendFileToAPI(token, apiFunction, pathToFile, fileName)
    return prospecStudy

# -----------------------------------------------------------------------------
# Get list of Decks from Study | Obter lista de Decks do Estudo
# -----------------------------------------------------------------------------


def getListOfDecks(idStudy):
    apiFunction = '/api/prospectiveStudies/' + str(idStudy)
    url = api_url_base + apiFunction
    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }
    response = requests.get(url, headers=headers, verify=verifyCertificate)
    response_json = response.json()

    return response_json['Decks']


# -----------------------------------------------------------------------------
# Send files to REST API | Enviar arquivos para REST API
# -----------------------------------------------------------------------------


def sendFileToAPI(*args):

    if len(args) < 4:
        print('Sao necessarios quatro argumentos: token, apiFunction, '
              'pathToFile e nameToFile')
        return ''
    elif len(args) == 4:
        token = args[0]
        apiFunction = args[1]
        pathToFile = args[2]
        nameToFile = args[3]
    else:
        print('Sao necessarios quatro argumentos: token, apiFunction, '
              'pathToFile e nameToFile')
        return ''

    # Specify URL | Especificar URL
    url = api_url_base + apiFunction

    headers = {
        'Authorization': 'Bearer ' + token
    }

    # Call REST API | Chamar via REST API
    files = {
        'file': (nameToFile, open(pathToFile, 'rb'),
                 'multipart/form-data', {'Expires': '0'})
    }
    response = requests.post(url, headers=headers, files=files,
                             verify=verifyCertificate)

    print(response.status_code)
    print(response.text)

#    if (response.status_code == 401):
#        token = getToken(username, password)
#
#        headers = {
#            'Authorization': 'Bearer ' + token,
#            "Content-Type": "application/json"
#        }
#
#        response = requests.post(url, headers=headers, files=files,
#                                 verify=verifyCertificate)

#        print(response.status_code)
#        print(response.text)

    if (response.status_code == 200):
        return response.json()
    elif (response.status_code == 201):
        return response

    return ''

# -----------------------------------------------------------------------------
# Generate decks to a prospective study
# Gerar decks para um estudo prospectivo
# -----------------------------------------------------------------------------


def generateStudyDecks(idStudy, initialYear, initialMonth, duration, month,
                       year, multipleStages, multipleRevision, firstNewaveFile,
                       otherNewaveFiles, decompFile, spreadsheetFile, tags):
    api_function = '/api/prospectiveStudies/' + str(idStudy) + '/Generate'
    listOfDeckConfiguration = []
    listOfTags = []

    i = 0
    for deck in month:
        deckConfiguration = {}
        deckConfiguration['Year'] = year[i]
        deckConfiguration['Month'] = month[i]
        deckConfiguration['MultipleStages'] = multipleStages[i]
        deckConfiguration['MultipleRevisions'] = multipleRevision[i]
        if (i > 0):
            if (otherNewaveFiles[i] != ''):
                deckConfiguration['NewaveUploaded'] = otherNewaveFiles[i]
        listOfDeckConfiguration.append(deckConfiguration)
        i = i + 1

    for tag in tags:
        tagsConfiguration = {}
        tagsConfiguration['Text'] = tag
        listOfTags.append(tagsConfiguration)

    parameter = ''
    data = {
        "InitialYear": initialYear,
        "InitialMonth": initialMonth,
        "Duration": duration,
        "DeckCreationConfigurations": listOfDeckConfiguration,
        "Tags": listOfTags,
        "InitialFiles": {
            "NewaveFileName": firstNewaveFile,
            "DecompFileName": decompFile,
            "SpreadsheetFileName": spreadsheetFile
        }
    }

    print("Gerando decks com as seguintes configuracoes para o estudo: ",
          str(idStudy))
    print(data)

    postInAPI(token, api_function, parameter, data)

# -----------------------------------------------------------------------------
# Post in REST API | Postar via REST API
# -----------------------------------------------------------------------------


def postInAPI(*args):

    if len(args) < 4:
        print('Sao necessarios quatro argumentos: token, apiFunction, parametros'
              ' e dados')
        return ''
    elif len(args) >= 4:
        token = args[0]
        apiFunction = args[1]
        params = args[2]
        data = args[3]
    else:
        print('Sao necessarios quatro argumentos: token, apiFunction, parametros'
              ' e dados')
        return ''

    # Specify URL | especificar URL
    url = api_url_base + apiFunction

    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }

    # Call REST API | Chamar REST API
    response = requests.post(url, headers=headers, params=params,
                             data=json.dumps(data), verify=verifyCertificate)

    print(response.status_code)
    print(response.text)

    if (response.status_code == 401):
        token = autenticar_prospec ()

        headers = {
            'Authorization': 'Bearer ' + token,
            "Content-Type": "application/json"
        }

        response = requests.post(url, headers=headers, params=params,
                                 data=json.dumps(data), verify=verifyCertificate)

        print(response.status_code)
        print(response.text)
        print('ERRO FOI AQUI')

    if (response.status_code == 200):
        return response.text
    elif (response.status_code == 201):
        return response.text

    return ''

# -----------------------------------------------------------------------------
# Send files | Enviar arquivos
# -----------------------------------------------------------------------------


def sendFiles(*args):

    if len(args) < 3:
        print('Sao necessarios tres argumentos: token, pathToFile e '
              'nameToFile')
        return ''
    elif len(args) == 3:
        token = args[0]
        apiFunction = args[1]
        files = args[2]
    else:
        print('Sao necessarios tres argumentos: token, pathToFile e '
              'nameToFile')
        return ''

    # Specify url | Especificar URL
    url = api_url_base + apiFunction

    headers = {
        'Authorization': 'Bearer ' + token
    }

    response = requests.post(url, headers=headers, files=files,
                             verify=verifyCertificate)

    print(response.status_code)
    print(response.text)
    print(json.dumps(response.json(), indent=2))

    if (response.status_code == 401):
        token = autenticar_prospec()

        headers = {
            'Authorization': 'Bearer ' + token,
            "Content-Type": "application/json"
        }

        response = requests.post(url, headers=headers, files=files,
                                 verify=verifyCertificate)

        print(response.status_code)
        print(response.text)

    if (response.status_code == 200):
        return response.json()
    elif (response.status_code == 201):
        return response

    return ''

# -----------------------------------------------------------------------------
# Send all prevs files to a study | Enviar todos os arquivos prevs de um estudo
# -----------------------------------------------------------------------------


def sendAllPrevsToStudy(idStudy, pathToAllPrevs):
    apifunction = '/api/prospectiveStudies/' + str(idStudy) + '/UploadMultiplePrevs'
    listOfPrevs = {}

    for file in os.listdir(pathToAllPrevs):
        if 'prevs' in file.lower():
            listOfPrevs[file] = [file, open((pathToAllPrevs + '\\' + file), 'rb'), 'multipart/form-data', {'Expires': '0'}]
    print(listOfPrevs)
    sendFiles(token, apifunction, listOfPrevs)

# -----------------------------------------------------------------------------
# Get file from S3 via REST API | Obter arquivos do S3 via REST API
# -----------------------------------------------------------------------------

def getFileFromS3viaAPI(*args):
    
    if len(args) < 3:
        print('Sao necessarios ao menos tres argumentos: token e apiFunction')
        return ''
    elif len(args) == 3:
        token = args[0]
        apiFunction = args[1]
        fileName = args[2]
        pathToDownload = ''
    elif len(args) == 4:
        token = args[0]
        apiFunction = args[1]
        fileName = args[2]
        pathToDownload = args[3]
    else:
        print('Sao aceitos no máximo quatro argumentos: token, apiFunction e'
                ' parametros')
        return ''
    
    #Specify URL | Especificar URL
    url = api_url_base + apiFunction

    headers = {
        'Authorization': 'Bearer ' + token, 
        "Content-Type": "application/json"
        }
    
    #Call REST API | Chamar REST API
    response = requests.get(url, headers=headers, stream=True, 
                             verify=verifyCertificate, allow_redirects=True)
        
    print(response.status_code)
      
    if (response.status_code == 401):
        token = getToken(username, password)
        
        headers = {
            'Authorization': 'Bearer ' + token, 
            "Content-Type": "application/json"
            }
        
        response = requests.get(url, headers=headers, stream=True, 
                                 verify=verifyCertificate, allow_redirects=True)
        
        print(response.status_code)
    
    if (response.status_code == 200):
        with open((pathToDownload + fileName), 'wb') as file:
            for chunk in response.iter_content(chunk_size=1024): 
                if chunk: # filter out keep-alive new chunks
                    file.write(chunk)
        
    return ''

# -----------------------------------------------------------------------------
# Get file from S3 via REST API | Obter arquivos do S3 via REST API
# -----------------------------------------------------------------------------

def getFileFromS3viaAPIV2(*args):
    
    if len(args) < 5:
        print('Sao necessarios ao menos cinco argumentos: token e apiFunction')
        return ''
    elif len(args) == 5:
        token = args[0]
        apiFunction = args[1]
        fileNames = list(args[2])
        fileName = args[3]
        pathToDownload = args[4]
    else:
        print('Sao aceitos no máximo cinco argumentos: token, apiFunction, fileNames, fileName e'
                ' pathToDownload')
        return ''
    
    #Specify URL | Especificar URL
    url = api_url_base + apiFunction
    
    headers = {
        'Authorization': 'Bearer ' + token, 
        "Content-Type": "application/json"
        }

    #Call REST API | Chamar REST API

    dado = {"FileNames": fileNames}

    response = requests.post(url, headers=headers, data=json.dumps(dado), verify=verifyCertificate)
        
    print(response.status_code)
      
    if (response.status_code == 401):
        token = getToken(username, password)
        
        headers = {
            'Authorization': 'Bearer ' + token, 
            "Content-Type": "application/json"
            }
        
        response = requests.get(url, headers=headers, stream=True, 
                                 verify=verifyCertificate, allow_redirects=True)
        
        print(response.status_code)
    
    if (response.status_code == 200):
        with open((pathToDownload + fileName), 'wb') as file:
            for chunk in response.iter_content(chunk_size=1024): 
                if chunk: # filter out keep-alive new chunks
                    file.write(chunk)
        
    return ''

# -----------------------------------------------------------------------------
# Download study | Download de um estudo
# -----------------------------------------------------------------------------


def downloadDecksOfStudy(idStudy, pathToDownload, fileName):
    response = getFileFromAPI(token, '/api/prospectiveStudies/' + str(idStudy)
                              + '/DeckDownload', fileName, pathToDownload)

# -----------------------------------------------------------------------------
# Download File From Deck Results | Download de um arquivo de um resultado do deck
# ----------------------------------------------------------------------------- #NEW

def downloadFileFromDeck(idDeck, pathToDownload, fileNameDownload, fileNames):
    filesToGet = 'fileNames=' + '&fileNames='.join(fileNames)
    response = getFileFromS3viaAPI(token, '/api/prospectiveStudies/DownloadResultFiles/' + str(idDeck) 
                                + '?' + filesToGet, fileNameDownload, pathToDownload)

# -----------------------------------------------------------------------------
# Download File From Deck Results V2 | Download de um arquivo de um resultado do deck V2
# ----------------------------------------------------------------------------- #NEW

def downloadFileFromDeckV2(idDeck, pathToDownload, fileNameDownload, fileName, fileNames):
    filesToGet = fileNames
    response = getFileFromS3viaAPIV2(token, '/api/v2/prospectiveStudies/DownloadResultFiles/' + str(idDeck), filesToGet, fileName, pathToDownload)



# -----------------------------------------------------------------------------
#  Download dos estudos compilados a partir do Excel que gerou esses arquivos
# -----------------------------------------------------------------------------



def download_compilado(token, idStudy, pathdownload, nome_arquivo):
    api_function = '/api/prospectiveStudies/' + str(idStudy) + '/CompilationDownload'
    url = api_url_base + api_function
    #verifyCertificate = False
    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }
    caminho_arquivo = os.path.join(pathdownload, nome_arquivo)
#    nome_arquivo = inicio_pasta + '_' + idStudy + '.zip'

    # verifica se o caminho existe, senão ele cria as pastas
    if not os.path.exists(pathdownload):
        os.makedirs(pathdownload)
        print(str(idStudy) + ' - Pasta criada')
    else:
        #print(str(idStudy) + ' - Pasta já existe')
        validador = 1

    response_api = requests.post(url, headers=headers, stream=True, verify=verifyCertificate)

    # salva o arquivo zipado na pasta determinada
    with open(caminho_arquivo, 'wb') as file:
        for chunk in response_api.iter_content(chunk_size=1024):
            if chunk:  # filter out keep-alive new chunks
                file.write(chunk)
        file.close()

# -----------------------------------------------------------------------------
#  Pega Status dos estudos criados
# -----------------------------------------------------------------------------

def GetStatusOfStudy (token, idStudy):
    api_function = '/api/prospectiveStudies/' + str(idStudy)
    url = api_url_base + api_function
    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }
    #print('token: ', token)
    #print('idStudy: ', idStudy)
    response_api = requests.get(url, headers=headers, verify=verifyCertificate)
    #print(response_api)
    response_api_status = response_api.json()['Status']
    return response_api_status

def cria_pasta_rodada_prevs():
    data_atual = datetime.datetime.today()
    dia = data_atual.day
    mes = data_atual.month
    ano = data_atual.year
    if dia < 10:
        dia = str(0) + str(dia)
    if mes < 10:
        mes = str(0) + str(mes)
    caminho = Path(r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS')
    caminho_dia = os.path.join(caminho, str(ano), str(mes), str(dia),'02.Prevs')
    if not(os.path.exists(caminho_dia)):
        os.makedirs(caminho_dia)
    return caminho_dia

def cria_pastas_rodada (**kwargs):
    if kwargs.get('caminho_arquivo_estudos'):
        if kwargs.get('tipo_rodada') == 'Diario':
            caminho_padrao = Path(r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\05. Diario\07.Rodadas')
    else:
        caminho_padrao = Path(r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS')
    hoje = datetime.datetime.today()
    
    dia = '%02d' % hoje.day
    mes = '%02d' % hoje.month
    ano = str(hoje.year)
    caminho_rodada = os.path.join(caminho_padrao, ano, mes, dia)
    pastas = ['01.Decks', '02.Prevs', '03.GEVAZP', '04.Download Estudos']
    for pasta in pastas:
        os.makedirs(os.path.join(caminho_rodada, pasta), exist_ok=True)
    print('Pastas padrão criadas no caminho da rodada:', caminho_rodada)
    return caminho_rodada


#DESCOBRE O CÓDIGO DO PRIMEIRO DECK DO DECOMP
def first_deck_decomp (idStudy):

    deckid = []
    mes_ano_deck = []
    nome_arquivo_deck = []
    lista_decks = getListOfDecks(idStudy)
    #lista_decks = resposta['Decks']
    print(lista_decks)
    for deck in lista_decks:
        if deck['Model'] == 'DECOMP': #and deck['FileName'] == 'DC202003-sem3.zip':
            deckid.append(deck['Id'])
            mes_ano_deck.append(str(deck['Month']) + '-' + str(deck['Year']))
            nome_arquivo_deck.append(deck['FileName'])
    primeiro_deck_decomp = min(deckid)
    #nome_arquivo_prim_deck_decomp =
    print('idStudy: ', idStudy, ' - deckid (prim. DECOMP): ', primeiro_deck_decomp)#, ' - nome_arquivo_deck: ', nome_arquivo_deck)
    return primeiro_deck_decomp


#DESCOBRE O CÓDIGO DO PRIMEIRO DECK DO NEWAVE
def first_deck_newave (idStudy):

    deckid = []
    mes_ano_deck = []
    nome_arquivo_deck = []
    lista_decks = getListOfDecks(idStudy)
    #lista_decks = resposta['Decks']

    for deck in lista_decks:
        if deck['Model'] == 'NEWAVE':
            deckid.append(deck['Id'])
            mes_ano_deck.append(str(deck['Month']) + '-' + str(deck['Year']))
            nome_arquivo_deck.append(deck['FileName'])
    primeiro_deck_newave = [min(deckid)]
    #nome_arquivo_prim_deck_decomp =
    print('idStudy: ', idStudy, ' - deckid (prim. DECOMP): ', primeiro_deck_newave)#, ' - nome_arquivo_deck: ', nome_arquivo_deck)
    return primeiro_deck_newave


#ENVIA ARQUIVOS DO GEVAZP PARA PRIMEIRO DECK DO DECOMP
def send_GEVAZP(caminho_pasta_dia, idStudy):
    fileName1 = 'MODIF.DAT'
    fileName2 = 'POSTOS.DAT'
    fileName3 = 'REGRAS.DAT'
    fileName4 = 'VAZOES.DAT'
    filename_GEVAZP = [fileName1, fileName2, fileName3, fileName4]

    pathToGVAZP = Path(caminho_pasta_dia + '\\03.GEVAZP')
    idDeck = first_deck_decomp(idStudy)

    for num in range(len(filename_GEVAZP)):
        pathToFile = os.path.join(pathToGVAZP, filename_GEVAZP[num])
        sendFileToDeck(idStudy, idDeck, pathToFile, filename_GEVAZP[num])
    print('Arquivos GEVAZP enviados! ID do Estudo: ', idStudy, ' ID do Deck: ', idDeck)

    # -----------------------------------------------------------------------------
    # Associate cuts | Reaproveitar (associar) cortes
    # -----------------------------------------------------------------------------

def cutAssociation(idStudy, destinationIds, sourceStudyId):
    api_function = '/api/prospectiveStudies/' + str(idStudy) + '/Associate'
    url = api_url_base + api_function
    listOfAssociation = []

    for deck in destinationIds:
        associationConfiguration = {}
        associationConfiguration['DestinationDeckId'] = deck
        associationConfiguration['SourceStudyId'] = sourceStudyId
        listOfAssociation.append(associationConfiguration)

    parameter = ''
    data = {
        "cortesAssociation": listOfAssociation,
    }

    print("Usando a seguinte configuracao do estudo: ", str(idStudy))
    print(data)

    postInAPI(token, api_function, parameter, data)

# -----------------------------------------------------------------------------
# Obter lista de servidores - necessário para adicionar um estudo em uma fila
# -----------------------------------------------------------------------------


def getListOfServers():
    api_function = '/api/Servers'
    url = api_url_base + api_function
    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }
    response  = requests.get(url, headers=headers, verify=verifyCertificate)
    print(response.status_code)
    lista_servidores = response.json()
    print(lista_servidores)
    return lista_servidores

def getIdOfServer(serverName):
    listOfServers = getListOfServers()

    for server in listOfServers:
        if server['Name'] == serverName:
            idServer = server['Id']
            return idServer

    return 0

# -----------------------------------------------------------------------------
# Get list of Spot Instances Types
# Obter lista de tipos de instâncias SPOT
# -----------------------------------------------------------------------------


def getListOfSpotInstancesTypes():
    api_function = '/api/Servers/SpotInstances'
    url = api_url_base + api_function
    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }
    response  = requests.get(url, headers=headers, verify=verifyCertificate)
    print(response.status_code)
    lista_servidores = response.json()
    print(lista_servidores)
    return lista_servidores

# -----------------------------------------------------------------------------
# Get list of Spot Instances Types    and choose one
# Obter lista de tipos de instâncias SPOT e escolher um
# -----------------------------------------------------------------------------


def getIdOfSpotInstancesType(serverType):
    listOfSpotInstances = getListOfSpotInstancesTypes()

    idSpotInstances = ''
    for spotInstances in listOfSpotInstances:
        if spotInstances['InstanceType'] == serverType:
            idSpotInstances = spotInstances['Id']
            return idSpotInstances

    return 0

## RODAR NWLISTOP ###################
def runNwlistop(idStudy, idDeck, spotInstanceType, pathToFile = '', fileName = '', idServer = 0):

    if pathToFile != '' and fileName != '':
        response = sendFileToAPI(token, '/api/prospectiveStudies/' + str(idStudy)
                             + '/RunNwlistop' + '?deckId=' + str(idDeck)
                             + '&spotInstanceType=' + str(spotInstanceType)
                             + '&serverId=' + str(idServer), pathToFile, fileName)
    else:
        parameter = ''
        data = ''
        response = postInAPI(token,'/api/prospectiveStudies/' + str(idStudy)
                             + '/RunNwlistop' + '?deckId=' + str(idDeck)
                             + '&spotInstanceType=' + str(spotInstanceType)
                             + '&serverId=' + str(idServer), parameter, data)

    print(response)

# -----------------------------------------------------------------------------
# Get list of Spot Instances Types    and choose one
# Obter lista de tipos de instâncias SPOT e escolher um
# -----------------------------------------------------------------------------

def escolher_sevidor(idStudy):
    lista_decks = getListOfDecks(idStudy)
    # print(json.dumps(lista_decks, indent=2))
    # Lista os decks e verifica se tem algum Newave
    contains_newave = False
    contains_multiple_prevs = True  # PENSAR COMO VERIFICAR SE TEM MAIS DE UM PREVS POR DECOMP
    contains_one_decomp = True
    for deck in lista_decks:
        if deck['Model'] == 'NEWAVE':
            contains_newave = True
    if contains_newave:
        serverType = 'c5.18xlarge'
    elif contains_multiple_prevs:
        serverType = 'm5.12xlarge'
    elif contains_one_decomp:
        serverType = 'm5.4xlarge'

    #lista = getListOfSpotInstancesTypes()
    idservidor = getIdOfSpotInstancesType(serverType)
    print('idservidor: ', idservidor, ' - nome do Servidor: ', serverType)
    return {'idservidor': idservidor, 'nome_servidor': serverType}

def executa_estudos (idStudy, serverType, ExecutionMode, InfeasibilityHandling, InfeasibilityHandlingSensibility, maxRestarts):
    api_function = '/api/prospectiveStudies/' + str(idStudy) + '/Run'
    url = api_url_base + api_function
    idQueue = 0
    ExecutionMode = 0  # Modo de execução(integer): 0 - Modo Pdrão, 1 - Consistência, 2 - Padrão + consistência
    InfeasibilityHandling = 3  # InfeasibilityHandling(integer): 0 - Parar estudo, 1 - Tratar inviabilidades, 2 - Ignorar inviabilidades, 3 - Tratar + Ignorar inviabilidades
    InfeasibilityHandlingSensibility = 3  # InfeasibilityHandlingSensibility(integer): 0 - Parar estudo, 1 - Tratar inviabilidades, 2 - Ignorar inviabilidades, 3 - Tratar + Ignorar inviabilidades
    maxRestarts = 10
    maxRestartsSensibility = 10
    data = {
        #"SpotInstanceType": serverType,
        "EphemeralInstanceType": serverType,
        "ExecutionMode": ExecutionMode,
        "InfeasibilityHandling": InfeasibilityHandling,
        "InfeasibilityHandlingSensibility": InfeasibilityHandlingSensibility,
        "MaxTreatmentRestarts": maxRestarts,
        "ServerPurchaseOption": 0,
        "MaxTreatmentRestartsSensibility":maxRestartsSensibility,
        "MaxExtraTreatmentRestarts":3
    }
    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }

    parameter = ''

    response = requests.post(url, headers=headers, params=parameter,
                             data=json.dumps(data), verify=verifyCertificate)
    print("A seguinte configuracao sera usada para iniciar a execucao o estudo: ", str(idStudy))
    print(data)

def download_estudos_finalizados ():
    caminho_arquivo_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\Criacao_Estudos.xlsx'
    arquivo_excel = load_workbook(caminho_arquivo_base)
    formulario = arquivo_excel['formulario']

    token = autenticar_prospec()
    estudos_pendentes = 0
    linha = 2
    caminho_rodada = formulario.cell(row=2, column=11).value
    # Loop que verifica linha a linha do arquivo em excel a vai baixando os estudos pra quantas linhas tiverem
    while (formulario.cell(row=linha, column=3).value != 0 and formulario.cell(row=linha, column=3).value != None):
        inicio_pasta = formulario.cell(row=linha, column=2).value
        idStudy = formulario.cell(row=linha, column=3).value
        modelo = formulario.cell(row=linha, column=13).value
        status_dowload = formulario.cell(row=linha, column=15).value

        if status_dowload != 'Download realizado':
            status = GetStatusOfStudy(token, idStudy)
            if status == 'Finished':
                pathdownload = os.path.join(formulario.cell(row=linha, column=11).value, '04.Download Estudos')
                nome_arquivo = inicio_pasta + '_' + str(idStudy) + '.zip'
                download_compilado(token, idStudy, pathdownload, nome_arquivo)
                formulario.cell(row=linha, column=15).value = 'Download realizado'
                formulario.cell(row=linha, column=16).value = datetime.datetime.now()
                print('Download realizado para o estudo ', idStudy, ' - ', inicio_pasta, ' - modelo ', modelo)
            else:
                estudos_pendentes = estudos_pendentes + 1
                formulario.cell(row=linha, column=15).value = status
                formulario.cell(row=linha, column=16).value = datetime.datetime.now()
                print('Dowload não realizado - IdStudy: ', idStudy, 'Status: ', status)
        else:
            print('Estudo com download já realizado anteriormente - IdStudy: ', idStudy)
        arquivo_excel.save(caminho_arquivo_base)
        linha = linha + 1

    arquivo_excel.save(caminho_arquivo_base)
    arquivo_excel.close()
    return {"Estudos pendente": estudos_pendentes, "Caminho rodada": caminho_rodada}


def download_monitoramento_estudos_pendentes():
    #resposta = download_estudos_finalizados()
    #qtd_estudos_pendentes = resposta['Estudos pendente']
    urllib3.disable_warnings((urllib3.exceptions.InsecureRequestWarning))
    caminho_rede = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\Boletim_Diario.xlsm'
    excel_caminho_local = r'C:\Users\fernando.fidalgo\Desktop\Docs_Fidalgo\10. Eneva Com\Boletim_Diario.xlsm'

    #verifica se todos downloads foram feitos, caso sim, executa macro, caso não, verifica de novo em 30 minutos
    while True:
        print('\n' ,'Hora de início da verificação: ', datetime.datetime.now())
        resposta = download_estudos_finalizados()
        qtd_estudos_pendentes = resposta['Estudos pendente']
        caminho_da_rodada = resposta['Caminho rodada']
        if qtd_estudos_pendentes == 0:
            shutil.copy2(Path(caminho_rede), Path(excel_caminho_local))
            print('Arquivo original copiado da rede para execução local, caminho de origem: ', caminho_rede)
            time.sleep(30)
            caminho_novo_arquivo = caminho_da_rodada + excel_caminho_local[(excel_caminho_local.rfind('\\')):]
            roda_vba (excel_caminho_local, 1, 'CompiladorResultados_python', caminho_novo_arquivo)
            shutil.copy2(Path(excel_caminho_local) , Path(caminho_rede))
            print('Arquivo copiado para a rede no caminho: ', caminho_rede)
            shutil.copy2(Path(excel_caminho_local), Path(caminho_novo_arquivo))
            print('Arquivo copiado para a pasta da rodada no caminho: ', caminho_novo_arquivo)
            quit()
        else:
            time.sleep(60*20)

#roda o VBA de um excel e salva no mesmo local
def roda_vba (excel_caminho, num_modulo, nome_macro, caminho_novo_arquivo):
    #Essa função usa o caminho do excel sem a função Path
    #caminho para arquivo em excel com a Macro
    #A macro precisa estar marcada como Public
    #excel_caminho = r'C:\Users\fernando.fidalgo\Desktop\Docs_Fidalgo\10. Eneva Com\Boletim_ Diario - Rv2.xlsm'
    #excel_temp = excel_caminho.replace('.xlsm', '_temp.xlsm')
    caracter = excel_caminho.rfind('\\') + 1
    nome_arquivo = excel_caminho[caracter:]
    caminho_pasta_rede = caminho_novo_arquivo[:(caminho_novo_arquivo.rfind('\\'))]
    # DispatchEx is required in the newest versions of Python.
    excel_macro = client.DispatchEx("Excel.Application")
    wb = excel_macro.Workbooks.Open(Filename = excel_caminho)
    ws = wb.Worksheets('Resumo do Caso')
    excel_macro.Visible = True
    time.sleep(5)
    #salva o caminho do pasta da rodada na rede no próprio arquivo para que a macro leia os estudos da rodada
    ws.Cells(4, 35).Value = caminho_pasta_rede + r'\04.Download Estudos'
    macro = 'Módulo' + str(num_modulo) + '.' + nome_macro
    print('Iniciando Execução da Macro: ' + macro)
    wb.Application.Run("\'" + nome_arquivo + "\'" + '!' + macro)
    print('Macro executada com sucesso')
    wb.Save() #, FileFormat = 52
    print('Arquivo Salvo Localmente: ', excel_caminho)
    time.sleep(5)
    excel_macro.Application.Quit()
    del excel_macro
    #shutil.move(Path(excel_temp), Path(caminho_novo_arquivo))
    #print('Arquivo salvo na pasta de destino: ', caminho_novo_arquivo[(caminho_novo_arquivo.rfind('\\') + 1):])

def download_monitoramento_estudos_pendentes_sem_macro():
    # resposta = download_estudos_finalizados()
    # qtd_estudos_pendentes = resposta['Estudos pendente']
    urllib3.disable_warnings((urllib3.exceptions.InsecureRequestWarning))
    caminho_rede = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\Boletim_Diario.xlsm'
    excel_caminho_local = r'C:\Users\fernando.fidalgo\Desktop\Docs_Fidalgo\10. Eneva Com\Boletim_Diario.xlsm'

    # verifica se todos downloads foram feitos, caso sim, executa macro, caso não, verifica de novo em 30 minutos
    while True:
        print('\n', 'Hora de início da verificação: ', datetime.datetime.now())
        resposta = download_estudos_finalizados()
        qtd_estudos_pendentes = resposta['Estudos pendente']
        caminho_da_rodada = resposta['Caminho rodada']
        if qtd_estudos_pendentes == 0:
           # shutil.copy2(Path(caminho_rede), Path(excel_caminho_local))
           # print('Arquivo original copiado da rede para execução local, caminho de origem: ', caminho_rede)
           # time.sleep(30)
           # caminho_novo_arquivo = caminho_da_rodada + excel_caminho_local[(excel_caminho_local.rfind('\\')):]
           # roda_vba(excel_caminho_local, 1, 'CompiladorResultados_python', caminho_novo_arquivo)
           # shutil.copy2(Path(excel_caminho_local), Path(caminho_rede))
           # print('Arquivo copiado para a rede no caminho: ', caminho_rede)
           # shutil.copy2(Path(excel_caminho_local), Path(caminho_novo_arquivo))
           # print('Arquivo copiado para a pasta da rodada no caminho: ', caminho_novo_arquivo)
            print('Todos estudos baixados')
            quit()
        else:
            time.sleep(60 * 60)

def download_estudos_nao_finalizados (**kwargs):
    if kwargs.get('caminho_arquivo_estudos'):
        caminho_arquivo_base = kwargs.get('caminho_arquivo_estudos')
    else:
        caminho_arquivo_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\Criacao_Estudos.xlsx'
    arquivo_excel = load_workbook(caminho_arquivo_base)
    formulario = arquivo_excel['formulario']

    token = autenticar_prospec()
    estudos_pendentes = 0
    linha = 2
    caminho_rodada = formulario.cell(row=2, column=11).value
    # Loop que verifica linha a linha do arquivo em excel a vai baixando os estudos pra quantas linhas tiverem
    while (formulario.cell(row=linha, column=3).value != 0 and formulario.cell(row=linha, column=3).value != None):
        inicio_pasta = formulario.cell(row=linha, column=2).value
        idStudy = formulario.cell(row=linha, column=3).value
        modelo = formulario.cell(row=linha, column=13).value
        status_dowload = formulario.cell(row=linha, column=15).value

        if status_dowload != 'Download realizado':
            status = GetStatusOfStudy(token, idStudy)

            pathdownload = os.path.join(formulario.cell(row=linha, column=11).value, '04.Download Estudos')
            nome_arquivo = inicio_pasta + '_' + str(idStudy) + '.zip'
            download_compilado(token, idStudy, pathdownload, nome_arquivo)
            formulario.cell(row=linha, column=15).value = 'Download realizado'
            formulario.cell(row=linha, column=16).value = datetime.datetime.now()
            print('Download realizado para o estudo ', idStudy, ' - ', inicio_pasta, ' - modelo ', modelo)

        else:
            print('Estudo com download já realizado anteriormente - IdStudy: ', idStudy)
        arquivo_excel.save(caminho_arquivo_base)
        linha = linha + 1

    arquivo_excel.save(caminho_arquivo_base)
    arquivo_excel.close()
    return {"Estudos pendente": estudos_pendentes, "Caminho rodada": caminho_rodada}


def envia_email_python(assunto, corpo, caminho_anexo, anexo, destinatario):
    login_email = 'middle@eneva.com.br'
    password = 'T5Yx*CCuRM8@'
    smtpsrv = "smtp.office365.com"
    smtpserver = smtplib.SMTP(smtpsrv, 587)
    msg = EmailMessage()
    msg['Subject'] = assunto
    msg['From'] = 'middle@eneva.com.br'
    msg['To'] = destinatario #'fernando.fidalgo@eneva.com.br'#; renata.hunder@eneva.com.br' #comercializacao@eneva.com.br; camila.schoti@eneva.com.br; caio.picanco@eneva.com.br
    msg.set_content(corpo)
    if caminho_anexo != None and caminho_anexo != '':
        with open(caminho_anexo, 'rb') as content_file:
            content = content_file.read()
            msg.add_attachment(content, maintype='application/pdf', subtype='pdf',
                               filename=anexo)

    smtpserver.ehlo()
    smtpserver.starttls()
    smtpserver.login(login_email, password)
    smtpserver.send_message(msg)
    smtpserver.close()
    print('E-mail enviado com sucesso!')

def envia_email_pdf_gif (assunto, corpo, caminho_pdf, caminho_gif, destinatario):
    login_email = 'middle@eneva.com.br'
    password = 'T5Yx*CCuRM8@'
    smtpsrv = "smtp.office365.com"
    smtpserver = smtplib.SMTP(smtpsrv, 587)
    msg = EmailMessage()
    msg['Subject'] = assunto
    msg['From'] = 'middle@eneva.com.br' #'middle@eneva.com.br'
    msg['To'] = destinatario #'todosenevacom@eneva.com.br' #'comercializacao@eneva.com.br; camila.schoti@eneva.com.br; haruki.moraes@eneva.com.br; caio.picanco@eneva.com.br'
    msg.set_content(corpo)
    #anexando arquivo PDF
    posicao = len(caminho_pdf) - caminho_pdf.rfind('\\') - 1
    nome_arquivo = caminho_pdf[-posicao:]
    with open(caminho_pdf, 'rb') as content_file:
        content = content_file.read()
        msg.add_attachment(content, maintype='application/pdf', subtype='pdf',
                           filename=nome_arquivo)
    for file in caminho_gif:
        posicao = len(file) - file.rfind('\\') - 1
        nome_arquivo = file[-posicao:]

        with open(file, 'rb') as content_file:
            content = content_file.read()
            msg.add_attachment(content, maintype='image', subtype=imghdr.what(None, content), filename=nome_arquivo)

    smtpserver.ehlo()
    smtpserver.starttls()
    smtpserver.login(login_email, password)
    smtpserver.send_message(msg)
    smtpserver.close()
    print('E-mail enviado com sucesso!')

def monitor_estudos ():
    lista_estudos = [3032, 3033, 3034, 3035, 3036, 3037, 3038]
    token = autenticar_prospec()
    envia_email = False
    corpo_email = 'ATENÇÃO!!\n\nAlgum caso apresenta erro no PROSPEC. Abaixo lista de casos em andamento:\n\n\n'
    for studyid in lista_estudos:
        status = GetStatusOfStudy(token, studyid)
        if status == 'Failed':
            envia_email = True

        corpo_email = corpo_email + 'Estudo ' + str(studyid) + ' - Status: ' + status + '\n'
    if envia_email:
        assunto = 'RELATÓRIO DE FALHAS NO PROSPEC'
        qtd_reqs = conta_requisicoes(token)
        corpo_email = corpo_email + '\nQUANTIDADE DE REQUISIÇÕES UTILIZADAS: ' + str(qtd_reqs)
        envia_email_python(assunto, corpo_email, caminho_anexo='', anexo='')
    else:
        agora = datetime.datetime.now()
        qtd_reqs = conta_requisicoes(token)
        print('Não houve casos com erro em: ' + str(agora))
        print('Quantidade de Requisições utizadas: ', qtd_reqs)

def cria_rodada (caminho_arquivo_criacao_estudo, **kwargs):
    #caminho_arquivo_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\Criacao_Estudos.xlsm'
    if kwargs.get('Diario'):
        rel_diario = kwargs.get('Diario')
        diario = True
        if rel_diario == 'Definitivo':
            linha = 17
            print('Iniciando execução do relatório diário definitivo.')
        elif rel_diario == 'Preliminar':
            linha = 7
            print('Iniciando execução do relatório diário preliminar.')
    else:
        rel_diario = False
        linha = 7
    executa_rodada = True
    if kwargs.get('Executa_rodada'):
        if kwargs.get('Executa_rodada') == 'False':
            executa_rodada = False
            print('Executar rodada:', executa_rodada)
            print('Rodada não será executada')
        else:
            print('Rodada será executada')
    if kwargs.get('Caminho_rodada'):
        caminho_rodada = os.path.realpath(kwargs.get('Caminho_rodada'))
        print('caminho da rodada:', caminho_rodada)
    else:
        caminho_rodada = False
        print('Sem caminho da rodada declarado')
    token = autenticar_prospec()

    # abre arquivo excel com as informações dos estudos
    arquivo_excel = client.DispatchEx("Excel.Application")
    arquivo_excel.Visible = True
    arquivo_excel.DisplayAlerts = False
    wb = arquivo_excel.Workbooks.Open(Filename=caminho_arquivo_criacao_estudo)
    formulario = wb.Worksheets('formulario')
    aba_other_NW = wb.Worksheets('otherNewavefiles')
    aba_volume = wb.Worksheets('volume')
    aba_cortes = wb.Worksheets('associar_cortes')
    aba_estagios = wb.Worksheets('multiple_stages')
    aba_aux = wb.Worksheets('aux')
    print('arquivo excel aberto com sucesso')
    if caminho_rodada:
        aba_aux.Cells(2,2).Value = caminho_rodada

    # Loop que verifica linha a linha do arquivo em excel a vai criando os estudos pra quantas linhas tiverem
    while (formulario.Cells(linha, 6).Value != 0 and formulario.Cells(linha, 6).Value != None and formulario.Cells(
            linha, 6).Value != ""):  #and linha <=9:

        qtd_reqs = conta_requisicoes(token)
        print('Quantidade de requisições utilizadas: ', qtd_reqs)

        # apaga dados da última rodada (caso exista
        formulario.Cells(linha, 2).ClearContents()
        formulario.Cells(linha, 3).ClearContents()
        formulario.Cells(linha, 4).ClearContents()
        formulario.Cells(linha, 5).ClearContents()


        # copia as variáveis de criação do Estudo do Excel
        ordem_estudo = formulario.Cells(linha, 6).Value
        nome_estudo = formulario.Cells(linha, 7).Value

        fileNameNW = formulario.Cells(linha, 8).Value  # nome do arquivo do NEWAVE
        fileNameDC = formulario.Cells(linha, 9).Value  # nome do arquivo do DECOMP
        fileNameDP = formulario.Cells(linha, 10).Value  # nome do arquivo de Dados Prospectivos
        mes_ano_inicial = formulario.Cells(linha,
                                           11).Value  # Data inicial do estudo (o imput é a data, mas só é utilizado o mês e o ano)
        initialMonth = mes_ano_inicial.month  # mês incial da data informada como início do estudo
        initialYear = mes_ano_inicial.year  # ano incial da data informada como início do estudo

        duration = int(formulario.Cells(linha, 12).Value)  # duração em meses do estudo a ser executado
        meses_fonte_pluvia = formulario.Cells(linha, 13).Value
        associar_cortes = formulario.Cells(linha, 23).Value  # variável que define a associação de cortes do caso
        caminho_pastas = formulario.Cells(linha,
                                          14).Value  # caminho base da rodada onde os dados dos estudos estão armazenados
        print('caminho pasta:', caminho_pastas)
        # Caminho do arquivo de NEWAVE a ser enviado ao estudo
        pathToFileNW = Path(caminho_pastas + '\\01.Decks\\' + fileNameNW)
        # Caminho do arquivo de DECOMP a ser enviado ao estudo
        pathToFileDC = Path(caminho_pastas + '\\01.Decks\\' + fileNameDC)

        #ajusta prevs na pasta da rodada
        if meses_fonte_pluvia is not None and meses_fonte_pluvia:
            #apaga as prevs não desejadas no estudo no nome da pasta de prevs pluvia informado na planilha do intersemanal
            apaga_prevs_nao_mapeadas(meses_fonte_pluvia,caminho_pastas + r'\02.Prevs' + '\\' + formulario.Cells(linha, 16).Value )

        # função que cria o estudo
        idStudy = cria_estudo(nome_estudo)  # Cria o estudo e salva o código do estudo na variável
        formulario.Cells(linha, 3).Value = int(idStudy)  # anota o número do estudo criado na planilha do prospec
        formulario.Cells(linha, 3).NumberFormat = '0'  # formata a célula como número inteiro sem casas decimais
        # anota data e hora da criação do estudo na planilha do Prospec, está anotando ao contrário devido a incompatibilidade de win32com
        formulario.Cells(linha, 2).Value = datetime.datetime.now().strftime('%m/%d/%Y  %H:%M:%S')

        # envia os arquivos para criar os decks no Prospec
        sendFileToStudy(idStudy, pathToFileDC, fileNameDC)  # envia o arquivo do DECOMP para o estudo
        sendFileToStudy(idStudy, pathToFileNW, fileNameNW)  # envia o arquivo do NEWAVE para o estudo
        if fileNameDP != None and fileNameDP != 0 and fileNameDP != '':  # Verifica se tem planilha de DADOS PROSPECTIVOS para ser enviada ao estudo, caso esxista, evia
            pathToFileDP = Path(caminho_pastas + '\\01.Decks\\' + fileNameDP)
            sendFileToStudy(idStudy, pathToFileDP, fileNameDP)
        else:
            fileNameDP = ''

        # cria listas de variáveis de ano, mês, outros NEWAVES, Multiplestages, Multiple revision, volume e associação de cortes
        month = []
        year = []
        otherNewaveFiles = []
        multipleStages = []
        multipleRevision = []
        volume = []
        cortes = []
        # cria os vetores de mês e ano de acordo com a duração do estudo.
        # dur começa no 0
        col = 6
        coluna_NW = 0
        coluna_vol = 0
        coluna_cortes = 0
        coluna_stag = 0
        # verifica se é necessário criar lista do newave, ou seja se haverá envio de newave adicional
        if formulario.Cells(linha, 18).Value:
            while aba_other_NW.Cells(6, col).Value:
                mes_other_NW = aba_other_NW.Cells(6, col).Value
                if (mes_other_NW.month == mes_ano_inicial.month) and (mes_other_NW.year == mes_ano_inicial.year):
                    coluna_NW = col
                col = col + 1

        # verifica se é necessário criar lista de volume, ou seja se haverá envio de volumes para algum deck
        col = 6
        if formulario.Cells(linha, 19).Value:
            while aba_volume.Cells(6, col).Value:
                mes_vol = aba_volume.Cells(6, col).Value
                if (mes_vol.month == mes_ano_inicial.month) and (mes_vol.year == mes_ano_inicial.year):
                    coluna_vol = col
                col = col + 1

        # verifica se é necessário criar lista de associação de cortes, ou seja se haverá associação de cortes em algum mês do estudo
        col = 6
        if associar_cortes:
            while aba_cortes.Cells(6, col).Value:
                mes_cortes = aba_cortes.Cells(6, col).Value
                if (mes_cortes.month == mes_ano_inicial.month) and (mes_cortes.year == mes_ano_inicial.year):
                    coluna_cortes = col
                col = col + 1

        # verifica se é necessário criar lista de multiplos estágios para informar se será necessária a do mês ou apenas RV0
        # se a coluna na aba formulário não for marcada com alguma informação, será setado com multiplas estágios apenas no primeiro mês
        # se a coluna na aba formulário for marcada com alguma informação, então ele habilita a marcação na aba multiple_stages
        # na aba multiple_stages deve ser marcado com 1 ou true nos meses onde se quer essa abertura
        col = 6
        if formulario.Cells(linha,
                            15).Value == 'X':  # verifica a configuração de abertura em até três meses ou customizado caso preenchico com "x"
            while aba_estagios.Cells(6, col).Value:
                mes_stag = aba_estagios.Cells(6, col).Value
                if (mes_stag.month == mes_ano_inicial.month) and (mes_stag.year == mes_ano_inicial.year):
                    coluna_stag = col
                col = col + 1

        # primeiro dur do for é zero
        qtd_abert_meses = 1
        for dur in range(duration):
            # Adiciona um mês ao mês incial
            proxima_data = mes_ano_inicial + relativedelta(months=dur)
            month.append(proxima_data.month)
            year.append(proxima_data.year)
            # Cria a lista de Newaves adicionais do Estudo, caso a coluna 17 esteja preenchida
            if coluna_NW == 0:
                otherNewaveFiles.append('')
            else:
                otherNewaveFiles.append(aba_other_NW.Cells(linha, (coluna_NW + dur)).Value)
                if aba_other_NW.Cells(linha, (coluna_NW + dur)).Value != None:
                    filenameotherNW = aba_other_NW.Cells(linha, (coluna_NW + dur)).Value
                    pathToFileotherNW = Path(caminho_pastas + '\\01.Decks\\' + filenameotherNW)
                    sendFileToStudy(idStudy, pathToFileotherNW, filenameotherNW)
                    print('Deck adicional (', filenameotherNW, ')de NEWAVE enviado para o estudo ', idStudy)
            # Cria a lista de arquivos de volumes adicionais do Estudo, caso a coluna 17 esteja preenchida
            if coluna_vol == 0:
                volume = []
            else:
                volume.append(aba_volume.Cells(linha, (coluna_vol + dur)).Value)

            if coluna_cortes == 0:
                cortes = []
            else:
                corte_mes = aba_cortes.Cells(linha, (coluna_cortes + dur)).Value
                if corte_mes == None or corte_mes == '':
                    cortes.append(corte_mes)
                elif corte_mes == 'Deck Newave':
                    cortes.append(int(formulario.Cells(3, 3).Value))
                elif str(corte_mes)[:1] == 'E':
                    corte_mes_string = str(corte_mes)
                    caso = int(corte_mes_string.replace('E', ''))
                    corte_mes = int(aba_cortes.Cells(6 + caso, 3).Value)
                    cortes.append(corte_mes)
                else:
                    cortes.append(int(corte_mes))

            # CRIA LISTA DE MULTIPLAS REVISÕES
            if coluna_stag == 0:
                if formulario.Cells(linha,
                                    15).Value != None:  # se a coluna que diz a quantidade de meses aberto por semana for preenchida com algo que não seja o X, significa que será aberto por semanas a quantidade de meses que for preenchido

                    if qtd_abert_meses <= int(formulario.Cells(linha, 15).Value):
                        multipleStages.append(True)
                        multipleRevision.append(True)
                    else:
                        multipleStages.append(False)
                        multipleRevision.append(False)
                    qtd_abert_meses = qtd_abert_meses + 1


            else:
                if aba_estagios.Cells(linha, (coluna_stag + dur)).Value == '' or aba_estagios.Cells(linha, (
                        coluna_stag + dur)).Value == None or aba_estagios.Cells(linha, (
                        coluna_stag + dur)).Value == 'Mensal':  # se o campo estiver em branco ou preenchido com "Mensal", ele abre o mês
                    valor_multp_stages = False
                    valor_multp_revision = False
                elif aba_estagios.Cells(linha, (
                        coluna_stag + dur)).Value == 'Semanal':  # se o campo estiver preenchido com "Semanal", ele abre por semana
                    valor_multp_stages = True
                    valor_multp_revision = True
                elif aba_estagios.Cells(linha, (
                        coluna_stag + dur)).Value == 'RV0':  # se o campo estiver preenchido com "RV0", ele abre para a primeira semana do mês
                    valor_multp_stages = True
                    valor_multp_revision = False
                multipleStages.append(valor_multp_stages)
                multipleRevision.append(valor_multp_revision)

        firstNewaveFile = fileNameNW
        decompFile = fileNameDC
        spreadsheetFile = fileNameDP
        tags = [formulario.Cells(linha, 6).Value]

        generateStudyDecks(idStudy, initialYear, initialMonth, duration, month,
                           year, multipleStages, multipleRevision, firstNewaveFile,
                           otherNewaveFiles, decompFile, spreadsheetFile, tags)

        # ENVIA OS ARQUIVOS DO GEVAZP
        send_GEVAZP(caminho_pastas, idStudy)

        # ENVIA OS ARQUIVOS DE PREVS PARA O ESTUDO
        modelo = formulario.Cells(linha, 16).Value
        prevs_eneva = formulario.Cells(linha, 17).Value
        print(prevs_eneva)
        caminho_prevs = caminho_pastas + r'\02.Prevs'
        pathToAllPrevs_modelo = caminho_prevs + '\\' + modelo
        if prevs_eneva == None:
            pathToAllPrevs_eneva = 'VAZIO'
        else:
            pathToAllPrevs_eneva = caminho_prevs + '\\' + prevs_eneva
        Matriz1 = formulario.Cells(linha, 20).Value
        Matriz2 = formulario.Cells(linha, 21).Value
        Matriz3 = formulario.Cells(linha, 22).Value
        matrizes = [Matriz1, Matriz2, Matriz3]
        # envia os arquivos de Matriz quando as colunas 19, 20 e 21 estão prenchidas com "Enviar"
        for n in range(len(matrizes)):
            if matrizes[n] == 'Enviar':
                matrizes[n] = 'Matriz' + str(n + 1)
            else:
                matrizes[n] == ''

        if prevs_eneva == None or prevs_eneva == '':
            print('estudo: ', idStudy, ' - não foram eviadas prevs-eneva (sensibilidades).')
        else:
            sendAllPrevsToStudy(idStudy, pathToAllPrevs_eneva)
            print('estudo: ', idStudy, ' - modelo: ', prevs_eneva, 'caminho de envio: ', pathToAllPrevs_eneva)

        sendAllPrevsToStudy(idStudy, pathToAllPrevs_modelo)
        print('estudo: ', idStudy, ' - modelo: ', modelo, 'caminho de envio: ', pathToAllPrevs_modelo)

        for n in range(len(matrizes)):
            if matrizes[n] == None:
                print('Matriz não enviada: Matriz', n + 1)
            else:
                pathToAllPrevs_matriz = caminho_prevs + '\\' + matrizes[n]
                sendAllPrevsToStudy(idStudy, pathToAllPrevs_matriz)
                print('Matriz enviada: Matriz', n + 1)
        # Envia os arquivos de volumes quando a coluna 18 estiver preenchida
        if coluna_vol != 0:

            lista_decks = getListOfDecks(idStudy)
            # Itera a lista de volumes e usa o índice para pegar o mês e ano do volume e comparar com o mês e ano do deck
            # necessário comparar o tamanho da lista volume com o tamanho da lista mes e tamanho da lista ano
            # ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!
            for volu in range(len(volume)):
                print('Id Lista Volume:', volu, ' - mês/ano:', month[volu], '/', year[volu])
                if volume[volu] != None:
                    print('nome volume: ', volume[volu])
                    pathTovolume = caminho_pastas + '\\01.Decks\\' + volume[volu]
                    print(pathTovolume)
                    # RODAR TODOS OS DECKS E BUSCAR O DECK QUE  MÊS FOR IGUAL AO MÊS ONDE O VOLUME NÃO FOR VAZIO
                    for deck in lista_decks:
                        if month[volu] == deck['Month'] and year[volu] == deck['Year'] and deck[
                            'SensibilityInfo'] == 'Original':
                            sendFileToDeck(idStudy, deck['Id'], pathTovolume, volume[volu])
                            print('Arquivo de volume enviado para o deck ', deck['Id'], ' - ', deck['Model'],
                                  ' - mês/ano: ', deck['Month'], '/', deck['Year'])
                else:
                    print('Sem envio de volume para o ID da lista acima')
        else:
            print('Não houve envio de arquivo de volumes para o estudo: ', idStudy)

        # Associar Corte ao primeiro deck Newave do estudo
        if coluna_cortes != 0:

            lista_decks = getListOfDecks(idStudy)
            # Itera a lista de cortes e usa o índice para pegar o mês e ano do corte e comparar com o mês e ano do deck
            # necessário comparar o tamanho da lista volume com o tamanho da lista mes e tamanho da lista ano
            # ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!ATENÇÃO!
            for cort in range(len(cortes)):
                print('Posição na Lista Cortes:', cort, ' - mês/ano:', month[cort], '/', year[cort])
                if cortes[cort] != None:
                    print('Código do Estudo a qual corte será associado: ', cortes[cort])

                    # RODAR TODOS OS DECKS E BUSCAR O DECK QUE  MÊS FOR IGUAL AO MÊS ONDE O CORTE NÃO FOR VAZIO
                    for deck in lista_decks:
                        if deck['Model'] == 'NEWAVE' and month[cort] == deck['Month'] and year[cort] == deck['Year'] and \
                                deck['SensibilityInfo'] == 'Original':
                            # ASSOCIAR CORTE AQUI
                            print('Cortes do estudo ', idStudy, 'associados ao deck do', deck['Model'], ' do mês/ano',
                                  deck['Month'], '/', deck['Year'], '. Id do Deck: ', deck['Id'])
                            destinationIds = [deck['Id']]
                            cutAssociation(idStudy, destinationIds, cortes[cort])
                            print('Cortes do estudo ', idStudy, 'associados ao deck do', deck['Model'], ' do mês/ano',
                                  deck['Month'], '/', deck['Year'], '. Id do Deck: ', deck['Id'])
                else:
                    print('No estudo', idStudy, ' no deck de ', month[cort], '/', year[cort],
                          'não houve associação de cortes')
        else:
            print('Não houve nenhuma associação para o seguinte estudo: ', idStudy)

        if kwargs.get('Envia_volume_UHE'):
            if kwargs.get('Envia_volume_UHE') == 'True':
                print('Será enviado arquivo de volume para a semana seguinte.')
                idDeck = deck_decomp_semana_seguinte(idStudy)
                nome_arquivo_UH = 'volume_uhe.csv'

                caminho_arquivo_UH = Path(caminho_pastas + '\\01.Decks\\' + nome_arquivo_UH)
                sendFileToDeck(idStudy, idDeck, caminho_arquivo_UH, nome_arquivo_UH)



        servidor = escolher_sevidor(idStudy)
        idservidor = servidor['idservidor']
        serverType = servidor['nome_servidor']
        idQueue = 0
        ####ALTERADO de 2 para 3
        ExecutionMode = 0  # Modo de execução(integer): 0 - Modo Pdrão, 1 - Consistência, 2 - Padrão + consistência
        InfeasibilityHandling = 3  # InfeasibilityHandling(integer): 0 - Parar estudo, 1 - Tratar inviabilidades, 2 - Ignorar inviabilidades, 3 - Tratar + Ignorar inviabilidades
        InfeasibilityHandlingSensibility = 3  # InfeasibilityHandlingSensibility(integer): 0 - Parar estudo, 1 - Tratar inviabilidades, 2 - Ignorar inviabilidades, 3 - Tratar + Ignorar inviabilidades
        maxRestarts = 10

        if executa_rodada:
            executa_estudos(idStudy, serverType, ExecutionMode, InfeasibilityHandling, InfeasibilityHandlingSensibility, maxRestarts)
        else:
            print('Rodada criada, mas não executada.')

        wb.Save()
        time.sleep(10)
        linha = linha + 1
    wb.Close(False)
    arquivo_excel.Application.Quit()

    if rel_diario:
        nome_arquivo_backup = 'Criacao_Estudos_diario.xlsm'
    else:
        nome_arquivo_backup = 'Criacao_Estudos.xlsm'
#    shutil.copy2(Path(caminho_arquivo_criacao_estudo), Path.joinpath(Path(caminho_pastas), nome_arquivo_backup))
    qtd_reqs = conta_requisicoes(token)
    print('Quantidade de requisições utilizadas: ', qtd_reqs)

def download_estudos_nao_finalizados_novo_arquivo(**kwargs):
    if kwargs.get('caminho_arquivo_estudos'):
        caminho_arquivo_base = kwargs.get('caminho_arquivo_estudos')
        tipo_estudo = 'INTERSEMANAL'
    else:
        caminho_arquivo_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\06.Semanal\Criacao_Estudos.xlsm'
        tipo_estudo = 'SEMANAL'
    #caminho_arquivo_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\Criacao_Estudos.xlsm' #caminho do arquivo base onde estão registrados os casos
    aplicacao_excel = client.DispatchEx("Excel.Application")
    aplicacao_excel.Visible = True
    arquivo_excel = aplicacao_excel.Workbooks.Open(Filename =caminho_arquivo_base)
    formulario = arquivo_excel.Worksheets('formulario')
    #try:
    token = autenticar_prospec()
    #except:
        ##arquivo_excel.DisplayAlerts = False
        #del arquivo_excel
        #arquivo_excel.Application.Quit()
        #print('Erro de login no prospec')
    print('Quantidade de requisições utilizadas: ', conta_requisicoes(token))
    estudos_pendentes = 0

    linha = 7 #7 - linha do primeiro caso a ser baixado
    caminho_rodada = formulario.Cells(linha, 14).Value
    # Loop que verifica linha a linha do arquivo em excel a vai baixando os estudos pra quantas linhas tiverem preenchidas
    #esse loop usa como base o preenchimento ou não da coluna do nome de arquivo newave
    while (formulario.Cells(linha,8).Value != 0 and formulario.Cells(linha, 8).Value != None and formulario.Cells(linha, 8).Value != ''):

        caminho_rodada = formulario.Cells(linha, 14).Value #caminho da rede onde a rodada será armazenada
        inicio_pasta = formulario.Cells(linha, 6).Value #Salva a ordem do estudo (tag no prospec). E1, E2 etc
        idStudy = int(formulario.Cells(linha, 3).Value) #salva em variável o StudyID do estudo que será baixado
        modelo = formulario.Cells(linha, 16).Value # Salva o modelo/mapa usado na criação do estudo
        status_download = formulario.Cells(linha, 4).Value #salva em variável o status de download anterior (para verirficar se já foi feito download ou não)
        data_criacao_estudo = formulario.Cells(linha, 2).Value
        #data_criacao_estudo = data_criacao_estudo.strftime('%d/%m/%Y')
        if status_download != 'Download realizado': #verifica se na planilha está sinalizado como download realizado, caso não tenha sido, faz o download, caso tenha sido realizado, nõa faz nada e passa para o próximo
            status = GetStatusOfStudy(token, idStudy) #pega status do estudo

            pathdownload = os.path.join(caminho_rodada, '04.Download Estudos')
            nome_arquivo = inicio_pasta + '_' + str(idStudy) + '.zip' #salva na variável o nome do arquivo zipado que será criado
            download_compilado(token, idStudy, pathdownload, nome_arquivo) #função que faz o download do estudo
            formulario.Cells(linha, 4).Value = 'Download realizado' #escreve no excel o status como download realizado
            formulario.Cells(linha, 5).Value = datetime.datetime.now().strftime('%m/%d/%Y  %H:%M:%S') #escreve no excel a data e hora do dowload
            print('Download realizado para o estudo ', idStudy, ' - ', inicio_pasta, ' - modelo ', modelo)
            fonte_pluvia = formulario.Cells(linha, 16).Value
            cenario_eneva = formulario.Cells(linha, 17).Value
            observacao = formulario.Cells(linha, 24).Value
            arquiva_estudos_intersemanal_semanal(idStudy, data_criacao_estudo, os.path.join(pathdownload, nome_arquivo) , tipo_estudo, fonte_pluvia, cenario_eneva, observacao)
        else:
            print('Estudo com download já realizado anteriormente - IdStudy: ', idStudy)
        arquivo_excel.Save() #salva o arquivo excel que tem os estudos a cada novo estudo baixado, isso faz com que caso dê algum erro, ele sempre tem o status do últmo download salvo
        linha = linha + 1

    arquivo_excel.Save() #salva o escel uma última vez
    arquivo_excel.Application.Quit() #fecha o escel
    print('Quantidade de requisições utilizadas: ', conta_requisicoes(token))
    return {"Estudos pendente": estudos_pendentes, "Caminho rodada": caminho_rodada}

def download_estudos_nao_finalizados_novo_arquivo_preco_medio_auto(pld_semanal, **kwargs):
    if kwargs.get('caminho_arquivo_estudos'):
        caminho_arquivo_base = kwargs.get('caminho_arquivo_estudos')
        tipo_estudo = 'INTERSEMANAL'
    else:
        caminho_arquivo_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\06.Semanal\Criacao_Estudos.xlsm'
        tipo_estudo = 'SEMANAL'
    #caminho_arquivo_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\Criacao_Estudos.xlsm' #caminho do arquivo base onde estão registrados os casos
    aplicacao_excel = client.DispatchEx("Excel.Application")
    aplicacao_excel.Visible = True
    arquivo_excel = aplicacao_excel.Workbooks.Open(Filename =caminho_arquivo_base)
    formulario = arquivo_excel.Worksheets('formulario')
    #try:
    token = autenticar_prospec()
    #except:
        ##arquivo_excel.DisplayAlerts = False
        #del arquivo_excel
        #arquivo_excel.Application.Quit()
        #print('Erro de login no prospec')
    print('Quantidade de requisições utilizadas: ', conta_requisicoes(token))
    estudos_pendentes = 0

    linha = 7 #7 - linha do primeiro caso a ser baixado
    caminho_rodada = formulario.Cells(linha, 14).Value
    # Loop que verifica linha a linha do arquivo em excel a vai baixando os estudos pra quantas linhas tiverem preenchidas
    #esse loop usa como base o preenchimento ou não da coluna do nome de arquivo newave
    while (formulario.Cells(linha,8).Value != 0 and formulario.Cells(linha, 8).Value != None and formulario.Cells(linha, 8).Value != ''):

        caminho_rodada = formulario.Cells(linha, 14).Value #caminho da rede onde a rodada será armazenada
        inicio_pasta = formulario.Cells(linha, 6).Value #Salva a ordem do estudo (tag no prospec). E1, E2 etc
        idStudy = int(formulario.Cells(linha, 3).Value) #salva em variável o StudyID do estudo que será baixado
        modelo = formulario.Cells(linha, 16).Value # Salva o modelo/mapa usado na criação do estudo
        status_download = formulario.Cells(linha, 4).Value #salva em variável o status de download anterior (para verirficar se já foi feito download ou não)
        data_criacao_estudo = formulario.Cells(linha, 2).Value
        #data_criacao_estudo = data_criacao_estudo.strftime('%d/%m/%Y')
        if status_download != 'Download realizado': #verifica se na planilha está sinalizado como download realizado, caso não tenha sido, faz o download, caso tenha sido realizado, nõa faz nada e passa para o próximo
            status = GetStatusOfStudy(token, idStudy) #pega status do estudo

            pathdownload = os.path.join(caminho_rodada, '04.Download Estudos')
            nome_arquivo = inicio_pasta + '_' + str(idStudy) + '.zip' #salva na variável o nome do arquivo zipado que será criado
            download_compilado(token, idStudy, pathdownload, nome_arquivo) #função que faz o download do estudo
            formulario.Cells(linha, 4).Value = 'Download realizado' #escreve no excel o status como download realizado
            formulario.Cells(linha, 5).Value = datetime.datetime.now().strftime('%m/%d/%Y  %H:%M:%S') #escreve no excel a data e hora do dowload
            print('Download realizado para o estudo ', idStudy, ' - ', inicio_pasta, ' - modelo ', modelo)
            fonte_pluvia = formulario.Cells(linha, 16).Value
            cenario_eneva = formulario.Cells(linha, 17).Value
            observacao = formulario.Cells(linha, 24).Value
            arquiva_estudos_intersemanal_semanal_preco_medio_auto(idStudy, data_criacao_estudo, os.path.join(pathdownload, nome_arquivo) , tipo_estudo, fonte_pluvia, cenario_eneva, observacao, pld_semanal)

        else:
            print('Estudo com download já realizado anteriormente - IdStudy: ', idStudy)
        arquivo_excel.Save() #salva o arquivo excel que tem os estudos a cada novo estudo baixado, isso faz com que caso dê algum erro, ele sempre tem o status do últmo download salvo
        linha = linha + 1

    arquivo_excel.Save() #salva o escel uma última vez
    arquivo_excel.Application.Quit() #fecha o escel
    print('Quantidade de requisições utilizadas: ', conta_requisicoes(token))
    return {"Estudos pendente": estudos_pendentes, "Caminho rodada": caminho_rodada}


#####################################################################################################################
#                                                   PLD MEDIO SEMANAL
#####################################################################################################################

def download_estudos_nao_finalizados_novo_arquivo_semanal(**kwargs):
    criar_copia_semanal ()
    apagar_linhas_semanal()
    #pld_realizado ()
    if kwargs.get('caminho_arquivo_estudos'):
        caminho_arquivo_base = kwargs.get('caminho_arquivo_estudos')
        tipo_estudo = 'INTERSEMANAL'
    else:
        caminho_arquivo_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\06.Semanal\Criacao_Estudos.xlsm'
        tipo_estudo = 'SEMANAL'
    #caminho_arquivo_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\Criacao_Estudos.xlsm' #caminho do arquivo base onde estão registrados os casos
    aplicacao_excel = client.DispatchEx("Excel.Application")
    aplicacao_excel.Visible = True
    arquivo_excel = aplicacao_excel.Workbooks.Open(Filename =caminho_arquivo_base)
    formulario = arquivo_excel.Worksheets('formulario')
    #try:
    token = autenticar_prospec()
    #except:
        ##arquivo_excel.DisplayAlerts = False
        #del arquivo_excel
        #arquivo_excel.Application.Quit()
        #print('Erro de login no prospec')
    print('Quantidade de requisições utilizadas: ', conta_requisicoes(token))
    estudos_pendentes = 0

    linha = 7 #7 - linha do primeiro caso a ser baixado
    caminho_rodada = formulario.Cells(linha, 14).Value
    # Loop que verifica linha a linha do arquivo em excel a vai baixando os estudos pra quantas linhas tiverem preenchidas
    #esse loop usa como base o preenchimento ou não da coluna do nome de arquivo newave
    while (formulario.Cells(linha,8).Value != 0 and formulario.Cells(linha, 8).Value != None and formulario.Cells(linha, 8).Value != ''):

        caminho_rodada = formulario.Cells(linha, 14).Value #caminho da rede onde a rodada será armazenada
        inicio_pasta = formulario.Cells(linha, 6).Value #Salva a ordem do estudo (tag no prospec). E1, E2 etc
        idStudy = int(formulario.Cells(linha, 3).Value) #salva em variável o StudyID do estudo que será baixado
        modelo = formulario.Cells(linha, 16).Value # Salva o modelo/mapa usado na criação do estudo
        status_download = formulario.Cells(linha, 4).Value #salva em variável o status de download anterior (para verirficar se já foi feito download ou não)
        data_criacao_estudo = formulario.Cells(linha, 2).Value
        observacao = formulario.Cells(linha, 24).Value
        #data_criacao_estudo = data_criacao_estudo.strftime('%d/%m/%Y')
        if status_download != 'Download realizado': #verifica se na planilha está sinalizado como download realizado, caso não tenha sido, faz o download, caso tenha sido realizado, nõa faz nada e passa para o próximo
            status = GetStatusOfStudy(token, idStudy) #pega status do estudo

            pathdownload = os.path.join(caminho_rodada, '04.Download Estudos')
            nome_arquivo = inicio_pasta + '_' + str(idStudy) + '.zip' #salva na variável o nome do arquivo zipado que será criado
            download_compilado(token, idStudy, pathdownload, nome_arquivo) #função que faz o download do estudo
            formulario.Cells(linha, 4).Value = 'Download realizado' #escreve no excel o status como download realizado
            formulario.Cells(linha, 5).Value = datetime.datetime.now().strftime('%m/%d/%Y  %H:%M:%S') #escreve no excel a data e hora do dowload
            print('Download realizado para o estudo ', idStudy, ' - ', inicio_pasta, ' - modelo ', modelo)
            fonte_pluvia = formulario.Cells(linha, 16).Value
            cenario_eneva = formulario.Cells(linha, 17).Value
            StudyId_semanal (StudyId = idStudy)
            arquiva_estudos_intersemanal_semanal_semanal(idStudy, data_criacao_estudo, os.path.join(pathdownload, nome_arquivo) , tipo_estudo, fonte_pluvia, cenario_eneva)
            arquiva_estudos_intersemanal_semanal(idStudy, data_criacao_estudo, os.path.join(pathdownload, nome_arquivo) , tipo_estudo, fonte_pluvia, cenario_eneva, observacao)
            funcao_copia_cola_semanal(StudyId = idStudy)
        else:
            print('Estudo com download já realizado anteriormente - IdStudy: ', idStudy)
        arquivo_excel.Save() #salva o arquivo excel que tem os estudos a cada novo estudo baixado, isso faz com que caso dê algum erro, ele sempre tem o status do últmo download salvo
        linha = linha + 1
    #arquiva_pld ()
    #arquiva_valor (StudyId = idStudy) 
    arquivo_excel.Save() #salva o escel uma última vez
    arquivo_excel.Application.Quit() #fecha o excel
    arquiva_pld_semanal ()
    pld_realizado_semanal ()
    #ena_realizada_semanal ()
    arquiva_valor_semanal ()
    salvar_copia_semanal ()
    retorna_arquivo_semanal ()  
    print('Quantidade de requisições utilizadas: ', conta_requisicoes(token))
    return {"Estudos pendente": estudos_pendentes, "Caminho rodada": caminho_rodada}

def arquiva_estudos_intersemanal_semanal_semanal (StudyId, data_estudo, caminho_estudo_zip, tipo_estudo, fonte_pluvia, cenario_eneva):
    caminho_estudo = os.path.realpath(caminho_estudo_zip)
    caminho_pasta_unzip = caminho_estudo.replace('.zip','')
    if type(data_estudo) != str:
        data_criacao = data_estudo.strftime('%d/%m/%Y')
        print('Data de criação foi convertida para String')
    else:
        print('Data de criação já recebida como string')
        data_criacao = data_estudo

    #verifica se pasta descompacta ja existe
    if os.path.exists(caminho_pasta_unzip):
        shutil.rmtree(caminho_pasta_unzip)
    #descompacta arquivo zip do estudo
    with ZipFile(caminho_estudo_zip, 'r') as zipObj:
        zipObj.extractall(caminho_pasta_unzip)
    time.sleep(10)
    deleta_linhas_base_prospec_diario(StudyId)
    time.sleep(5)
    arquiva_dados_rodada_diaria_semanal(StudyId, data_criacao, caminho_pasta_unzip, tipo_estudo, Fonte_Pluvia = fonte_pluvia, Cenario_Eneva = cenario_eneva)

    #shutil.rmtree(caminho_pasta_unzip)

def arquiva_dados_rodada_diaria_semanal(StudyId, data, caminho_estudo, tipo_preliminar_definitivo, **kwargs):
    caminho_estudo = os.path.realpath(caminho_estudo)
    caminho_arquivo_base_prospec = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario_v1.xlsx'
    print('Iniciando a importação de dados da pasta:', caminho_estudo)
    arquivo_excel_rodada = load_workbook(caminho_arquivo_base_prospec)
    aba_cadastro = arquivo_excel_rodada['Cadastro_estudos']
    arquivos_csv = ['compila_cmo_medio', 'compila_ena_ons','compila_ena','compila_ea', 'compila_ea_inicial', 'compila_ena_ree', 'compila_gh', 'compila_gt', 'compila_intercambio_medio']
    #Verifica se já existem dados desse estudo na planilha
    #imputa informações da aba de cadastro do estudo
    #arquivos_csv_temp = deepcopy(arquivos_csv)
   # arquivos_csv_temp.append('Cadastro_estudos')
    #salvar_base = False

    data_criacao = data
    linha = aba_cadastro.max_row + 1
    print('Número da primeira linha vazia na aba Cadastro:', linha)
    aba_cadastro.cell(row=linha, column=2).value = data_criacao
    aba_cadastro.cell(row=linha, column=1).value = aba_cadastro.cell(row=linha - 1, column=1).value + 1
    aba_cadastro.cell(row=linha, column=2).number_format = 'dd/mm/yyyy'
    aba_cadastro.cell(row=linha, column=3).value = StudyId
    aba_cadastro.cell(row=linha, column=6).value = tipo_preliminar_definitivo
    if kwargs.get('Fonte_Pluvia'):
        aba_cadastro.cell(row=linha, column=17).value = kwargs.get('Fonte_Pluvia')
    if kwargs.get('Cenario_Eneva'):
        aba_cadastro.cell(row=linha, column=18).value = kwargs.get('Cenario_Eneva')
    for nome_csv in arquivos_csv:
        nome_arquivo = nome_csv + '.csv'
        nome_aba = arquivo_excel_rodada[nome_csv]

        with open(caminho_estudo + '\\' + nome_arquivo) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=';')
            line_count = 0
            linha = nome_aba.max_row + 1
            for row in csv_reader:
                if line_count == 0:
                    print(f'Column names are {", ".join(row)}')
                    line_count += 1
                else:
                    for coluna in range(len(row)):
                        nome_aba.cell(row=linha, column=coluna + 2).value = row[coluna]
                    nome_aba.cell(row=linha, column=1).value = StudyId
                    linha = linha + 1
                    line_count += 1
            print(f'Processed {line_count} lines.')


    arquivo_excel_rodada.save(caminho_arquivo_base_prospec)
    arquivo_excel_rodada.close()
    time.sleep(10)

    #?#salvar copia no caminho para banco de dados
    #endereco_rede_prospec2 = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\25.BANCO_DE_DADOS\Planilhas\Prospec\Base_prospec_diario.xlsx'
    #shutil.copy2(caminho_arquivo_base_prospec, endereco_rede_prospec2)

#####################################################################################################################
#                                                       PLD MEDIO INTERSEMANAL
#####################################################################################################################

def download_estudos_nao_finalizados_novo_arquivo_intersemanal(**kwargs): 
    criar_copia ()
    apagar_linhas()
    if kwargs.get('caminho_arquivo_estudos'):
        caminho_arquivo_base = kwargs.get('caminho_arquivo_estudos')
        tipo_estudo = 'INTERSEMANAL'
    else:
        caminho_arquivo_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\06.Semanal\Criacao_Estudos.xlsm'
        tipo_estudo = 'SEMANAL'
    #caminho_arquivo_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\Criacao_Estudos.xlsm' #caminho do arquivo base onde estão registrados os casos
    aplicacao_excel = client.DispatchEx("Excel.Application")
    aplicacao_excel.Visible = True
    arquivo_excel = aplicacao_excel.Workbooks.Open(Filename =caminho_arquivo_base)
    formulario = arquivo_excel.Worksheets('formulario')
    #try:
    token = autenticar_prospec()
    #except:
        ##arquivo_excel.DisplayAlerts = False
        #del arquivo_excel
        #arquivo_excel.Application.Quit()
        #print('Erro de login no prospec')
    print('Quantidade de requisições utilizadas: ', conta_requisicoes(token))
    estudos_pendentes = 0

    linha = 7 #7 - linha do primeiro caso a ser baixado
    caminho_rodada = formulario.Cells(linha, 14).Value
    # Loop que verifica linha a linha do arquivo em excel a vai baixando os estudos pra quantas linhas tiverem preenchidas
    #esse loop usa como base o preenchimento ou não da coluna do nome de arquivo newave
    while (formulario.Cells(linha,8).Value != 0 and formulario.Cells(linha, 8).Value != None and formulario.Cells(linha, 8).Value != ''):

        caminho_rodada = formulario.Cells(linha, 14).Value #caminho da rede onde a rodada será armazenada
        inicio_pasta = formulario.Cells(linha, 6).Value #Salva a ordem do estudo (tag no prospec). E1, E2 etc
        idStudy = int(formulario.Cells(linha, 3).Value) #salva em variável o StudyID do estudo que será baixado
        modelo = formulario.Cells(linha, 16).Value # Salva o modelo/mapa usado na criação do estudo
        status_download = formulario.Cells(linha, 4).Value #salva em variável o status de download anterior (para verirficar se já foi feito download ou não)
        data_criacao_estudo = formulario.Cells(linha, 2).Value
        #data_criacao_estudo = data_criacao_estudo.strftime('%d/%m/%Y')
        if status_download != 'Download realizado': #verifica se na planilha está sinalizado como download realizado, caso não tenha sido, faz o download, caso tenha sido realizado, nõa faz nada e passa para o próximo
            status = GetStatusOfStudy(token, idStudy) #pega status do estudo

            pathdownload = os.path.join(caminho_rodada, '04.Download Estudos')
            nome_arquivo = inicio_pasta + '_' + str(idStudy) + '.zip' #salva na variável o nome do arquivo zipado que será criado
            download_compilado(token, idStudy, pathdownload, nome_arquivo) #função que faz o download do estudo
            formulario.Cells(linha, 4).Value = 'Download realizado' #escreve no excel o status como download realizado
            formulario.Cells(linha, 5).Value = datetime.datetime.now().strftime('%m/%d/%Y  %H:%M:%S') #escreve no excel a data e hora do dowload
            print('Download realizado para o estudo ', idStudy, ' - ', inicio_pasta, ' - modelo ', modelo)
            fonte_pluvia = formulario.Cells(linha, 16).Value
            cenario_eneva = formulario.Cells(linha, 17).Value
            observacao = formulario.Cells(linha, 24).Value
            StudyId (StudyId = idStudy)
            arquiva_estudos_intersemanal_semanal_intersemanal(idStudy, data_criacao_estudo, os.path.join(pathdownload, nome_arquivo) , tipo_estudo, fonte_pluvia, cenario_eneva)
            arquiva_estudos_intersemanal_semanal(idStudy, data_criacao_estudo, os.path.join(pathdownload, nome_arquivo) , tipo_estudo, fonte_pluvia, cenario_eneva, observacao)    
            #funcao_copia_cola(StudyId = idStudy)
            #arquiva_pld (StudyId = idStudy)
            #pld_realizado (StudyId = idStudy)
            #arquiva_valor (StudyId = idStudy) 
        else:
            print('Estudo com download já realizado anteriormente - IdStudy: ', idStudy)
        arquivo_excel.Save() #salva o arquivo excel que tem os estudos a cada novo estudo baixado, isso faz com que caso dê algum erro, ele sempre tem o status do últmo download salvo
        linha = linha + 1
    #arquiva_pld ()
    #arquiva_valor (StudyId = idStudy) 
    arquivo_excel.Save() #salva o escel uma última vez
    arquivo_excel.Application.Quit() #fecha o excel
    arquiva_pld ()
    pld_realizado ()
    #ena_realizada ()
    arquiva_valor ()
    salvar_copia ()
      
    print('Quantidade de requisições utilizadas: ', conta_requisicoes(token))
    return {"Estudos pendente": estudos_pendentes, "Caminho rodada": caminho_rodada}

def arquiva_estudos_intersemanal_semanal_intersemanal (StudyId, data_estudo, caminho_estudo_zip, tipo_estudo, fonte_pluvia, cenario_eneva):
    caminho_estudo = os.path.realpath(caminho_estudo_zip)
    caminho_pasta_unzip = caminho_estudo.replace('.zip','')
    if type(data_estudo) != str:
        data_criacao = data_estudo.strftime('%d/%m/%Y')
        print('Data de criação foi convertida para String')
    else:
        print('Data de criação já recebida como string')
        data_criacao = data_estudo

    #verifica se pasta descompacta ja existe
    if os.path.exists(caminho_pasta_unzip):
        shutil.rmtree(caminho_pasta_unzip)
    #descompacta arquivo zip do estudo
    with ZipFile(caminho_estudo_zip, 'r') as zipObj:
        zipObj.extractall(caminho_pasta_unzip)
    time.sleep(10)
    deleta_linhas_base_prospec_diario(StudyId)
    time.sleep(5)
    arquiva_dados_rodada_diaria_intersemanal(StudyId, data_criacao, caminho_pasta_unzip, tipo_estudo, Fonte_Pluvia = fonte_pluvia, Cenario_Eneva = cenario_eneva)

    #shutil.rmtree(caminho_pasta_unzip)

def arquiva_dados_rodada_diaria_intersemanal(StudyId, data, caminho_estudo, tipo_preliminar_definitivo, **kwargs):
    caminho_estudo = os.path.realpath(caminho_estudo)
    caminho_arquivo_base_prospec = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario.xlsx'
    print('Iniciando a importação de dados da pasta:', caminho_estudo)
    arquivo_excel_rodada = load_workbook(caminho_arquivo_base_prospec)
    aba_cadastro = arquivo_excel_rodada['Cadastro_estudos']
    arquivos_csv = ['compila_cmo_medio', 'compila_ena_ons','compila_ena','compila_ea', 'compila_ea_inicial', 'compila_ena_ree', 'compila_gh', 'compila_gt', 'compila_intercambio_medio']
    #Verifica se já existem dados desse estudo na planilha
    #imputa informações da aba de cadastro do estudo
    #arquivos_csv_temp = deepcopy(arquivos_csv)
   # arquivos_csv_temp.append('Cadastro_estudos')
    #salvar_base = False

    data_criacao = data
    linha = aba_cadastro.max_row + 1
    print('Número da primeira linha vazia na aba Cadastro:', linha)
    aba_cadastro.cell(row=linha, column=2).value = data_criacao
    aba_cadastro.cell(row=linha, column=1).value = aba_cadastro.cell(row=linha - 1, column=1).value + 1
    aba_cadastro.cell(row=linha, column=2).number_format = 'dd/mm/yyyy'
    aba_cadastro.cell(row=linha, column=3).value = StudyId
    aba_cadastro.cell(row=linha, column=6).value = tipo_preliminar_definitivo
    if kwargs.get('Fonte_Pluvia'):
        aba_cadastro.cell(row=linha, column=17).value = kwargs.get('Fonte_Pluvia')
    if kwargs.get('Cenario_Eneva'):
        aba_cadastro.cell(row=linha, column=18).value = kwargs.get('Cenario_Eneva')
    for nome_csv in arquivos_csv:
        nome_arquivo = nome_csv + '.csv'
        nome_aba = arquivo_excel_rodada[nome_csv]

        with open(caminho_estudo + '\\' + nome_arquivo) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=';')
            line_count = 0
            linha = nome_aba.max_row + 1
            for row in csv_reader:
                if line_count == 0:
                    print(f'Column names are {", ".join(row)}')
                    line_count += 1
                else:
                    for coluna in range(len(row)):
                        nome_aba.cell(row=linha, column=coluna + 2).value = row[coluna]
                    nome_aba.cell(row=linha, column=1).value = StudyId
                    linha = linha + 1
                    line_count += 1
            print(f'Processed {line_count} lines.')


    arquivo_excel_rodada.save(caminho_arquivo_base_prospec)
    arquivo_excel_rodada.close()
    time.sleep(10)

    #?#salvar copia no caminho para banco de dados
    #endereco_rede_prospec2 = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\25.BANCO_DE_DADOS\Planilhas\Prospec\Base_prospec_diario.xlsx'
    #shutil.copy2(caminho_arquivo_base_prospec, endereco_rede_prospec2)


def arquiva_estudos_intersemanal_semanal_preco_medio_auto (StudyId, data_estudo, caminho_estudo_zip, tipo_estudo, fonte_pluvia, cenario_eneva, observacao, pld_semanal):
    caminho_estudo = os.path.realpath(caminho_estudo_zip)
    caminho_pasta_unzip = caminho_estudo.replace('.zip','')
    if type(data_estudo) != str:
        data_criacao = data_estudo.strftime('%d/%m/%Y')
        print('Data de criação foi convertida para String')
    else:
        print('Data de criação já recebida como string')
        data_criacao = data_estudo

    #verifica se pasta descompacta ja existe
    if os.path.exists(caminho_pasta_unzip):
        shutil.rmtree(caminho_pasta_unzip)
    #descompacta arquivo zip do estudo
    with ZipFile(caminho_estudo_zip, 'r') as zipObj:
        zipObj.extractall(caminho_pasta_unzip)
    print('Descompactado')
    time.sleep(10)
    #deleta_linhas_base_prospec_diario(StudyId)
    time.sleep(5)
    arquiva_dados_rodada_diaria_preco_mensal_auto(StudyId, data_criacao, caminho_pasta_unzip, tipo_estudo, Fonte_Pluvia = fonte_pluvia, Cenario_Eneva = cenario_eneva, observacao = observacao, pld_semanal=pld_semanal)
    shutil.rmtree(caminho_pasta_unzip)

#####################################################################################################################
# DESCOMPACTA ESTUDO, ARQUIVA NA BASE DE DADOS, APAGA PASTA DESCOMPACTADA
#####################################################################################################################
def arquiva_estudos_intersemanal_semanal (StudyId, data_estudo, caminho_estudo_zip, tipo_estudo, fonte_pluvia, cenario_eneva, observacao):
    caminho_estudo = os.path.realpath(caminho_estudo_zip)
    caminho_pasta_unzip = caminho_estudo.replace('.zip','')
    if type(data_estudo) != str:
        data_criacao = data_estudo.strftime('%d/%m/%Y')
        print('Data de criação foi convertida para String')
    else:
        print('Data de criação já recebida como string')
        data_criacao = data_estudo

    #verifica se pasta descompacta ja existe
    if os.path.exists(caminho_pasta_unzip):
        shutil.rmtree(caminho_pasta_unzip)
    #descompacta arquivo zip do estudo
    with ZipFile(caminho_estudo_zip, 'r') as zipObj:
        zipObj.extractall(caminho_pasta_unzip)
    print('Descompactado')
    time.sleep(10)
    #deleta_linhas_base_prospec_diario(StudyId)
    time.sleep(5)
    arquiva_dados_rodada_diaria(StudyId, data_criacao, caminho_pasta_unzip, tipo_estudo, Fonte_Pluvia = fonte_pluvia, Cenario_Eneva = cenario_eneva, observacao = observacao)

    shutil.rmtree(caminho_pasta_unzip)

#####################################################################################################################
# APAGA PREVS QUE NÃO ESTÃO MAPEADAS NA CRIAÇÃO DO ESTUDO, MANTÉM SÓ OS MESES DESEJADOS
#####################################################################################################################

def apaga_prevs_nao_mapeadas (qtd_meses, caminho_prevs):
    #caminho_prevs = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\2020\12\11\02.Prevs'
    mes = int(qtd_meses)
    meses = []
    for file in os.listdir(caminho_prevs): # cria uma lista de meses de prevs que existem na pasta
        meses.append(file[:6])

    #print('meses completos:', meses)
    meses = list(set(meses)) #traz uma lista com registro únicos dos meses que constam na pasta
    meses.sort() #ordena de forma ascendente os meses na pasta
    print(meses)
    x = len(meses) - 1 #salva variável do tamanho da lista dos meses
    while x > (mes - 1): #deleta da lista os meses que não devem se utilizados no estudo
        meses.pop(x)

        x = x - 1
    print('mes restantes:', meses)
    for file in os.listdir(caminho_prevs): #remove as prevs que não serão utilizadas no estudo
        if file[:6] not in meses:
            os.remove(caminho_prevs + '\\' + file)
            print('arquivo removido:', file)

def renomeia_prevs(pasta_prevs):
    caminho_pasta = Path(pasta_prevs)
    for arquivo in os.listdir(pasta_prevs):
        novo_nome = arquivo[:12] + arquivo[-4:]
        os.rename(os.path.join(pasta_prevs, arquivo), os.path.join(pasta_prevs, novo_nome))

###########################################################################################
#Esta função descobre o dia de término da semana operativa, e a semana operativa em si, a partir de uma data em datetime
############################################################################################
def descobre_sem_op_dia_term (data_datetime):
    #dia da semana conforme pythom, segunda-feira = 0, domingo = 6
    dia_semana = data_datetime.weekday()
    coef = 4 - dia_semana
    #print('Coeficiente:', coef)
    if dia_semana > 4:
        term_semana_op = data_datetime + relativedelta(days=7 + coef)
    else:
        term_semana_op = data_datetime + relativedelta(days=coef)

    #verifica qual o dia do término da semana operativa
    dia = term_semana_op.day

    #descobre qual é a semana operativa
    if dia <=7:
        semana_op = 'rv0'
    elif dia <= 14:
        semana_op = 'rv1'
    elif dia <= 21:
        semana_op = 'rv2'
    elif dia <= 28:
        semana_op = 'rv3'
    else:
        semana_op = 'rv4'
    print('Data considerada:', data_datetime.strftime('%d/%m/%Y'),' - Término Semana Operativa:', term_semana_op.strftime('%d/%m/%Y'), ' - Semana Operativa:', semana_op)
    return (term_semana_op, semana_op)

def limpa_prevs_semana_atual (caminho_prevs):
    caminho_prevs = os.path.realpath(caminho_prevs)
    resposta = descobre_sem_op_dia_term(datetime.datetime.today())
    termino_semana_operativa = resposta[0]
    semana_operativa = resposta[1]
    ano = str(termino_semana_operativa.year)
    mes = '%02d' %termino_semana_operativa.month
    for prevs in os.listdir(caminho_prevs):
        if prevs.endswith(semana_operativa) and prevs.startswith(ano + mes):
            if prevs != (ano + mes + '-prevs.' + semana_operativa):
                os.remove(caminho_prevs + '\\' + prevs)
                print('Arquivo deletado:', prevs, 'Caminho do arquivo:', caminho_prevs)

def gera_arquivo_UH_atualizado (**kwargs):
    caminho_arquivo_UH = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\05. Diario\04.Arquivos Padrão'
    nome_processador_UH = 'Decomp - UH_VI_LV - Treinamento_Rv1_V0.xlsm'
    nome_arquivo_UH = 'volume_uhe.csv'
    excel_macro = client.DispatchEx("Excel.Application")
    wb = excel_macro.Workbooks.Open(Filename = caminho_arquivo_UH + '\\' + nome_processador_UH)
    ws = wb.Worksheets('UH')
    excel_macro.Visible = True
    num_modulo = 2
    lista_macros = ('atualiza_armazenamento','cria_arquivo_UH')
    for nome_macro in lista_macros:
        macro = 'Módulo' + str(num_modulo) + '.' + nome_macro
        print('Iniciando Execução da Macro: ' + macro)
        wb.Application.Run("\'" + nome_processador_UH + "\'" + '!' + macro)
        print('Macro executada com sucesso')
    #wb.Save()  # , FileFormat = 52
    print('Arquivo processador salvo e Planilha UH gerada.')
    time.sleep(5)
    wb.Close(False)
    excel_macro.Application.Quit()
    del excel_macro
    if kwargs.get('caminho_rodada'):
        caminho_rodada_decks =  os.path.join(Path(kwargs.get('caminho_rodada')),'01.Decks')
        shutil.copy2(os.path.join(Path(caminho_arquivo_UH), nome_processador_UH), caminho_rodada_decks)
        shutil.copy2(os.path.join(Path(caminho_arquivo_UH), nome_arquivo_UH), caminho_rodada_decks)

#retorna o deck ID da semana seguinte do estudo para subir o arquivo UH

def deck_decomp_semana_seguinte(idStudy):
    lista_decks = getListOfDecks(int(idStudy))
    lista_mes_rev = []
    for deck in lista_decks:
        if deck['Model'] == "DECOMP" and deck['SensibilityInfo'] == 'Original':
            lista_mes_rev.append(str(deck['Year']) + '%02d' %deck['Month'] + str(deck['Revision']))
            print('Código do deck da semana seguinte:',deck['Id'], ' - Nome arquivo do deck:', deck['FileName'])
            #return deck['Id']
    print('Lista de Meses completa:', lista_mes_rev)
    lista_mes_rev = list(set(lista_mes_rev))
    lista_mes_rev.sort()
    print('lista com meses distintos e ordenados:' ,lista_mes_rev)
    ano = lista_mes_rev[1][:4]
    mes = lista_mes_rev[1][4:6]
    rev = lista_mes_rev[1][-1:]
    print('Ano:', ano, 'Mês:', mes, 'Revisão:', rev)
    for deck in lista_decks:
        if deck['Model'] == "DECOMP" and deck['SensibilityInfo'] == 'Original' and deck['Year'] == int(ano) and deck['Month'] == int(mes) and deck['Revision']== int(rev):
            print(deck)
            print('DeckID:', deck['Id'])
            return deck['Id']

def download_estudos_finalizados_diario (Tipo_relatorio):
    caminho_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\05. Diario'
    criador_rodada_diaria = 'Criacao_Estudos_diario.xlsm'
    print('Abrindo arquivo criador da rodada diária')
    excel_diario = client.DispatchEx("Excel.Application")
    excel_diario.Visible = True
    excel_diario.DisplayAlerts = False
    wb_criador_diario = excel_diario.Workbooks.Open(Filename = caminho_base + '\\' + criador_rodada_diaria)

    formulario = wb_criador_diario.Worksheets('formulario')

    token = autenticar_prospec()
    print('Token:', token)
    estudos_pendentes = 0
    if Tipo_relatorio == 'Preliminar':
        linha = 7
    elif Tipo_relatorio == 'Definitivo':
        linha = 17
    caminho_rodada = formulario.Cells(linha, 14).Value
    # Loop que verifica linha a linha do arquivo em excel a vai baixando os estudos pra quantas linhas tiverem
    #while (formulario.Cells( linha, 3).Value != 0 and formulario.Cells(linha, 3).Value != None):
    inicio_pasta = str(formulario.Cells(linha, 1).Value) + '_' + formulario.Cells(linha, 7).Value
    idStudy = int(formulario.Cells(linha, 3).Value)
    modelo = formulario.Cells(linha, 13).Value
    status_dowload = formulario.Cells(linha, 4).Value
    print('Inicio pasta:', inicio_pasta)
    print('idStudy:', idStudy, ' - ',str(idStudy))
    print('Modelo:', modelo)
    print('Status Download:', status_dowload)
    if status_dowload != 'Download realizado_2':
        status = GetStatusOfStudy(token, idStudy)
        if status == 'Finished':

            pathdownload = os.path.join(caminho_rodada, '04.Download Estudos', Tipo_relatorio)
            os.makedirs(pathdownload, exist_ok=True)
            nome_arquivo = inicio_pasta + '_' + str(idStudy) + '.zip'
            download_compilado(token, idStudy, pathdownload, nome_arquivo)
            formulario.Cells(linha, 4).Value = 'Download realizado'
            formulario.Cells(linha, 5).Value = datetime.datetime.now().strftime('%m/%d/%Y  %H:%M:%S')
            caminho_zip = os.path.join(pathdownload, nome_arquivo)
            print('Download realizado para o estudo ', idStudy, ' - ', inicio_pasta, ' - modelo ', modelo)
            #formulario.Cells(linha, 2).ClearContents
            #formulario.Cells(linha, 3).ClearContents
            #formulario.Cells(linha, 4).ClearContents
            #formulario.Cells(linha, 5).ClearContents
            excel_diario.DisplayAlerts = True
            wb_criador_diario.Save()
            time.sleep(5)
            wb_criador_diario.Close(False)
            excel_diario.Application.Quit()
            del excel_diario
            return {"Estudos pendentes": estudos_pendentes, "Caminho rodada": caminho_rodada, "Caminho ZIP": caminho_zip}
        else:
            estudos_pendentes = estudos_pendentes + 1
            formulario.Cells(linha, 4).Value = status
            formulario.Cells(linha, 5).Value = datetime.datetime.now().strftime('%m/%d/%Y  %H:%M:%S')
            print('Dowload não realizado - IdStudy: ', idStudy, 'Status: ', status)
    else:
        print('Estudo com download já realizado anteriormente - IdStudy: ', idStudy)
    excel_diario.DisplayAlerts = True
    wb_criador_diario.Save()
        #linha = linha + 1

    time.sleep(5)

    wb_criador_diario.Close(False)
    excel_diario.Application.Quit()
    del excel_diario

    return {"Estudos pendentes": estudos_pendentes, "Caminho rodada": caminho_rodada}

def download_estudos_finalizados_diario2_com_mensal(Tipo_relatorio, pld_semanal):
    caminho_base =r'C:\Users\alex.lourenco\OneDrive - Eneva S.A\Documentos\processos_alex\diario'
        #r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\05. Diario'
    criador_rodada_diaria = 'Criacao_Estudos_diario.xlsm'
    print('Abrindo arquivo criador da rodada diária')
    excel_diario = client.DispatchEx("Excel.Application")
    excel_diario.Visible = True
    excel_diario.DisplayAlerts = False
    wb_criador_diario = excel_diario.Workbooks.Open(Filename = caminho_base + '\\' + criador_rodada_diaria)

    formulario = wb_criador_diario.Worksheets('formulario')

    token = autenticar_prospec()
    print('Token:', token)
    estudos_pendentes = 0
    if Tipo_relatorio == 'Definitivo':
        linha = 7
    elif Tipo_relatorio == 'Definitivo':
        linha = 17

    caminho_rodada = formulario.Cells(linha, 14).Value
    # Loop que verifica linha a linha do arquivo em excel a vai baixando os estudos pra quantas linhas tiverem
    while (formulario.Cells( linha, 3).Value != 0 and formulario.Cells(linha, 3).Value != None and formulario.Cells(linha, 3).Value != ''):
        inicio_pasta = str(formulario.Cells(linha, 1).Value) + '_' + formulario.Cells(linha, 7).Value
        idStudy = int(formulario.Cells(linha, 3).Value)
        modelo = formulario.Cells(linha, 13).Value
        status_dowload = formulario.Cells(linha, 4).Value
        print('Inicio pasta:', inicio_pasta)
        print('idStudy:', idStudy, ' - ',str(idStudy))
        print('Modelo:', modelo)
        print('Status Download:', status_dowload)
        if status_dowload != 'Download realizado':
            status = GetStatusOfStudy(token, idStudy)
            if status == 'Finished':

                pathdownload = os.path.join(caminho_rodada, '04.Download Estudos', Tipo_relatorio)
                os.makedirs(pathdownload, exist_ok=True)
                nome_arquivo = inicio_pasta + '_' + str(idStudy) + '.zip'
                download_compilado(token, idStudy, pathdownload, nome_arquivo)
                formulario.Cells(linha, 4).Value = 'Download realizado'
                formulario.Cells(linha, 5).Value = datetime.datetime.now().strftime('%m/%d/%Y  %H:%M:%S')
                caminho_zip = os.path.join(pathdownload, nome_arquivo)
                print('Download realizado para o estudo ', idStudy, ' - ', inicio_pasta, ' - modelo ', modelo)
                #formulario.Cells(linha, 2).ClearContents
                #formulario.Cells(linha, 3).ClearContents
                #formulario.Cells(linha, 4).ClearContents
                #formulario.Cells(linha, 5).ClearContents
                #salva
                #excel_diario.DisplayAlerts = True
                #wb_criador_diario.Save()
                print("salvo")
                #NEW
                caminho_estudo_descompactado = caminho_zip.replace('.zip','')
                shutil.unpack_archive(caminho_zip, caminho_estudo_descompactado,'zip')
                print("tira zip")
                StudyId = int(caminho_estudo_descompactado[caminho_estudo_descompactado.rfind('_') + 1:])
                print(StudyId)
                now = datetime.datetime.now()
                data_mapas = now.strftime('%d/%m/%Y')
                print('antes do arquivado')
                fonte_pluvia = formulario.Cells(linha, 16).Value
                cenario_eneva = formulario.Cells(linha, 17).Value
                observacao = formulario.Cells(linha, 24).Value
                arquiva_dados_rodada_diaria2_com_mensal(StudyId, data_mapas,caminho_estudo_descompactado,
                                                        Tipo_relatorio.upper(),Fonte_Pluvia = fonte_pluvia,Cenario_Eneva = cenario_eneva,
                                                        pld_semanal=pld_semanal, observacao=observacao)
                print('arquivado')

                # #envia arquivo da rodada diária por e-mail para substituir a base corrente no onedrive
                # caminho_pasta_base_prospec = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec'
                # nome_base_prospec = 'Base_prospec_diario_rodadadiaria.xlsx'
                # assunto = 'DIARIO - BASE PROSPEC PARA ATUALIZAR NO POWER BI'# + ' - ' +  datetime.datetime.today().strftime('%d/%m/%Y')
                # corpo = 'Prezados,\n\nSegue em anexo base do prospec atualizada para update do Power Bi.\n\n'
                # caminho_anexo = caminho_pasta_base_prospec + '\\' + nome_base_prospec
                # anexo = nome_base_prospec
                # destinatario = 'maria.barbosa@eneva.com.br'
                # envia_email_python(assunto, corpo, caminho_anexo, anexo, destinatario)
                # print('E-mail com base de dados do prospec enviado para ativação do fluxo do Power Automate')

                #return {"Estudos pendentes": estudos_pendentes, "Caminho rodada": caminho_rodada, "Caminho ZIP": caminho_zip}
            else:
                estudos_pendentes = estudos_pendentes + 1
                formulario.Cells(linha, 4).Value = status
                formulario.Cells(linha, 5).Value = datetime.datetime.now().strftime('%m/%d/%Y  %H:%M:%S')
                print('Dowload não realizado - IdStudy: ', idStudy, 'Status: ', status)
        else:
            print('Estudo com download já realizado anteriormente - IdStudy: ', idStudy)
        excel_diario.DisplayAlerts = True
        wb_criador_diario.Save()
        linha = linha + 1

    time.sleep(5)

    wb_criador_diario.Close(False)
    excel_diario.Application.Quit()
    del excel_diario

    return {"Estudos pendentes": estudos_pendentes, "Caminho rodada": caminho_rodada}


def download_estudos_finalizados_diario2 (Tipo_relatorio):
    caminho_base = r'C:\Users\alex.lourenco\OneDrive - Eneva S.A\Documentos\processos_alex\diario'
        #r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\05. Diario'
    criador_rodada_diaria = 'Criacao_Estudos_diario.xlsm'
    print('Abrindo arquivo criador da rodada diária')
    excel_diario = client.DispatchEx("Excel.Application")
    excel_diario.Visible = True
    excel_diario.DisplayAlerts = False
    wb_criador_diario = excel_diario.Workbooks.Open(Filename = caminho_base + '\\' + criador_rodada_diaria)

    formulario = wb_criador_diario.Worksheets('formulario')

    token = autenticar_prospec()
    print('Token:', token)
    estudos_pendentes = 0
    if Tipo_relatorio == 'Preliminar':
        linha = 7
    elif Tipo_relatorio == 'Definitivo':
        linha = 17
    caminho_rodada = formulario.Cells(linha, 14).Value
    # Loop que verifica linha a linha do arquivo em excel a vai baixando os estudos pra quantas linhas tiverem
    while (formulario.Cells( linha, 3).Value != 0 and formulario.Cells(linha, 3).Value != None and formulario.Cells(linha, 3).Value != ''):
        inicio_pasta = str(formulario.Cells(linha, 1).Value) + '_' + formulario.Cells(linha, 7).Value
        idStudy = int(formulario.Cells(linha, 3).Value)
        modelo = formulario.Cells(linha, 13).Value
        status_dowload = formulario.Cells(linha, 4).Value
        print('Inicio pasta:', inicio_pasta)
        print('idStudy:', idStudy, ' - ',str(idStudy))
        print('Modelo:', modelo)
        print('Status Download:', status_dowload)
        if status_dowload != 'Download realizado':
            status = GetStatusOfStudy(token, idStudy)
            if status == 'Finished':

                pathdownload = os.path.join(caminho_rodada, '04.Download Estudos', Tipo_relatorio)
                os.makedirs(pathdownload, exist_ok=True)
                nome_arquivo = inicio_pasta + '_' + str(idStudy) + '.zip'
                download_compilado(token, idStudy, pathdownload, nome_arquivo)
                formulario.Cells(linha, 4).Value = 'Download realizado'
                formulario.Cells(linha, 5).Value = datetime.datetime.now().strftime('%m/%d/%Y  %H:%M:%S')
                caminho_zip = os.path.join(pathdownload, nome_arquivo)
                print('Download realizado para o estudo ', idStudy, ' - ', inicio_pasta, ' - modelo ', modelo)
                #formulario.Cells(linha, 2).ClearContents
                #formulario.Cells(linha, 3).ClearContents
                #formulario.Cells(linha, 4).ClearContents
                #formulario.Cells(linha, 5).ClearContents
                #salva
                #excel_diario.DisplayAlerts = True
                #wb_criador_diario.Save()
                print("salvo")
                #NEW
                caminho_estudo_descompactado = caminho_zip.replace('.zip','')
                shutil.unpack_archive(caminho_zip, caminho_estudo_descompactado,'zip')
                print("tira zip")
                StudyId = int(caminho_estudo_descompactado[caminho_estudo_descompactado.rfind('_') + 1:])
                print(StudyId)
                now = datetime.datetime.now()
                data_mapas = now.strftime('%d/%m/%Y')
                print('antes do arquivado')
                fonte_pluvia = formulario.Cells(linha, 16).Value
                cenario_eneva = formulario.Cells(linha, 17).Value
                observacao = formulario.Cells(linha, 24).Value
                arquiva_dados_rodada_diaria2(StudyId, data_mapas,caminho_estudo_descompactado, Tipo_relatorio.upper(),Fonte_Pluvia = fonte_pluvia,Cenario_Eneva = cenario_eneva, observacao=observacao)
                print('arquivado')

                #envia arquivo da rodada diária por e-mail para substituir a base corrente no onedrive
                #caminho_pasta_base_prospec = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec'
                #nome_base_prospec = 'Base_prospec_diario_rodadadiaria.xlsx'
                #assunto = 'DIARIO - BASE PROSPEC PARA ATUALIZAR NO POWER BI'# + ' - ' +  datetime.datetime.today().strftime('%d/%m/%Y')
                #corpo = 'Prezados,\n\nSegue em anexo base do prospec atualizada para update do Power Bi.\n\n'
                #caminho_anexo = caminho_pasta_base_prospec + '\\' + nome_base_prospec
                # anexo = nome_base_prospec
                # destinatario = 'maria.barbosa@eneva.com.br'
                #envia_email_python(assunto, corpo, caminho_anexo, anexo, destinatario)
                print('E-mail com base de dados do prospec enviado para ativação do fluxo do Power Automate')

                #return {"Estudos pendentes": estudos_pendentes, "Caminho rodada": caminho_rodada, "Caminho ZIP": caminho_zip}
            else:
                estudos_pendentes = estudos_pendentes + 1
                formulario.Cells(linha, 4).Value = status
                formulario.Cells(linha, 5).Value = datetime.datetime.now().strftime('%m/%d/%Y  %H:%M:%S')
                print('Dowload não realizado - IdStudy: ', idStudy, 'Status: ', status)
        else:
            print('Estudo com download já realizado anteriormente - IdStudy: ', idStudy)
        excel_diario.DisplayAlerts = True
        wb_criador_diario.Save()
        linha = linha + 1

    time.sleep(5)

    wb_criador_diario.Close(False)
    excel_diario.Application.Quit()
    del excel_diario

    return {"Estudos pendentes": estudos_pendentes, "Caminho rodada": caminho_rodada}

def download_estudos_finalizados_diario3 (Tipo_relatorio):
    caminho_base = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\05. Diario'
    criador_rodada_diaria = 'Criacao_Estudos_diario_prevsENEVA.xlsm'
    print('Abrindo arquivo criador da rodada diária')
    excel_diario = client.DispatchEx("Excel.Application")
    excel_diario.Visible = True
    excel_diario.DisplayAlerts = False
    wb_criador_diario = excel_diario.Workbooks.Open(Filename = caminho_base + '\\' + criador_rodada_diaria)

    formulario = wb_criador_diario.Worksheets('formulario')

    token = autenticar_prospec()
    print('Token:', token)
    estudos_pendentes = 0
    if Tipo_relatorio == 'Preliminar':
        linha = 7
    elif Tipo_relatorio == 'Definitivo':
        linha = 17
    caminho_rodada = formulario.Cells(linha, 14).Value
    # Loop que verifica linha a linha do arquivo em excel a vai baixando os estudos pra quantas linhas tiverem
    while (formulario.Cells( linha, 3).Value != 0 and formulario.Cells(linha, 3).Value != None and formulario.Cells(linha, 3).Value != ''):
        inicio_pasta = str(formulario.Cells(linha, 1).Value) + '_' + formulario.Cells(linha, 7).Value
        idStudy = int(formulario.Cells(linha, 3).Value)
        modelo = formulario.Cells(linha, 13).Value
        status_dowload = formulario.Cells(linha, 4).Value
        print('Inicio pasta:', inicio_pasta)
        print('idStudy:', idStudy, ' - ',str(idStudy))
        print('Modelo:', modelo)
        print('Status Download:', status_dowload)
        if status_dowload != 'Download realizado':
            status = GetStatusOfStudy(token, idStudy)
            if status == 'Finished':

                pathdownload = os.path.join(caminho_rodada, '04.Download Estudos', Tipo_relatorio)
                os.makedirs(pathdownload, exist_ok=True)
                nome_arquivo = inicio_pasta + '_' + str(idStudy) + '.zip'
                download_compilado(token, idStudy, pathdownload, nome_arquivo)
                formulario.Cells(linha, 4).Value = 'Download realizado'
                formulario.Cells(linha, 5).Value = datetime.datetime.now().strftime('%m/%d/%Y  %H:%M:%S')
                caminho_zip = os.path.join(pathdownload, nome_arquivo)
                print('Download realizado para o estudo ', idStudy, ' - ', inicio_pasta, ' - modelo ', modelo)
                #formulario.Cells(linha, 2).ClearContents
                #formulario.Cells(linha, 3).ClearContents
                #formulario.Cells(linha, 4).ClearContents
                #formulario.Cells(linha, 5).ClearContents
                #salva
                #excel_diario.DisplayAlerts = True
                #wb_criador_diario.Save()
                print("salvo")
                #NEW
                caminho_estudo_descompactado = caminho_zip.replace('.zip','')
                shutil.unpack_archive(caminho_zip, caminho_estudo_descompactado,'zip')
                print("tira zip")
                StudyId = int(caminho_estudo_descompactado[caminho_estudo_descompactado.rfind('_') + 1:])
                print(StudyId)
                now = datetime.datetime.now()
                data_mapas = now.strftime('%d/%m/%Y')
                print('antes do arquivado')
                fonte_pluvia = formulario.Cells(linha, 16).Value
                cenario_eneva = formulario.Cells(linha, 17).Value
                observacao = formulario.Cells(linha, 24).Value
                arquiva_dados_rodada_diaria2(StudyId, data_mapas,caminho_estudo_descompactado, Tipo_relatorio.upper(),Fonte_Pluvia = fonte_pluvia,Cenario_Eneva = cenario_eneva, observacao=observacao)
                print('arquivado')

                #envia arquivo da rodada diária por e-mail para substituir a base corrente no onedrive
                caminho_pasta_base_prospec = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec'
                nome_base_prospec = 'Base_prospec_diario_rodadadiaria.xlsx'
                assunto = 'DIARIO - BASE PROSPEC PARA ATUALIZAR NO POWER BI'# + ' - ' +  datetime.datetime.today().strftime('%d/%m/%Y')
                corpo = 'Prezados,\n\nSegue em anexo base do prospec atualizada para update do Power Bi.\n\n'
                caminho_anexo = caminho_pasta_base_prospec + '\\' + nome_base_prospec
                anexo = nome_base_prospec
                destinatario = 'maria.barbosa@eneva.com.br' 
                envia_email_python(assunto, corpo, caminho_anexo, anexo, destinatario)
                print('E-mail com base de dados do prospec enviado para ativação do fluxo do Power Automate')

                #return {"Estudos pendentes": estudos_pendentes, "Caminho rodada": caminho_rodada, "Caminho ZIP": caminho_zip}
            else:
                estudos_pendentes = estudos_pendentes + 1
                formulario.Cells(linha, 4).Value = status
                formulario.Cells(linha, 5).Value = datetime.datetime.now().strftime('%m/%d/%Y  %H:%M:%S')
                print('Dowload não realizado - IdStudy: ', idStudy, 'Status: ', status)
        else:
            print('Estudo com download já realizado anteriormente - IdStudy: ', idStudy)
        excel_diario.DisplayAlerts = True
        wb_criador_diario.Save()
        linha = linha + 1

    time.sleep(5)

    wb_criador_diario.Close(False)
    excel_diario.Application.Quit()
    del excel_diario

    return {"Estudos pendentes": estudos_pendentes, "Caminho rodada": caminho_rodada}


def copia_arquivos_padrão_rodada_diaria(caminho_rodada):
    caminho_rodada = os.path.realpath(caminho_rodada)
    caminho_arquivos_padrao = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\05. Diario\05.Arquivos Semanais'
    lista_pasta = ['01.Decks', '03.GEVAZP']
    for pasta in lista_pasta:
        caminho_pasta_padrao = caminho_arquivos_padrao + '\\' + pasta
        for file in os.listdir(caminho_pasta_padrao):
            print(file)
            caminho_arquivo_local = caminho_pasta_padrao + '\\' + file
            caminho_arquivo_rede = caminho_rodada + '\\' + pasta #+ '\\' + file
            shutil.copy2(caminho_arquivo_local, caminho_arquivo_rede)
            print('Arquivo padrão copiado para pasta da rodada:', file, 'Caminho da Rodada:', caminho_rodada)
            print('ok')

def processa_rodada_diaria(caminho_estudo_zip, tipo_relatorio):
    caminho_estudo_zip = os.path.realpath(caminho_estudo_zip)
    caminho_processador = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\05. Diario'
    nome_processador = 'Relatorio_Diario_V4.xlsm'
    caminho_arquivo_processador = caminho_processador + '\\' + nome_processador
    excel_processador_diario = client.DispatchEx("Excel.Application")
    excel_processador_diario.DisplayAlerts = False
    excel_processador_diario.Visible = True
    wb_pross_dia = excel_processador_diario.Workbooks.Open(Filename = caminho_arquivo_processador)

    ws_rel = wb_pross_dia.Worksheets('Relatorio')
    ws_aux = wb_pross_dia.Worksheets('aux')
    ws_rel.Range("AI10").Value = tipo_relatorio.upper()
    ws_aux.Cells(2,2).Value = caminho_estudo_zip
    nome_macro = 'relatorio_diario_automatico_python'
    num_modulo = 1
    macro = 'Módulo' + str(num_modulo) + '.' + nome_macro
    print('Iniciando Execução da Macro: ' + macro)
    #roda a macro
    wb_pross_dia.Application.Run("\'" + nome_processador + "\'" + '!' + macro)
    print('Macro excecutada com sucesso: ' + macro)
    caminho_estudo_descompac = ws_aux.Cells(3, 2).Value
    idStudy = int(caminho_estudo_descompac[caminho_estudo_descompac.rfind('_') + 1:])
    nome_macro = 'Gera_PDF_range'
    num_modulo = 2
    macro = 'Módulo' + str(num_modulo) + '.' + nome_macro
    print('Iniciando Execução da Macro: ' + macro)
    wb_pross_dia.Application.Run("\'" + nome_processador + "\'" + '!' + macro)
    print('Macro excecutada com sucesso: ' + macro)


    caminho_relatorio = ws_aux.Cells(4,2).Value
    nome_relatorio = ws_aux.Cells(5,2).Value

    excel_processador_diario.DisplayAlerts = True
    wb_pross_dia.Close(True)
    excel_processador_diario.Application.Quit()
    del excel_processador_diario
    return (caminho_relatorio, nome_relatorio, idStudy, caminho_estudo_descompac)

def arquiva_dados_rodada_diaria2_com_mensal(StudyId, data, caminho_estudo, tipo_preliminar_definitivo, observacao, pld_semanal, **kwargs):
    caminho_estudo = os.path.realpath(caminho_estudo)
    caminho_arquivo_base_prospec = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario_rodadadiaria.xlsx'
    print('Iniciando a importação de dados da pasta:', caminho_estudo)
    arquivo_excel_rodada = load_workbook(caminho_arquivo_base_prospec)
    aba_cadastro = arquivo_excel_rodada['Cadastro_estudos']
    arquivos_csv = ['compila_cmo_medio', 'compila_ena_ons','compila_ea', 'compila_ea_inicial']
    # ['compila_cmo_medio', 'compila_ena_ons','compila_ena','compila_ea', 'compila_ea_inicial', 'compila_ena_ree', 'compila_gh', 'compila_gt', 'compila_intercambio_medio']
    #Verifica se já existem dados desse estudo na planilha
    #imputa informações da aba de cadastro do estudo
    #arquivos_csv_temp = deepcopy(arquivos_csv)
   # arquivos_csv_temp.append('Cadastro_estudos')
    #salvar_base = False

    data_criacao = data
    linha = aba_cadastro.max_row + 1
    print('Número da primeira linha vazia na aba Cadastro:', linha)
    aba_cadastro.cell(row=linha, column=2).value = data_criacao
    aba_cadastro.cell(row=linha, column=1).value = aba_cadastro.cell(row=linha - 1, column=1).value + 1
    aba_cadastro.cell(row=linha, column=2).number_format = 'dd/mm/yyyy'
    aba_cadastro.cell(row=linha, column=3).value = StudyId
    aba_cadastro.cell(row=linha, column=6).value = tipo_preliminar_definitivo
    aba_cadastro.cell(row=linha, column=25).value = observacao 
    if kwargs.get('Fonte_Pluvia'):
        aba_cadastro.cell(row=linha, column=17).value = kwargs.get('Fonte_Pluvia')
    if kwargs.get('Cenario_Eneva'):
        aba_cadastro.cell(row=linha, column=18).value = kwargs.get('Cenario_Eneva')
    for nome_csv in arquivos_csv:
        nome_arquivo = nome_csv + '.csv'
        nome_aba = arquivo_excel_rodada[nome_csv]

        # TODO: Adicionado por Alexandre
        if nome_csv == 'compila_cmo_medio':
            df_compila = pd.read_csv(os.path.join(caminho_estudo, nome_arquivo), sep=';')
            df_compila[['SUDESTE', 'SUL', 'NORDESTE', 'NORTE']] = df_compila[['SUDESTE', 'SUL', 'NORDESTE', 'NORTE']].clip(lower=piso, upper=teto_diario)
            df_aux = df_compila[df_compila['MEN=0-SEM=1'] == 1][['Deck', 'SUDESTE', 'SUL', 'NORDESTE', 'NORTE']]

            pld_pivot = pld_semanal.pivot(index='SemOp', columns='submercado', values='pld').reset_index()
            pld_pivot['SemOp'] = [f"{x}_s1" for x in pld_pivot['SemOp']]
            pld_pivot = pld_pivot.rename(columns={'SemOp': 'Deck'})

            df_all = pd.concat([pld_pivot, df_aux], axis=0)

            if actual_week_historical:
                df_all = df_all.drop_duplicates(subset='Deck', keep='first')
            else:
                df_all = df_all.drop_duplicates(subset='Deck', keep='last')

            year = int(df_all.iloc[0, 0][2:6])
            month = int(df_all.iloc[0, 0][6:8])
            dia_inicio = next_friday(date(year, month, 1)) - delta(days=6)
            df_all['dia'] = [dia_inicio + delta(days=7*i) for i in range(df_all.shape[0])]

            year_final = df_all['dia'].values[-1].year
            month_final = df_all['dia'].values[-1].month
            day_final = monthrange(year_final, month_final)[1]
            data_final = date(year_final, month_final, day_final)
            new_days = [dia_inicio + delta(days=i) for i in range((data_final - dia_inicio).days+1)]

            df_days = pd.DataFrame(data=[], columns=df_all.columns)
            df_days['dia'] = new_days

            df_all_days = pd.concat([df_all, df_days], axis=0)
            df_all_days = df_all_days.drop_duplicates(subset='dia', keep='first')
            df_all_days = df_all_days.sort_values(by='dia').ffill(axis=0)
            df_all_days['anomes'] = df_all_days['dia'].apply(lambda x: x.strftime("%Y%m"))
            df_mes = df_all_days.groupby('anomes').agg('mean')
        ####

        with open(caminho_estudo + '\\' + nome_arquivo) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=';')
            line_count = 0
            linha = nome_aba.max_row + 1
            for row in csv_reader:
                if line_count == 0:
                    print(f'Column names are {", ".join(row)}')
                    line_count += 1
                else:
                    for coluna in range(len(row)):
                        nome_aba.cell(row=linha, column=coluna + 2).value = row[coluna]
                    nome_aba.cell(row=linha, column=1).value = StudyId
                    linha = linha + 1
                    line_count += 1

        if nome_csv == 'compila_cmo_medio':

            months = sorted(set(int(f"{df_compila.loc[i, 'Deck'][2:8]}")
                                for i in list(df_compila.index) if df_compila.loc[i, 'MEN=0-SEM=1'] == 1),
                            reverse=False)
            for month in months:
                nome_aba.cell(row=linha, column=1).value = StudyId
                nome_aba.cell(row=linha, column=2).value = 'Original'
                nome_aba.cell(row=linha, column=3).value = f"DC{month}-sem1_s1"
                nome_aba.cell(row=linha, column=4).value = '0'
                nome_aba.cell(row=linha, column=5).value = f"{round(df_mes.loc[str(month), 'SUDESTE'], 2):02}"
                nome_aba.cell(row=linha, column=6).value = f"{round(df_mes.loc[str(month), 'SUL'], 2):02}"
                nome_aba.cell(row=linha, column=7).value = f"{round(df_mes.loc[str(month), 'NORDESTE'], 2):02}"
                nome_aba.cell(row=linha, column=8).value = f"{round(df_mes.loc[str(month), 'NORTE'], 2):02}"

                linha += 1
                line_count += 1

            print(f'Processed {line_count} lines.')

    arquivo_excel_rodada.save(caminho_arquivo_base_prospec)
    arquivo_excel_rodada.close()
    time.sleep(10)

    # salvar copia no caminho para banco de dados
    endereco_rede_prospec2 = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\25.BANCO_DE_DADOS\Planilhas\Prospec\Base_prospec_diario_rodadadiaria.xlsx'
    shutil.copy2(caminho_arquivo_base_prospec, endereco_rede_prospec2)


def arquiva_dados_rodada_diaria_preco_mensal_auto(StudyId, data, caminho_estudo, tipo_preliminar_definitivo, observacao, pld_semanal, **kwargs):
    caminho_estudo = os.path.realpath(caminho_estudo)
    caminho_arquivo_base_prospec = r'C:\Users\alex.lourenco\OneDrive - Eneva S.A\Documentos\processos_alex\intersemanal\Base_prospec_diario.xlsx'
        #r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario.xlsx'
    print('Iniciando a importação de dados da pasta:', caminho_estudo)
    arquivo_excel_rodada = load_workbook(caminho_arquivo_base_prospec)
    aba_cadastro = arquivo_excel_rodada['Cadastro_estudos']
    arquivos_csv = ['compila_cmo_medio', 'compila_ena_ons','compila_ea', 'compila_ea_inicial']
    # ['compila_cmo_medio', 'compila_ena_ons','compila_ena','compila_ea', 'compila_ea_inicial', 'compila_ena_ree', 'compila_gh', 'compila_gt', 'compila_intercambio_medio']
    #Verifica se já existem dados desse estudo na planilha
    #imputa informações da aba de cadastro do estudo
    #arquivos_csv_temp = deepcopy(arquivos_csv)
   # arquivos_csv_temp.append('Cadastro_estudos')
    #salvar_base = False

    data_criacao = data
    linha = aba_cadastro.max_row + 1
    print('Número da primeira linha vazia na aba Cadastro:', linha)
    aba_cadastro.cell(row=linha, column=2).value = data_criacao
    aba_cadastro.cell(row=linha, column=1).value = aba_cadastro.cell(row=linha - 1, column=1).value + 1
    aba_cadastro.cell(row=linha, column=2).number_format = 'dd/mm/yyyy'
    aba_cadastro.cell(row=linha, column=3).value = StudyId
    aba_cadastro.cell(row=linha, column=6).value = tipo_preliminar_definitivo
    aba_cadastro.cell(row=linha, column=25).value = observacao 
    if kwargs.get('Fonte_Pluvia'):
        aba_cadastro.cell(row=linha, column=17).value = kwargs.get('Fonte_Pluvia')
    if kwargs.get('Cenario_Eneva'):
        aba_cadastro.cell(row=linha, column=18).value = kwargs.get('Cenario_Eneva')
    for nome_csv in arquivos_csv:
        nome_arquivo = nome_csv + '.csv'
        nome_aba = arquivo_excel_rodada[nome_csv]

        # TODO: Adicionado por Alexandre
        if nome_csv == 'compila_cmo_medio':
            df_compila = pd.read_csv(os.path.join(caminho_estudo, nome_arquivo), sep=';')
            df_compila[['SUDESTE', 'SUL', 'NORDESTE', 'NORTE']] = df_compila[['SUDESTE', 'SUL', 'NORDESTE', 'NORTE']].clip(lower=piso, upper=teto_diario)
            df_aux = df_compila[df_compila['MEN=0-SEM=1'] == 1][['Deck', 'SUDESTE', 'SUL', 'NORDESTE', 'NORTE']]

            pld_pivot = pld_semanal.pivot(index='SemOp', columns='submercado', values='pld').reset_index()
            pld_pivot['SemOp'] = [f"{x}_s1" for x in pld_pivot['SemOp']]
            pld_pivot = pld_pivot.rename(columns={'SemOp': 'Deck'})

            df_all = pd.concat([pld_pivot, df_aux], axis=0)

            if actual_week_historical:
                df_all = df_all.drop_duplicates(subset='Deck', keep='first')
            else:
                df_all = df_all.drop_duplicates(subset='Deck', keep='last')

            year = int(df_all.iloc[0, 0][2:6])
            month = int(df_all.iloc[0, 0][6:8])
            dia_inicio = next_friday(date(year, month, 1)) - delta(days=6)
            df_all['dia'] = [dia_inicio + delta(days=7*i) for i in range(df_all.shape[0])]

            year_final = df_all['dia'].values[-1].year
            month_final = df_all['dia'].values[-1].month
            day_final = monthrange(year_final, month_final)[1]
            data_final = date(year_final, month_final, day_final)
            new_days = [dia_inicio + delta(days=i) for i in range((data_final - dia_inicio).days+1)]

            df_days = pd.DataFrame(data=[], columns=df_all.columns)
            df_days['dia'] = new_days

            df_all_days = pd.concat([df_all, df_days], axis=0)
            df_all_days = df_all_days.drop_duplicates(subset='dia', keep='first')
            df_all_days = df_all_days.sort_values(by='dia').ffill(axis=0)
            df_all_days['anomes'] = df_all_days['dia'].apply(lambda x: x.strftime("%Y%m"))
            df_mes = df_all_days.groupby('anomes').agg('mean')
        ####

        with open(caminho_estudo + '\\' + nome_arquivo) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=';')
            line_count = 0
            linha = nome_aba.max_row + 1
            for row in csv_reader:
                if line_count == 0:
                    print(f'Column names are {", ".join(row)}')
                    line_count += 1
                else:
                    for coluna in range(len(row)):
                        nome_aba.cell(row=linha, column=coluna + 2).value = row[coluna]
                    nome_aba.cell(row=linha, column=1).value = StudyId
                    linha = linha + 1
                    line_count += 1

        if nome_csv == 'compila_cmo_medio':

            months = sorted(set(int(f"{df_compila.loc[i, 'Deck'][2:8]}")
                                for i in list(df_compila.index) if df_compila.loc[i, 'MEN=0-SEM=1'] == 1), reverse=False)
            for month in months:
                nome_aba.cell(row=linha, column=1).value = StudyId
                nome_aba.cell(row=linha, column=2).value = 'Original'
                nome_aba.cell(row=linha, column=3).value = f"DC{month}-sem1_s1"
                nome_aba.cell(row=linha, column=4).value = '0'
                nome_aba.cell(row=linha, column=5).value = f"{round(df_mes.loc[str(month), 'SUDESTE'], 2):02}"
                nome_aba.cell(row=linha, column=6).value = f"{round(df_mes.loc[str(month), 'SUL'], 2):02}"
                nome_aba.cell(row=linha, column=7).value = f"{round(df_mes.loc[str(month), 'NORDESTE'], 2):02}"
                nome_aba.cell(row=linha, column=8).value = f"{round(df_mes.loc[str(month), 'NORTE'], 2):02}"

                linha += 1
                line_count += 1

            print(f'Processed {line_count} lines.')

    arquivo_excel_rodada.save(caminho_arquivo_base_prospec)
    arquivo_excel_rodada.close()
    time.sleep(10)

    #?#salvar copia no caminho para banco de dados
#    endereco_rede_prospec2 = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\25.BANCO_DE_DADOS\Planilhas\Prospec\Base_prospec_diario.xlsx'
#    shutil.copy2(caminho_arquivo_base_prospec, endereco_rede_prospec2)

def arquiva_dados_rodada_diaria(StudyId, data, caminho_estudo, tipo_preliminar_definitivo, observacao, **kwargs):
    caminho_estudo = os.path.realpath(caminho_estudo)
    caminho_arquivo_base_prospec = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario.xlsx'
    print('Iniciando a importação de dados da pasta:', caminho_estudo)
    arquivo_excel_rodada = load_workbook(caminho_arquivo_base_prospec)
    aba_cadastro = arquivo_excel_rodada['Cadastro_estudos']
    arquivos_csv = ['compila_cmo_medio', 'compila_ena_ons','compila_ea', 'compila_ea_inicial']
    # ['compila_cmo_medio', 'compila_ena_ons','compila_ena','compila_ea', 'compila_ea_inicial', 'compila_ena_ree', 'compila_gh', 'compila_gt', 'compila_intercambio_medio']
    #Verifica se já existem dados desse estudo na planilha
    #imputa informações da aba de cadastro do estudo
    #arquivos_csv_temp = deepcopy(arquivos_csv)
   # arquivos_csv_temp.append('Cadastro_estudos')
    #salvar_base = False

    data_criacao = data
    linha = aba_cadastro.max_row + 1
    print('Número da primeira linha vazia na aba Cadastro:', linha)
    aba_cadastro.cell(row=linha, column=2).value = data_criacao
    aba_cadastro.cell(row=linha, column=1).value = aba_cadastro.cell(row=linha - 1, column=1).value + 1
    aba_cadastro.cell(row=linha, column=2).number_format = 'dd/mm/yyyy'
    aba_cadastro.cell(row=linha, column=3).value = StudyId
    aba_cadastro.cell(row=linha, column=6).value = tipo_preliminar_definitivo
    aba_cadastro.cell(row=linha, column=25).value = observacao 
    if kwargs.get('Fonte_Pluvia'):
        aba_cadastro.cell(row=linha, column=17).value = kwargs.get('Fonte_Pluvia')
    if kwargs.get('Cenario_Eneva'):
        aba_cadastro.cell(row=linha, column=18).value = kwargs.get('Cenario_Eneva')
    for nome_csv in arquivos_csv:
        nome_arquivo = nome_csv + '.csv'
        nome_aba = arquivo_excel_rodada[nome_csv]

        with open(caminho_estudo + '\\' + nome_arquivo) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=';')
            line_count = 0
            linha = nome_aba.max_row + 1
            for row in csv_reader:
                if line_count == 0:
                    print(f'Column names are {", ".join(row)}')
                    line_count += 1
                else:
                    for coluna in range(len(row)):
                        nome_aba.cell(row=linha, column=coluna + 2).value = row[coluna]
                    nome_aba.cell(row=linha, column=1).value = StudyId
                    linha = linha + 1
                    line_count += 1
            print(f'Processed {line_count} lines.')


    arquivo_excel_rodada.save(caminho_arquivo_base_prospec)
    arquivo_excel_rodada.close()
    time.sleep(10)

    #?#salvar copia no caminho para banco de dados
    endereco_rede_prospec2 = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\25.BANCO_DE_DADOS\Planilhas\Prospec\Base_prospec_diario.xlsx'
    shutil.copy2(caminho_arquivo_base_prospec, endereco_rede_prospec2)

def arquiva_dados_rodada_diaria2(StudyId, data, caminho_estudo, tipo_preliminar_definitivo, observacao, **kwargs):
    caminho_estudo = os.path.realpath(caminho_estudo)
    caminho_arquivo_base_prospec = r'C:\SCRIPTS_\Base_prospec_diario_rodadadiaria.xlsx'
        #r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario_rodadadiaria.xlsx'
    print('Iniciando a importação de dados da pasta:', caminho_estudo)
    arquivo_excel_rodada = load_workbook(caminho_arquivo_base_prospec)
    aba_cadastro = arquivo_excel_rodada['Cadastro_estudos']
    arquivos_csv = ['compila_cmo_medio', 'compila_ena_ons','compila_ea', 'compila_ea_inicial']
    # ['compila_cmo_medio', 'compila_ena_ons','compila_ena','compila_ea', 'compila_ea_inicial', 'compila_ena_ree', 'compila_gh', 'compila_gt', 'compila_intercambio_medio']
    #Verifica se já existem dados desse estudo na planilha
    #imputa informações da aba de cadastro do estudo
    #arquivos_csv_temp = deepcopy(arquivos_csv)
   # arquivos_csv_temp.append('Cadastro_estudos')
    #salvar_base = False

    data_criacao = data
    linha = aba_cadastro.max_row + 1
    print('Número da primeira linha vazia na aba Cadastro:', linha)
    aba_cadastro.cell(row=linha, column=2).value = data_criacao
    aba_cadastro.cell(row=linha, column=1).value = aba_cadastro.cell(row=linha - 1, column=1).value + 1
    aba_cadastro.cell(row=linha, column=2).number_format = 'dd/mm/yyyy'
    aba_cadastro.cell(row=linha, column=3).value = StudyId
    aba_cadastro.cell(row=linha, column=6).value = tipo_preliminar_definitivo
    aba_cadastro.cell(row=linha, column=25).value = observacao 
    if kwargs.get('Fonte_Pluvia'):
        aba_cadastro.cell(row=linha, column=17).value = kwargs.get('Fonte_Pluvia')
    if kwargs.get('Cenario_Eneva'):
        aba_cadastro.cell(row=linha, column=18).value = kwargs.get('Cenario_Eneva')
    for nome_csv in arquivos_csv:
        nome_arquivo = nome_csv + '.csv'
        nome_aba = arquivo_excel_rodada[nome_csv]

        with open(caminho_estudo + '\\' + nome_arquivo) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=';')
            line_count = 0
            linha = nome_aba.max_row + 1
            for row in csv_reader:
                if line_count == 0:
                    print(f'Column names are {", ".join(row)}')
                    line_count += 1
                else:
                    for coluna in range(len(row)):
                        nome_aba.cell(row=linha, column=coluna + 2).value = row[coluna]
                    nome_aba.cell(row=linha, column=1).value = StudyId
                    linha = linha + 1
                    line_count += 1
            print(f'Processed {line_count} lines.')


    arquivo_excel_rodada.save(caminho_arquivo_base_prospec)
    arquivo_excel_rodada.close()
    time.sleep(10)

    #?#salvar copia no caminho para banco de dados
    #endereco_rede_prospec2 = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\25.BANCO_DE_DADOS\Planilhas\Prospec\Base_prospec_diario_rodadadiaria.xlsx'
    #shutil.copy2(caminho_arquivo_base_prospec, endereco_rede_prospec2)

def deleta_linhas_base_prospec_diario(StudyId):
    #caminho_estudo = os.path.realpath(caminho_estudo)
    caminho_arquivo_base_prospec = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_diario.xlsx'
    #print('Iniciando a importação de dados da pasta:', caminho_estudo)
    arquivo_excel_rodada = load_workbook(caminho_arquivo_base_prospec)
    #aba_cadastro = arquivo_excel_rodada['Cadastro_estudos']
    #arquivos_csv = ['compila_cmo_medio', 'compila_ena_ons','compila_ena','compila_ea', 'compila_ea_inicial', 'compila_ena_ree', 'compila_gh', 'compila_gt', 'compila_intercambio_medio']
    arquivos_csv = ['compila_cmo_medio', 'compila_ena_ons','compila_ea', 'compila_ea_inicial']
    #Verifica se já existem dados desse estudo na planilha
    #imputa informações da aba de cadastro do estudo
    arquivos_csv_temp = deepcopy(arquivos_csv)
    arquivos_csv_temp.append('Cadastro_estudos')
    salvar_base = False

    for aba in arquivos_csv_temp:
        if aba == 'Cadastro_estudos':
            coluna = 3
        else: coluna = 1
        nome_aba = arquivo_excel_rodada[aba]
        #ult_lin = nome_aba.max_row
        linha = 2
        qtd_lin_del = 0
        while nome_aba.cell(linha, coluna).value is not None:
            if int(nome_aba.cell(linha, coluna).value) == int(StudyId):
                nome_aba.delete_rows(linha, amount=1)
                #print(linha deletada)
                qtd_lin_del += 1
                salvar_base = True
            else:
                linha += 1
        print('Aba:', aba,'quantidade de linhas deletadas:', qtd_lin_del)
        if nome_aba.cell(linha, coluna).value is None:
            nome_aba.delete_rows(linha, amount=1)
            print('Linha adicional deletada')
    if salvar_base:
        arquivo_excel_rodada.save(caminho_arquivo_base_prospec)
        print('Base de dados Salva')
    arquivo_excel_rodada.close()

##
# PROCESSA A RODADA DIÁRIA A PARTIR do caminho da rodada
#
def processa_rodada_intersemanal (caminho_rodada):
    caminho_rodada = os.path.realpath(caminho_rodada)
    caminho_processador = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\04. Intersemanal'
    nome_processador = 'Boletim_ Intersemanal_v7.xlsm'

    #Abre arquivo processador da rodada intersemanal
    caminho_arquivo_processador = caminho_processador + '\\' + nome_processador
    excel_processador_intersemanal = client.DispatchEx("Excel.Application")
    excel_processador_intersemanal.DisplayAlerts = False
    excel_processador_intersemanal.Visible = True
    wb_pross_intersem = excel_processador_intersemanal.Workbooks.Open(Filename = caminho_arquivo_processador)
    wb_pross_intersem.Worksheets('Resumo do Caso'). Cells(4,35).Value = caminho_rodada + '\\' + '04.Download Estudos'

    #atualiza PLD
    wb_pross_intersem.RefreshAll()
    print('PLD Atualizado')

    #Atualiza ENA/EARM
    nome_macro = 'atualiza_ENA_acomph'
    num_modulo = 9
    macro = 'Módulo' + str(num_modulo) + '.' + nome_macro
    print('Iniciando Execução da Macro: ' + macro)
    wb_pross_intersem.Application.Run("\'" + nome_processador + "\'" + '!' + macro)
    print('Macro excecutada com sucesso: ' + macro)

    #Apaga dadas ENA BASE
    nome_macro = 'apaga_dados_ENA_BASE'
    num_modulo = 9
    macro = 'Módulo' + str(num_modulo) + '.' + nome_macro
    print('Iniciando Execução da Macro: ' + macro)
    wb_pross_intersem.Application.Run("\'" + nome_processador + "\'" + '!' + macro)
    print('Macro excecutada com sucesso: ' + macro)

    #Roda reusltados intersemanal
    nome_macro = 'CompiladorResultados_python'
    num_modulo = 1
    macro = 'Módulo' + str(num_modulo) + '.' + nome_macro
    print('Iniciando Execução da Macro: ' + macro)
    wb_pross_intersem.Application.Run("\'" + nome_processador + "\'" + '!' + macro)
    print('Macro excecutada com sucesso: ' + macro)
    time.sleep(10)

    #Ajusta área de impressão
    nome_macro = 'set_print_area_relatorio'
    num_modulo = 9
    macro = 'Módulo' + str(num_modulo) + '.' + nome_macro
    print('Iniciando Execução da Macro: ' + macro)
    wb_pross_intersem.Application.Run("\'" + nome_processador + "\'" + '!' + macro)
    print('Macro excecutada com sucesso: ' + macro)

    #salva excel, fecha arquivo excel e elimina a variável do arquivo
    wb_pross_intersem.Close(True)
    excel_processador_intersemanal.Application.Quit()
    del wb_pross_intersem
    assunto = 'Rodada Intersemanal - ' + datetime.datetime.today().strftime('%d/%m/%Y')
    corpo = 'Prezados,\n\nRodada intersemanal foi baixada e processada. O arquivo de com as configurações se encontra no caminho abaixo:\n\n' + caminho_arquivo_processador
    caminho_anexo = ''
    anexo = ''
    destinatario = 'fernando.fidalgo@eneva.com.br; renata.hunder@eeva.com.br'
    #envia_email_python(assunto, corpo, caminho_anexo, anexo, destinatario)


def calendario_semanas_operativas(data_string=None, hoje=False):
    """Gera dados do calendário operativo diário, segundo o critério adotado pelo ONS.

    Com base numa data de entrada é calculado os dados necessários para os processos de modelagem segundo o ONS.

    Args:
        data_string (str): (DD-MM-YYYY) data base na qual se deseja ter as informações.
        hoje (bool): Habilita o uso da data no momento da execução.
    |
    Returns:
        dict: Dados de calendário necessários para os processos de modelagem do ONS.
            
                | 'data-referencia': data usada como base (Datetime)
                | 'inicio': primeiro dia da semana operativa (DateTime)
                | 'final' : último dia da semana operativo (DateTime)
                | 'dias-realizados-semana': quantos dias da semana operativa são de realizado (int)
                | 'semana-operativa': valor da semana operativa do mes (int)
                | 'rev'   : numeração da revisão da semana (int)
                | 'revisao': retorna a concatenação de 'RV' + rev 
                | 'semana-operativa-ano': número da semana operativa do ano (int)
                | 'mes-operativo': mês operativo (DateTime) - YYYY-MM-01
                | 'inicio-mes-operativo': primeiro sábado do mês operativo
                | 'termino-mes-operativo': última sexta-feria do mês operativo

    """
    
    data_requerida = pendulum.today('America/Sao_Paulo') if hoje else pendulum.from_format(data_string, 'DD-MM-YYYY')
    
    #Todo mês que começa em um sábado ou domingo necessita de uma correção
    necessita_de_tratativa = [pendulum.SATURDAY, pendulum.SUNDAY]
    
    if data_requerida.day_of_week == pendulum.SATURDAY:

        inic_semana_operativa = data_requerida
        final_semana_operativa = data_requerida.add(days=6)
        semana_operativa_do_mes = final_semana_operativa.week_of_month
        semana_operativa_do_ano = final_semana_operativa.week_of_year
        dias_realizados = 0
        mes_operativo = final_semana_operativa.replace(day = 1)
        inicio_mes_operativo = mes_operativo.subtract(days=(mes_operativo.day_of_week+1)) 
        mes_seguinte = mes_operativo.add(months = 1)
        termino_mes_operativo = mes_seguinte.subtract(days=(mes_seguinte.day_of_week+2))
        #se o mês começa no  sábado e domingo, faz um correção
        correcao = 1 if mes_operativo.day_of_week in necessita_de_tratativa else 0


    else:

        # pega o primeiro dia da semana operativa
        # O "+1" corrige o início da semana de domingo para sábado 
        inic_semana_operativa = data_requerida.subtract(days=(data_requerida.day_of_week+1))

        # Calcula qual será o último dia da semana operativa
        # O 5 ao invés de 6(dias da semana sem contar o atual) corrige o início da semana operativa para o sábado
        final_semana_operativa = data_requerida.add(days=(5 - data_requerida.day_of_week))

        # Numeral que representa a semana operativa do mês
        semana_operativa_do_mes = final_semana_operativa.week_of_month

        # Numeral que representa a semana operativa do ano
        semana_operativa_do_ano = final_semana_operativa.week_of_year
        
        dias_realizados = data_requerida.day_of_week+1

        mes_operativo = final_semana_operativa.replace(day = 1)
        inicio_mes_operativo = mes_operativo.subtract(days=(mes_operativo.day_of_week+1)) 
        mes_seguinte = mes_operativo.add(months = 1)
        termino_mes_operativo = mes_seguinte.subtract(days=(mes_seguinte.day_of_week+2))
        #se a semana começa no sábado ou domingo, faz uma correção
        correcao = 1 if mes_operativo.day_of_week in necessita_de_tratativa else 0

    resultado = {'data-referencia':data_requerida,
                'inicio': inic_semana_operativa,
                'final': final_semana_operativa,
                'dias-realizados-semana': dias_realizados,
                'semana-operativa': semana_operativa_do_mes - correcao,
                'rev': semana_operativa_do_mes-1 - correcao,
                'revisao': 'RV' + str(semana_operativa_do_mes-1-correcao),
                'semana-operativa-ano': semana_operativa_do_ano,
                'mes-operativo': mes_operativo,
                'inicio-mes-operativo': inicio_mes_operativo,
                'termino-mes-operativo': termino_mes_operativo
                }
    
    
    return resultado

def arquiva_deck_dessem (caminho_arquivo_dessem_ZIP, **kwargs):
    caminho_arquivo_dessem_ZIP = os.path.realpath(caminho_arquivo_dessem_ZIP)
    caminho_raiz_rede = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\16.DECKS\03.DESSEM'
    if kwargs.get('data_dessem_datetime'):
        data_deck_dessem = kwargs.get('data_dessem_datetime').format('DD-MM-YYYY')
    else:
        data_deck_dessem = pendulum.today().format('DD-MM-YYYY')
    with ZipFile(caminho_arquivo_dessem_ZIP, 'r') as zipObj:
        #Extract all the contents of zip file in current directory and with the same name
        #caminho da pasta é  mesmo do zip, só que sem o ".zip"
        caminho_dessem_pasta = caminho_arquivo_dessem_ZIP.replace('.zip','') 
        zipObj.extractall(caminho_dessem_pasta)
        print('Arquivo descompactado:', caminho_arquivo_dessem_ZIP)
    
    infos_sem_operativa = calendario_semanas_operativas(data_deck_dessem)     #(hoje=True)
    mes_ano_dessem = infos_sem_operativa['mes-operativo'].format('MM') + str(infos_sem_operativa['mes-operativo'].year)
    revisao = infos_sem_operativa['revisao']
    dia_dessem = infos_sem_operativa['data-referencia'].format('DD')
    caminho_pasta_deck_rede = caminho_raiz_rede + '\\' + str(infos_sem_operativa['data-referencia'].format('YYYY')) + '\\' + str(infos_sem_operativa['data-referencia'].format('MM'))
    
    #nome padrão do deck do dessem do dia dentro do arquivo baixado do ONS
    nome_arquivo_dessem = 'DS_CCEE_' + mes_ano_dessem + '_SEMREDE_' + revisao + 'D' + dia_dessem + '.zip'
    caminho_deck_dessem = caminho_dessem_pasta + '\\' + nome_arquivo_dessem
    
    #verifica se o arquivo do dia existe para poder copiá-lo para a rede
    caminho_deck_rede = caminho_pasta_deck_rede + '\\' + nome_arquivo_dessem
    if os.path.exists(caminho_deck_rede):
        os.remove(caminho_deck_rede)
        print('Arquivo existente na rede foi deletado:', caminho_deck_rede)
    else:
        print('Deck do dessem do dia solicitado não existe na rede:', caminho_deck_dessem)
    print('Caminho do deck na rede:', caminho_deck_rede)
    os.makedirs(caminho_pasta_deck_rede,exist_ok=True)
    shutil.copy2(caminho_deck_dessem, caminho_deck_rede)
    print('Deck do Dessem copiado para pasta na rede:', caminho_deck_rede)



# -----------------------------------------------------------------------------
# Generate DESSEM decks to a prospective study
# Gerar decks DESSEM para um estudo prospectivo
# -----------------------------------------------------------------------------


def generateDessemStudyDecks(idStudy, initialYear, initialMonth, initialDay, duration ,firstDessemFile, firstDecompFile  = '', firstNewaveFile = ''):
    parameter = ''
    data = {
        "InitialDay": initialDay,
        "InitialMonth": initialMonth,
        "InitialYear": initialYear,
        "Duration": duration,
        "NewaveFileName": firstNewaveFile,
        "DecompFileName": firstDecompFile,
        "DessemFileName": firstDessemFile
    }
    print("Gerando decks com as seguintes configuracoes para o estudo: ",
          str(idStudy))
    print(data)
    postInAPI(token, '/api/prospectiveStudies/' + str(idStudy) + '/GenerateDessem',
              parameter, data)
              

def cria_pasta_rodada_dessem (data_dessem = None, hoje = False):
    caminho_raiz_dessem = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\07.Dessem'
    if hoje:
        data_dessem = pendulum.today('America/Sao_Paulo')
    else:
        if not(isinstance(data_dessem, datetime.datetime)):
            print('Forneça uma data no formato datetime para criação das pastas de rodada do dessem')
            quit()

    dia_dessem = data_dessem.format('DD')
    mes_dessem = data_dessem.format('MM')
    ano_dessem = data_dessem.format('YYYY')
    caminho_pasta = ano_dessem + '\\' + mes_dessem + '\\' + dia_dessem
    caminho_pasta_dia = caminho_raiz_dessem + '\\' + '02.Rodadas' + '\\' + caminho_pasta
    
    os.makedirs(caminho_pasta_dia, exist_ok=True)
    subpastas_dessem = ['01.Decks', '02.Prevs', '03.GEVAZP', '04.Download Estudos']
    for subpasta in subpastas_dessem:
        caminho_subpasta = caminho_pasta_dia + '\\' + subpasta
        os.makedirs(caminho_subpasta, exist_ok=True)
    print('Pastas e subpastas criadas no caminho:', caminho_pasta_dia)
    return {'data': data_dessem, 'caminho': caminho_pasta_dia}


# -----------------------------------------------------------------------------
# Get file from REST API | Obter arquivo via REST API
# -----------------------------------------------------------------------------


def getFileFromAPI(*args):

    if len(args) < 3:
        print('Sao necessarios ao menos tres argumentos: token e apiFunction')
        return ''
    elif len(args) == 3:
        token = args[0]
        apiFunction = args[1]
        fileName = args[2]
        pathToDownload = ''
    elif len(args) == 4:
        token = args[0]
        apiFunction = args[1]
        fileName = args[2]
        pathToDownload = args[3]
    else:
        print('Sao aceitos no máximo quatro argumentos: token, apiFunction e'
              ' parametros')
        return ''

    # Specify URL | Especificar URL
    url = api_url_base + apiFunction

    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }

    # Call REST API | Chamar REST API
    response = requests.get(url, headers=headers, stream=True,
                            verify=verifyCertificate)

    print(response.status_code)
    #print(response.text)

    if (response.status_code == 401):
        token = autenticar_prospec()

        headers = {
            'Authorization': 'Bearer ' + token,
            "Content-Type": "application/json"
        }

        response = requests.get(url, headers=headers, stream=True,
                                verify=verifyCertificate)

        print(response.status_code)
        #print(response.text)

    if (response.status_code == 200):
        with open((pathToDownload + '\\' + fileName), 'wb') as file:
            for chunk in response.iter_content(chunk_size=1024):
                if chunk:  # filter out keep-alive new chunks
                    file.write(chunk)

    return ''

###########################################################
# Download dos arquivos de entrada de um determinado estudo
#

def download_decks_iniciais(token, StudyId, pathdownload):
    
    api_function = '/api/prospectiveStudies/' + str(StudyId) + '/DeckDownload'
    url = api_url_base + api_function
    nome_deck_entrada_zip = str(StudyId) + '_decks_entrada.zip'

    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }
    caminho_arquivo = os.path.join(pathdownload, nome_deck_entrada_zip)


    # verifica se o caminho existe, senão ele cria as pastas
    if not os.path.exists(pathdownload):
        os.makedirs(pathdownload)
        print(str(StudyId) + ' - Pasta criada')
    else:
        print(str(StudyId) + ' - Pasta já existe')
        validador = 1

    response_api = requests.get(url, headers=headers, stream=True, verify=verifyCertificate)
    print(response_api.status_code)

    # salva o arquivo zipado na pasta determinada
    with open(caminho_arquivo, 'wb') as file:
        for chunk in response_api.iter_content(chunk_size=1024):
            if chunk:  # filter out keep-alive new chunks
                file.write(chunk)
        file.close()
    
    #descompacta o arquivo baixado e traz a pasta do deck para a raiz
    with ZipFile(caminho_arquivo, 'r') as zipObj:
        #Extract all the contents of zip file in current directory and with the same name
        #caminho da pasta é  mesmo do zip, só que sem o ".zip"
        caminho_pasta_deck_inicial = caminho_arquivo.replace('.zip','') 
        zipObj.extractall(caminho_pasta_deck_inicial)
        print('Arquivo descompactado:', caminho_arquivo)
    
    return { 
        'caminho deck pasta temporária': caminho_pasta_deck_inicial,
        'Nome pasta deck': nome_deck_entrada_zip.replace('.zip', '')
    }

def copia_arquivos_rodada_dessem(data_rodada_datetime, caminho_rodada_dessem):
    caminho_raiz_deck_dessem = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\16.DECKS\03.DESSEM'
    caminho_raiz_dadvaz = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\10.VAZÕES DIÁRIAS\04.DADVAZ'
    caminho_raiz_carga_dessem = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\01.CARGA\06.Carga Dessem'
    caminho_rodada_dessem = os.path.realpath(caminho_rodada_dessem)
    #busca o arquivo na pasta correspondente
    infos_sem_operativa = calendario_semanas_operativas(data_rodada_datetime.format('DD-MM-YYYY'))
    mes_ano_dessem = infos_sem_operativa['mes-operativo'].format('MM') + str(infos_sem_operativa['mes-operativo'].year)
    revisao = infos_sem_operativa['revisao']
    dia_dessem = infos_sem_operativa['data-referencia'].format('DD')
    mes_dessem = infos_sem_operativa['data-referencia'].format('MM')
    ano_dessem = infos_sem_operativa['data-referencia'].format('YYYY')
    dia_seguinte_dessem = infos_sem_operativa['data-referencia'].add(days=1).format('DD')
    mes_dia_seguinte_dessem = infos_sem_operativa['data-referencia'].add(days=1).format('MM')
    ano_dia_seguinte_dessem = infos_sem_operativa['data-referencia'].add(days=1).format('YYYY')
    
    caminho_dia = ano_dessem + '\\' + mes_dessem
    caminho_dia2= ano_dia_seguinte_dessem + '\\' + mes_dia_seguinte_dessem

    #copia deck dessem
    nome_arquivo_dessem = 'DS_CCEE_' + mes_ano_dessem + '_SEMREDE_' + revisao + 'D' + dia_dessem + '.zip'
    nome_arquivo_dessem_copia = 'DS_CCEE_' + mes_ano_dessem + '_SEMREDE_' + revisao + 'D' + dia_dessem +  '_original.zip'
    caminho_rede_deck_dessem = caminho_raiz_deck_dessem + '\\' + caminho_dia + '\\' + nome_arquivo_dessem
    shutil.copy2(caminho_rede_deck_dessem, caminho_rodada_dessem + '\\' + '01.Decks')
    shutil.copy2(caminho_rede_deck_dessem, caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + nome_arquivo_dessem_copia)

    #copia dadvaz para a pasta da rodada já renomeando o arquivo para dadvaz.dat
    nome_arquivo_dadvaz_final = infos_sem_operativa['data-referencia'].format('DD_MM_YYYY') + '.DAT'
    nome_arquivo_dadvaz_final_dia_seguinte = infos_sem_operativa['data-referencia'].add(days=1).format('DD_MM_YYYY') + '.DAT'
    caminho_rede_dadvaz = caminho_raiz_dadvaz + '\\' + ano_dessem + '\\' + mes_dessem
    caminho_rede_dadvaz_dia_seguinte = caminho_raiz_dadvaz + '\\' + ano_dia_seguinte_dessem + '\\' + mes_dia_seguinte_dessem
    for file in os.listdir(caminho_rede_dadvaz_dia_seguinte):
        if file.endswith(nome_arquivo_dadvaz_final_dia_seguinte):
            nome_arquivo_dadvaz = file
            shutil.copy2(caminho_rede_dadvaz_dia_seguinte + '\\' + nome_arquivo_dadvaz, caminho_rodada_dessem + '\\' + '02.Prevs' + '\\' + nome_arquivo_dadvaz)
            shutil.copy2(caminho_rede_dadvaz_dia_seguinte + '\\' + nome_arquivo_dadvaz, caminho_rodada_dessem + '\\' + '02.Prevs' + '\\' + 'dadvaz.dat')
            print('Arquivo movido para a pasta da rodada:', caminho_rede_dadvaz_dia_seguinte + '\\' + nome_arquivo_dadvaz)
    #copia carga dessem
    #alterei para pegar o dia seguinte da carga (considerando que vamos baixar 13h, após previsao da carga)
    #o arquivo carga dessem é salvo na data do arquivo
    #pasta_carga_rede = 'Blocos_' + infos_sem_operativa['data-referencia'].format('YYYY-MM-DD')
    pasta_carga_rede = 'Blocos_' + data_rodada_datetime.add(days=1).format('YYYY-MM-DD')
    caminho_pasta_carga_rede = caminho_raiz_carga_dessem + '\\' + caminho_dia2 + '\\' + pasta_carga_rede
    for file in os.listdir(caminho_pasta_carga_rede):
        if file == 'DE.txt' or file == 'DP.txt':
            shutil.copy2(caminho_pasta_carga_rede + '\\' + file, caminho_rodada_dessem + '\\' + '02.Prevs' + '\\' + file)
            print('Arquivo movido para a pasta da rodada:', caminho_pasta_carga_rede + '\\' + file)
            print('nome:', pasta_carga_rede)

def altera_operut_deck_ccee(data_rodada_datetime, caminho_rodada_dessem):
    caminho_rodada_dessem = os.path.realpath(caminho_rodada_dessem)
    #busca o arquivo na pasta correspondente
    infos_sem_operativa = calendario_semanas_operativas(data_rodada_datetime.format('DD-MM-YYYY'))
    mes_ano_dessem = infos_sem_operativa['mes-operativo'].format('MM') + str(infos_sem_operativa['mes-operativo'].year)
    revisao = infos_sem_operativa['revisao']
    dia_dessem = infos_sem_operativa['data-referencia'].format('DD')
    mes_dessem = infos_sem_operativa['data-referencia'].format('MM')
    ano_dessem = infos_sem_operativa['data-referencia'].format('YYYY')
    dia_seguinte_dessem = infos_sem_operativa['data-referencia'].add(days=1).format('DD')
    mes_dia_seguinte_dessem = infos_sem_operativa['data-referencia'].add(days=1).format('MM')
    ano_dia_seguinte_dessem = infos_sem_operativa['data-referencia'].add(days=1).format('YYYY')
    
    caminho_dia = ano_dessem + '\\' + mes_dessem
    caminho_dia2= ano_dia_seguinte_dessem + '\\' + mes_dia_seguinte_dessem

    #copia deck dessem
    nome_arquivo_dessem = 'DS_CCEE_' + mes_ano_dessem + '_SEMREDE_' + revisao + 'D' + dia_dessem + '.zip'
    nome_arquivo_dessem_semZIP = 'DS_CCEE_' + mes_ano_dessem + '_SEMREDE_' + revisao + 'D' + dia_dessem
    #tira o zip
    shutil.unpack_archive(caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + nome_arquivo_dessem, caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + nome_arquivo_dessem_semZIP,'zip')
    #deleta o arquivo zip 
    os.remove(caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + nome_arquivo_dessem)
    #extrai o operut
    #shutil.copy2(caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + nome_arquivo_dessem_semZIP +'\\'+'operut.dat', caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + 'operut.dat')
    shutil.copy2(caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + nome_arquivo_dessem_semZIP +'\\'+'operut.dat', caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + 'operut_original.dat')    #edita o operut
    os.remove(caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + nome_arquivo_dessem_semZIP +'\\'+'operut.dat')
    #edita o operut
    edita_operut(caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + nome_arquivo_dessem_semZIP)
    #move arquivo editado
    shutil.copy2(caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + 'operut.dat',caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + nome_arquivo_dessem_semZIP +'\\'+'operut.dat')    #edita o operut
    #zipa deck novamente
    shutil.make_archive(caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + nome_arquivo_dessem_semZIP,'zip',caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + nome_arquivo_dessem_semZIP)
    #deleta a pasta nao zipada
    shutil.rmtree(caminho_rodada_dessem + '\\' + '01.Decks'+ '\\' + nome_arquivo_dessem_semZIP)

def verifica_arquivos_dessem(data_dessem_datetime):
    #essa função verifica se os aruqivos necessários para criação do estudo do dessem constam na pasta da rodada do dia
    #deck do dessem
    #carga prevista do dessem
    #dadvaz
    
    
    caminho_raiz_rodada_dessem = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\13.RODADAS\07.Dessem\02.Rodadas'
    caminho_dia_dessem = data_dessem_datetime.format(r'YYYY\\MM\\DD')
    caminho_rodada_dessem = caminho_raiz_rodada_dessem + '\\' + caminho_dia_dessem
    #cria nome do arquivo do deck do dessem para verificar sua existência
    infos_sem_operativa = calendario_semanas_operativas(data_dessem_datetime.format('DD-MM-YYYY'))
    mes_ano_dessem = infos_sem_operativa['mes-operativo'].format('MM') + str(infos_sem_operativa['mes-operativo'].year)
    revisao = infos_sem_operativa['revisao']
    dia_dessem = infos_sem_operativa['data-referencia'].format('DD')
    nome_arquivo_deck_dessem = 'DS_CCEE_' + mes_ano_dessem + '_SEMREDE_' + revisao + 'D' + dia_dessem + '.zip'
    arquivos_nao_encontrados = 0
    if os.path.exists(caminho_rodada_dessem + '\\01.Decks\\' + nome_arquivo_deck_dessem):
        print('Deck do dessem encontrado:', nome_arquivo_deck_dessem)
    else:
        print('Deck do dessem não encontrado na pasta padrão:', nome_arquivo_deck_dessem)
        arquivos_nao_encontrados += 1
    
    if os.path.exists(caminho_rodada_dessem + '\\02.Prevs\\' + 'dadvaz.dat'):
        print('Arquivo dadvaz.dat encontrado:', caminho_rodada_dessem + '\\02.Prevs\\' + 'dadvaz.dat')
    else:
        print('Arquivo dadvaz.dat não encontrado:', caminho_rodada_dessem + '\\02.Prevs\\' + 'dadvaz.dat')
        arquivos_nao_encontrados += 1
    
    if os.path.exists(caminho_rodada_dessem + '\\02.Prevs\\' + 'DE.txt') and os.path.exists(caminho_rodada_dessem + '\\02.Prevs\\' + 'DP.txt'):
        print('Arquivos de carga encontrados (DE.txt e DP.txt) no caminho:', caminho_rodada_dessem + '\\02.Prevs')
    else:
        print('Um ou mais arquivos de carga não encontrados (DE.txt e DP.txt) no caminho:', caminho_rodada_dessem + '\\02.Prevs')
        arquivos_nao_encontrados += 1
    if arquivos_nao_encontrados == 0:
        arquivos_corretos = True
    else:
        arquivos_corretos = False

    return {'Arquivos Corretos':arquivos_corretos,
        'QTD Arquivos pendentes': arquivos_nao_encontrados,
        'Caminho rodada dessem': caminho_rodada_dessem,
        'Nome do Deck Dessem': nome_arquivo_deck_dessem,

        
        
        
        }


def edita_entdados_deleta_blocos(data_dessem_datetime, caminho_decks_iniciais):
    # a data_dessem_datetime do input é a data do segundo dia do dessem
    '''
    A partir do caminho da pasta dos decks de entrada e data da rodada do dessem,
    o arquivo entdados é copiado para fora da pasta de decks de entrada e subpasta do dessem da data da rodada,
    e colado na mesma pasta onde está a pasta do deck de entrada (dois níveis acima da subpasta do dessem).
    é criado um novo arquivo entdados.dat nessa mesma pasta raiz já com os blocos deletados
    bloco DP é deletado por inteiro
    bloco DE é deletado até onde não há dados preenchidos na coluna Justific (última coluna do bloco). Esse critério é baseado no tamanho da linha

    '''
    #determina qual a o nome da pasta do deck de entrada do segundo dia do dessem 
    subpasta_dessem = 'DS' + data_dessem_datetime.format('YYYYMMDD')
    #determina a pasta do primeiro dia
    subpasta_dessem_inicial='DS' + data_dessem_datetime.add(days=-1).format('YYYYMMDD')

    nome_novo_entdados = 'entdados.dat'
    nome_arquivo_operut = 'operut.dat'
    #nome do arquivo original que será copiado para fora da pasta, como evidência do processo
    nome_arquivo_entdados = nome_novo_entdados.replace('.dat', '_original_' + subpasta_dessem + '.dat')
    caminho_raiz = caminho_decks_iniciais[:caminho_decks_iniciais.rfind('\\')]

    #copia o entdados do primeiro dia para a pasta da rodada
    nome_arquivo_entdados = 'entdados_original_DS20210803.dat'
    caminho_arquivo = caminho_raiz + '\\' + nome_arquivo_entdados
    caminho_entdados_editado = caminho_raiz + '\\' + nome_novo_entdados
    shutil.copy2(caminho_decks_iniciais + '\\' + subpasta_dessem + '\\' + 'entdados.dat', caminho_arquivo)
    #copia o operut do arquivo do primeiro dia para pasta da rodada
    #shutil.copy2(caminho_decks_iniciais + '\\' + subpasta_dessem_inicial + '\\' + 'operut.dat', caminho_raiz + '\\' + 'operut_original.dat')
    
    with open(caminho_arquivo, 'r') as entdados:

        entdados_original = entdados.readlines()
        entdados.close()

    pri_linha_DP = 1
    pri_linha_DE = 1
    with open(caminho_entdados_editado, 'w') as novo_entdados:
        for line in entdados_original:
            if line.startswith('DP '):
                if pri_linha_DP == 1:
                    novo_entdados.write('&DP - INICIO BLOCO DE CARGA\n')
                    #print('primeira linha')
                    #print(line)
                    pri_linha_DP += 1
                    
                    continue

                else:
                    pri_linha_DP += 1
            
            elif line.startswith('DE '):
                if len(line) < 46:
                    if pri_linha_DE == 1:
                        novo_entdados.write('&DE - INICIO BLOCO DE DEMANDAS/CARGAS ESPECIAIS\n')
                        pri_linha_DE += 1
                        continue
                    else:
                        pri_linha_DE += 1
                else:
                    novo_entdados.write(line)
            else:
                novo_entdados.write(line)
    print('Linhas deletadas do bloco DP:', pri_linha_DP - 1)
    print('Linhas deletadas do bloco DE:', pri_linha_DE - 1)
    return caminho_entdados_editado

def edita_operut(caminho_decks_iniciais):
    # a data_dessem_datetime do input é a data do segundo dia do dessem
    '''
    A partir do caminho da pasta dos decks de entrada e data da rodada do dessem,
    o arquivo entdados é copiado para fora da pasta de decks de entrada e subpasta do dessem da data da rodada,
    e colado na mesma pasta onde está a pasta do deck de entrada (dois níveis acima da subpasta do dessem).
    é criado um novo arquivo entdados.dat nessa mesma pasta raiz já com os blocos deletados
    bloco DP é deletado por inteiro
    bloco DE é deletado até onde não há dados preenchidos na coluna Justific (última coluna do bloco). Esse critério é baseado no tamanho da linha

    '''

    caminho_raiz = caminho_decks_iniciais[:caminho_decks_iniciais.rfind('\\')]

    nome_novo_operut = 'operut.dat'
    nome_arquivo_operut ='operut_original.dat'     
    caminho_arquivo = caminho_raiz + '\\' + nome_arquivo_operut
    caminho_operut_editado = caminho_raiz + '\\' + nome_novo_operut

    with open(caminho_arquivo, 'r') as operutoriginal:
        operut_original = operutoriginal.readlines()
        operutoriginal.close()

    with open(caminho_operut_editado, 'w') as novo_operut:
        for line in operut_original:
            if line.startswith('&UCTERM 2'):
                    novo_operut.write('UCTERM 2\n')
            else:
                novo_operut.write(line)

    print('operut alterado')
    return caminho_operut_editado

def le_arquivos_prev_carga_dessem(caminho_raiz):
    '''
    Está função lê os arquivo DP.txt e DE.txt que estão em uma pasta dada como argumento e retorna as duas listas
    com dados desses arquivos
    '''
    caminho_raiz = os.path.realpath(caminho_raiz)
    caminho_arquivo_DP = caminho_raiz + '\\' + 'DP.txt'
    caminho_arquivo_DE = caminho_raiz + '\\' + 'DE.txt'

    #Lê arquivo DP original
    with open(caminho_arquivo_DP, 'r') as arquivo_DP:
        arquivo_DP_original = arquivo_DP.readlines()
        #print(arquivo_DP_original)
    novo_bloco_DP = []
    for item in arquivo_DP_original:
        if not(item.startswith('&')):
            novo_bloco_DP.append(item)
    print('Linhas do novo bloco DP:', len(novo_bloco_DP))
    print('Leitura do novo bloco DP concluída')


    #lê arquivo DE original
    with open(caminho_arquivo_DE, 'r') as arquivo_DE:
        arquivo_DE_original = arquivo_DE.readlines()
        #print(arquivo_DP_original)
    novo_bloco_DE = []
    for item in arquivo_DE_original:
        if not(item.startswith('&')):
            novo_bloco_DE.append(item)
    print('Linhas do novo bloco DE:', len(novo_bloco_DE))
    print('Leitura do novo bloco DE concluída')

    dados_carga = {
        'bloco_DP': novo_bloco_DP,
        'bloco_DE': novo_bloco_DE
    }
    return dados_carga

def busca_deck_dessem(IdStudy, data_deck_datetime):
    #a funcao nao funciona adequadamente, sempre pega o id do último dia
    #o mes que o getListOfDecks retorna é o mês operativo, por isso a funcao não tem como ficar certinha...estranho

    '''
    Esta função busca o código de um deck DESSEM a partir do código do estudo e da data do deck em formato datetime
    Esta função retorna o Modelo do deck (strimg), a data do deck (string) e Deck ID (inteiro)
    '''
    modelo_deck = 'DESSEM'
    ano_deck = data_deck_datetime.year
    mes_deck = data_deck_datetime.month
    dia_deck = data_deck_datetime.day
   
    lista_decks = getListOfDecks(IdStudy)
    #numid=None
    for deck in lista_decks:
        if deck['Model'] == modelo_deck and deck['Year'] == ano_deck and deck['Month'] == mes_deck and deck['Day'] == dia_deck:
            #numid=deck['Id']
            print('ID do deck do Dessem de', data_deck_datetime.format('DD/MM/YYYY'), ':', deck['Id'])
            #break
    dados_deck = {
        'Modelo deck':modelo_deck,
        'Data Deck': data_deck_datetime.format('DD/MM/YYYY'),
        'Deck ID': deck['Id']
        #'Deck ID': numid
    }
    
    return dados_deck

def executa_dessem (idStudy, serverType, ExecutionMode, InfeasibilityHandling, InfeasibilityHandlingSensibility, maxRestarts):
    api_function = '/api/prospectiveStudies/' + str(idStudy) + '/Run'
    url = api_url_base + api_function
    idQueue = 0
    data = {
        "SpotInstanceType": serverType,
        "ExecutionMode": ExecutionMode,
        "InfeasibilityHandling": InfeasibilityHandling,
        "InfeasibilityHandlingSensibility": InfeasibilityHandlingSensibility,
        "MaxTreatmentRestarts": maxRestarts
    }
    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }

    parameter = ''

    response = requests.post(url, headers=headers, params=parameter,
                             data=json.dumps(data), verify=verifyCertificate)
    print("A seguinte configuracao sera usada para iniciar a execucao o estudo: ", str(idStudy))
    print(data)
    print(response)

def arquiva_dados_rodada_dessem(StudyId, data, caminho_estudo, **kwargs):
    caminho_estudo = os.path.realpath(caminho_estudo)
    caminho_base_prospec_dessem = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_dessem.xlsx'

    print('Iniciando a importação de dados da pasta:', caminho_estudo)
    arquivo_excel_rodada = load_workbook(caminho_base_prospec_dessem)
    aba_cadastro = arquivo_excel_rodada['Cadastro_estudos']
    #arquivos_csv = ['compila_cmo_medio', 'compila_cmo_patamar']
    arquivos_csv = ['compila_cmo_medio', 'compila_cmo_patamar', 'compila_pld_horario', 'compila_pld', 'compila_ea', 'compila_ea_mw', 'compila_gh', 'compila_gt']    


    data_criacao = data
    linha = aba_cadastro.max_row + 1
    print('Número da primeira linha vazia na aba Cadastro:', linha)
    aba_cadastro.cell(row=linha, column=2).value = data_criacao
    aba_cadastro.cell(row=linha, column=1).value = aba_cadastro.cell(row=linha - 1, column=1).value + 1
    aba_cadastro.cell(row=linha, column=2).number_format = 'dd/mm/yyyy'
    aba_cadastro.cell(row=linha, column=3).value = StudyId
    aba_cadastro.cell(row=linha, column=6).value = 'DESSEM'
    for nome_csv in arquivos_csv:
        nome_arquivo = nome_csv + '.csv'
        nome_aba = arquivo_excel_rodada[nome_csv]

        with open(caminho_estudo + '\\' + nome_arquivo) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=';')
            line_count = 0
            linha = nome_aba.max_row + 1
            for row in csv_reader:
                if line_count == 0:
                    print(f'Column names are {", ".join(row)}')
                    line_count += 1
                else:
                    for coluna in range(len(row)):
                        nome_aba.cell(row=linha, column=coluna + 2).value = row[coluna]
                    nome_aba.cell(row=linha, column=1).value = StudyId
                    linha = linha + 1
                    line_count += 1
            print(f'Processed {line_count} lines.')

    #arquivo_excel_rodada.save(r'C:\Users\Middle\Desktop\arquivos_download\testando.xlsx')
    #arquivo_excel_rodada.save(r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_dessem.xlsx')
    arquivo_excel_rodada.save(caminho_base_prospec_dessem)
    arquivo_excel_rodada.close()


def deleta_linhas_base_prospec_dessem(StudyId):
    #caminho_estudo = os.path.realpath(caminho_estudo)
    caminho_base_prospec_dessem = r'J:\SEDE\Comercializadora de Energia\6. MIDDLE\12.ANALISES\Prospec\Base_prospec_dessem.xlsx'
    #print('Iniciando a importação de dados da pasta:', caminho_estudo)
    arquivo_excel_rodada = load_workbook(caminho_base_prospec_dessem)
    #aba_cadastro = arquivo_excel_rodada['Cadastro_estudos']
    #arquivos_csv = ['compila_cmo_medio', 'compila_cmo_patamar']
    arquivos_csv = ['compila_cmo_medio', 'compila_cmo_patamar', 'compila_pld_horario', 'compila_pld', 'compila_ea', 'compila_ea_mw', 'compila_gh', 'compila_gt']
    #Verifica se já existem dados desse estudo na planilha
    #imputa informações da aba de cadastro do estudo
    arquivos_csv_temp = deepcopy(arquivos_csv)
    arquivos_csv_temp.append('Cadastro_estudos')
    salvar_base = False

    for aba in arquivos_csv_temp:
        if aba == 'Cadastro_estudos':
            coluna = 3
        else: coluna = 1
        nome_aba = arquivo_excel_rodada[aba]
        #ult_lin = nome_aba.max_row
        linha = 2
        qtd_lin_del = 0
        while nome_aba.cell(linha, coluna).value is not None:
            if int(nome_aba.cell(linha, coluna).value) == int(StudyId):
                nome_aba.delete_rows(linha, amount=1)
                #print(linha deletada)
                qtd_lin_del += 1
                salvar_base = True
            else:
                linha += 1
        print('Aba:', aba,'quantidade de linhas deletadas:', qtd_lin_del)
        if nome_aba.cell(linha, coluna).value is None:
            nome_aba.delete_rows(linha, amount=1)
            print('Linha adicional deletada')
    if salvar_base:
        arquivo_excel_rodada.save(caminho_base_prospec_dessem)
        print('Base de dados Salva')
    arquivo_excel_rodada.close()

#Necessaria para o runExecution
def getInfoFromStudy(idStudy):
    prospecStudy = getInfoFromAPI(token, '/api/prospectiveStudies/'
                                  + str(idStudy))
    return prospecStudy

#Necessaria para o runExecution
def getToken(username_temp, password_temp):
    basicURL = 'https://api.prospec.app'
    url = basicURL + '/api/Token'

    global username
    username = username_temp
    global password
    password = password_temp

    headers = {
        'content-type': 'application/x-www-form-urlencoded'
    }

    data = {
        'grant_type': 'password',
        'username': username,
        'password': password
    }

    tokenResponse = requests.post(url, headers=headers, data=data,
                                  verify=verifyCertificate)
    token_json = tokenResponse.json()
    token = token_json["access_token"]

    return token

#Necessaria para o runExecution
def getInfoFromAPI(*args):
    basicURL = 'https://api.prospec.app'
    verifyCertificate = True
    username = ''
    password = ''

    if len(args) < 2:
        print('Sao necessarios ao menos dois argumentos: token e apiFunction')
        return ''
    elif len(args) == 2:
        token = args[0]
        apiFunction = args[1]
    elif len(args) == 3:
        token = args[0]
        apiFunction = args[1]
        params = args[2]
    else:
        print('Sao necessarios ao menos dois argumentos: token, apiFunction e'
              ' parametros')
        return ''

    # Specify URL | Especificar URL
    url = basicURL + apiFunction

    headers = {
        'Authorization': 'Bearer ' + token,
        "Content-Type": "application/json"
    }

    # Call REST API | Chamar Rest API
    if 'params' in locals():
        response = requests.get(url, headers=headers, params=params,
                                verify=verifyCertificate)
    else:
        response = requests.get(url, headers=headers, verify=verifyCertificate)

    print(response.status_code)

    if (response.status_code == 401):
        token = getToken(username, password)

        headers = {
            'Authorization': 'Bearer ' + token,
            "Content-Type": "application/json"
        }

        if 'params' in locals():
            response = requests.get(url, headers=headers, params=params,
                                    verify=verifyCertificate)
        else:
            response = requests.get(url, headers=headers,
                                    verify=verifyCertificate)

        print(response.status_code)

    if (response.status_code == 200):
        return response.json()
    return ''

#Executar Dessem
def runExecution(idStudy, idServer, idQueue, idNEWAVEJson, idDECOMPJson, idDESSEMJson, spotInstanceType, executionMode,
                 infeasibilityHandling, maxRestarts):
    parameter = ''

    if idServer == 0:
        if spotInstanceType == '':
            prospecStudy = getInfoFromStudy(idStudy)
            listOfDecks = prospecStudy['Decks']
            containsNEWAVE = False
            for deck in listOfDecks:
                if deck['Model'] == 'NEWAVE':
                    containsNEWAVE = True
                    break

            if containsNEWAVE:
                data = {
                    "SpotInstanceType": 'c5.9xlarge',
                    "ExecutionMode": executionMode,
                    "InfeasibilityHandling": infeasibilityHandling,
                    "MaxTreatmentRestarts": maxRestarts
                }
            else:
                data = {
                    "SpotInstanceType": 'm5.4xlarge',
                    "ExecutionMode": executionMode,
                    "InfeasibilityHandling": infeasibilityHandling,
                    "MaxTreatmentRestarts": maxRestarts
                }

        else:
            data = {
                "SpotInstanceType": spotInstanceType,
                "ExecutionMode": executionMode,
                "InfeasibilityHandling": infeasibilityHandling,
                "MaxTreatmentRestarts": maxRestarts
            }
    elif idQueue == 0:
        data = {
            "ServerId": int(idServer),
            "ExecutionMode": executionMode,
            "InfeasibilityHandling": infeasibilityHandling,
            "MaxTreatmentRestarts": maxRestarts
        }
    else:
        data = {
            "ServerId": int(idServer),
            "QueueId": int(idQueue),
            "ExecutionMode": executionMode,
            "InfeasibilityHandling": infeasibilityHandling,
            "MaxTreatmentRestarts": maxRestarts,
        }

    # deckModel = []

    # for deck in listOfDecks:
    #      if deck['Model'] == 'NEWAVE' and idNEWAVEJson["idNewaveJulho"] != 0 and deck["Month"] == 7 and deck["Year"] == 2021:
    #        deckModel.append({"DeckId": deck["Id"], "NewaveVersionId": idNEWAVEJson["idNewaveJulho"], "MaxTreatmentRestarts": 10})

    #      elif deck['Model'] == 'DECOMP' and idDECOMPJson["idDecomp"] != 0:
    #        deckModel.append({"DeckId": deck["Id"], "DecompVersionId": idDECOMPJson["idDecomp"] , "MaxTreatmentRestarts": 10})

    #      elif deck['Model'] == 'DESSEM' and idDESSEMJson["idDessem"] != 0:
    #        deckModel.append({"DeckId": deck["Id"], "DessemVersionId": idDESSEMJson["idDessem"], "MaxTreatmentRestarts": 10})

    # if (deckModel.count != 0):
    #     data["DecksRunModel"] = deckModel

    print("A seguinte configuracao sera usada para iniciar a execucao o estudo: ", str(idStudy))
    print(data)

    response = postInAPI(token, '/api/prospectiveStudies/' + str(idStudy)
                         + '/Run', parameter, data)
